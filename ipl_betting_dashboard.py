# ============================================================
# IPL In-Play Betting Hedge Calculator v4
# Multi-User · Portfolio Intelligence · Gemini AI
# ============================================================
# requirements.txt:
#   streamlit>=1.30.0
#   plotly>=5.18.0
#   scipy>=1.11.0
#   numpy>=1.24.0
#   google-genai>=1.0.0
#   openpyxl>=3.1.0
#
# Run:  streamlit run ipl_betting_dashboard.py
# ============================================================

import streamlit as st
import plotly.graph_objects as go
import numpy as np
from scipy.optimize import linprog
from datetime import datetime
import json, copy, os, uuid, time, io

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from google import genai as genai_client_module
    from google.genai import types as genai_types
    GENAI_AVAILABLE = True
except ImportError:
    try:
        import google.generativeai as genai_legacy
        GENAI_AVAILABLE = True
        genai_client_module = None
        genai_types = None
    except ImportError:
        GENAI_AVAILABLE = False
        genai_client_module = None
        genai_legacy = None
        genai_types = None

# ── Constants ────────────────────────────────────────────────

BG = "#0A0F1E"
CARD = "#0F1629"
CARD_BORDER = "#1A2340"
GREEN = "#00FF88"
AMBER = "#FFB347"
RED = "#FF4B4B"
TEXT = "#E2E8F0"
MUTED = "#64748B"
CYAN = "#22D3EE"
VIOLET = "#A78BFA"
MISC_PURPLE = "#7B6CF6"

USERS_INDEX_FILE = "users_index.json"
GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_MAX_RETRIES = 3
GEMINI_RETRY_DELAY = 5  # seconds between retries on 429
GEMINI_SYSTEM = (
    "You are a sharp IPL cricket betting strategist. You think in probabilities and expected value. "
    "You give specific numbers, not theory. You say DO or DON'T, never 'consider'. "
    "You reference the user's actual bets and capital. You are concise — 3 sentences max per point. "
    "When asked about predictions, give exact ranges with most likely numbers. "
    "When assessing bets, say if each one is LOOKING GOOD, BAD, or NEUTRAL with one line why. "
    "You never give generic cricket commentary. Every word must help the user make a decision."
)

OUTCOME_LABELS = {"t1": "Team 1 Win", "t2": "Team 2 Win", "tie": "Tie"}
OUTCOME_COLORS = {"t1": GREEN, "t2": CYAN, "tie": AMBER}


# ── Utility Functions ────────────────────────────────────────

def fmt_inr(amount, decimals=0):
    neg = amount < 0
    val = abs(amount)
    if decimals > 0:
        integer_part = int(val)
        frac = f".{int(round((val - integer_part) * (10 ** decimals))):0{decimals}d}"
    else:
        integer_part = int(round(val))
        frac = ""
    s = str(integer_part)
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            chunk = s[-2:] if len(s) >= 2 else s
            result = chunk + "," + result
            s = s[:-2]
    sign = "\u2212" if neg else ""
    return f"{sign}\u20b9{result}{frac}"


def fmt_pct(val, signed=True):
    sign = "+" if signed and val > 0 else ""
    return f"{sign}{val:.1f}%"


def get_all_bets():
    return st.session_state.pre_bets + st.session_state.bets


def get_all_bets_including_misc():
    active_misc = [b for b in st.session_state.get("misc_bets", []) if b.get("status", "active") == "active"]
    return st.session_state.pre_bets + st.session_state.bets + active_misc


def get_match_total_pnl(m):
    """Total P&L for a settled match = main bets + misc bets."""
    return m.get("realized_pnl", 0) + m.get("misc_realized_pnl", 0)


def compute_pnl(bets):
    pnl = {"t1": 0.0, "t2": 0.0, "tie": 0.0}
    for b in bets:
        outcome = b.get("outcome", "")
        if outcome not in pnl:
            continue
        stake = b["stake"]
        profit = stake * (b["odds"] - 1.0)
        for sc in pnl:
            if sc == outcome:
                pnl[sc] += profit
            else:
                pnl[sc] -= stake
    return pnl


def compute_misc_pnl():
    misc = st.session_state.get("misc_bets", [])
    if not misc:
        return {"best": 0.0, "worst": 0.0, "count": 0, "realized": 0.0, "active_count": 0}
    active = [b for b in misc if b.get("status", "active") == "active"]
    settled = [b for b in misc if b.get("status") == "settled"]
    realized = sum(b.get("realized_pnl", 0) for b in settled)
    return {
        "best": realized + sum(b["stake"] * (b["odds"] - 1.0) for b in active),
        "worst": realized - sum(b["stake"] for b in active),
        "count": len(misc),
        "realized": realized,
        "active_count": len(active),
    }


def compute_total_staked(bets):
    return sum(b["stake"] for b in bets)


def compute_edge_shift(entry_odds, current_odds, outcome):
    if entry_odds <= 1 or current_odds.get(outcome, 0) <= 1:
        return None
    curr = current_odds[outcome]
    impl_entry = 100.0 / entry_odds
    impl_now = 100.0 / curr
    edge_shift = impl_now - impl_entry
    return {
        "impl_entry": impl_entry, "impl_now": impl_now,
        "edge_shift": edge_shift, "odds_entry": entry_odds, "odds_now": curr,
        "direction": "favourable" if edge_shift > 0 else ("against" if edge_shift < 0 else "neutral"),
    }


def solve_optimal_hedge(pnl, odds, capital):
    o1, o2, ot = odds["t1"], odds["t2"], odds["tie"]
    p1, p2, pt = pnl["t1"], pnl["t2"], pnl["tie"]
    if o1 <= 1 or o2 <= 1 or ot <= 1:
        return None
    c = [0, 0, 0, -1]
    A_ub = [
        [-(o1 - 1), 1, 1, 1], [1, -(o2 - 1), 1, 1],
        [1, 1, -(ot - 1), 1], [1, 1, 1, 0],
    ]
    b_ub = [p1, p2, pt, capital]
    bounds = [(0, capital), (0, capital), (0, capital), (None, None)]
    try:
        res = linprog(c, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")
        if res.success:
            s1, s2, st_val, z = res.x
            s1 = max(0, round(s1, 2)); s2 = max(0, round(s2, 2)); st_val = max(0, round(st_val, 2))
            new_pnl = {
                "t1": p1 + s1 * (o1 - 1) - s2 - st_val,
                "t2": p2 - s1 + s2 * (o2 - 1) - st_val,
                "tie": pt - s1 - s2 + st_val * (ot - 1),
            }
            return {"stakes": {"t1": s1, "t2": s2, "tie": st_val}, "min_pnl": z,
                    "new_pnl": new_pnl, "total_stake": s1 + s2 + st_val, "success": True}
    except Exception:
        pass
    return None


# ── Conviction Mode Solver ───────────────────────────────────

def solve_conviction_hedge(pnl, odds, capital, conviction):
    """
    Conviction mode: user believes 'conviction' outcome will win.
    Find minimum hedge stakes on OTHER outcomes to guarantee break-even
    if conviction is wrong, while maximizing profit if conviction hits.
    """
    outcomes = ["t1", "t2", "tie"]
    if conviction not in outcomes:
        return None
    others = [o for o in outcomes if o != conviction]
    if not all(odds.get(k, 0) > 1 for k in outcomes):
        return None

    o1, o2 = others
    p1, p2 = pnl[o1], pnl[o2]
    od1, od2 = odds[o1], odds[o2]

    # Minimize s1+s2 (hedge cost) subject to:
    #   If o1 wins: p1 + s1*(od1-1) - s2 >= 0
    #   If o2 wins: p2 - s1 + s2*(od2-1) >= 0
    #   s1+s2 <= capital, s1>=0, s2>=0
    c_obj = [1, 1]
    A_ub = [[-(od1 - 1), 1], [1, -(od2 - 1)], [1, 1]]
    b_ub = [p1, p2, capital]
    bounds = [(0, capital), (0, capital)]

    try:
        res = linprog(c_obj, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")
        if res.success:
            s1, s2 = max(0, round(res.x[0], 2)), max(0, round(res.x[1], 2))
            total_hedge = s1 + s2
            conviction_profit = pnl[conviction] - total_hedge
            pnl_o1 = p1 + s1 * (od1 - 1) - s2
            pnl_o2 = p2 - s1 + s2 * (od2 - 1)
            return {
                "stakes": {o1: s1, o2: s2},
                "total_hedge": total_hedge,
                "conviction_profit": round(conviction_profit, 2),
                "other_pnls": {o1: round(pnl_o1, 2), o2: round(pnl_o2, 2)},
                "success": True,
                "feasible": pnl_o1 >= -1 and pnl_o2 >= -1,
            }
    except Exception:
        pass
    return None


def solve_ponr_hedge(pnl, odds, capital, dead_outcome):
    """
    Point of No Return: 'dead_outcome' cannot win anymore.
    Optimize hedge across remaining 2 outcomes only.
    """
    outcomes = ["t1", "t2", "tie"]
    if dead_outcome not in outcomes:
        return None
    alive = [o for o in outcomes if o != dead_outcome]
    if not all(odds.get(k, 0) > 1 for k in alive):
        return None

    o1, o2 = alive
    p1, p2 = pnl[o1], pnl[o2]
    od1, od2 = odds[o1], odds[o2]

    # maximize z where:
    #   p1 + s1*(od1-1) - s2 >= z
    #   p2 - s1 + s2*(od2-1) >= z
    #   s1+s2 <= capital
    c_obj = [0, 0, -1]
    A_ub = [[-(od1 - 1), 1, 1], [1, -(od2 - 1), 1], [1, 1, 0]]
    b_ub = [p1, p2, capital]
    bounds = [(0, capital), (0, capital), (None, None)]

    try:
        res = linprog(c_obj, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")
        if res.success:
            s1, s2, z = res.x
            s1 = max(0, round(s1, 2)); s2 = max(0, round(s2, 2))
            return {
                "stakes": {o1: s1, o2: s2},
                "total_stake": s1 + s2,
                "min_pnl": z,
                "new_pnl": {o1: p1 + s1 * (od1 - 1) - s2, o2: p2 - s1 + s2 * (od2 - 1)},
                "dead_outcome": dead_outcome,
                "success": True,
            }
    except Exception:
        pass
    return None


# ── Advanced Algorithm Layer ─────────────────────────────────

def compute_bet_ev(entry_odds, fair_prob_pct):
    """
    Expected Value per ₹100 staked.
    fair_prob_pct: overround-adjusted true probability (0-100).
    entry_odds: the odds at which the bet was placed.
    Returns EV as a percentage of stake.
    """
    p = fair_prob_pct / 100.0
    profit = entry_odds - 1.0
    ev = p * profit - (1 - p) * 1.0  # per unit staked
    return ev * 100  # as percentage


def compute_all_bets_ev(bets, odds):
    """Compute EV for each bet using overround-adjusted fair probabilities."""
    results = []
    # Compute fair probabilities (remove overround)
    if not all(odds.get(k, 0) > 1 for k in ["t1", "t2", "tie"]):
        return [{"bet_id": b.get("id", ""), "ev_pct": 0, "fair_prob": 0} for b in bets]
    raw_impl = {k: 100.0 / odds[k] for k in ["t1", "t2", "tie"]}
    impl_sum = sum(raw_impl.values())  # e.g. 105% with 5% overround
    fair_probs = {k: raw_impl[k] / impl_sum * 100 for k in raw_impl}  # sum to exactly 100%

    for b in bets:
        outcome = b.get("outcome", "")
        if outcome in fair_probs:
            entry_odds = b["odds"]
            ev = compute_bet_ev(entry_odds, fair_probs[outcome])
            results.append({"bet_id": b.get("id", ""), "ev_pct": ev, "fair_prob": fair_probs[outcome]})
        else:
            results.append({"bet_id": b.get("id", ""), "ev_pct": 0, "fair_prob": 0})
    return results


def compute_kelly_fraction(true_prob_pct, odds):
    """
    Kelly Criterion: optimal bet fraction.
    true_prob_pct: estimated true probability (0-100).
    odds: decimal odds offered.
    Returns fraction of bankroll to stake (0.0 to 1.0), capped at 0.25.
    """
    if odds <= 1 or true_prob_pct <= 0 or true_prob_pct >= 100:
        return 0.0
    p = true_prob_pct / 100.0
    b = odds - 1.0
    kelly = (p * b - (1 - p)) / b
    if kelly <= 0:
        return 0.0
    # Half-Kelly for safety (standard practice)
    return min(kelly * 0.5, 0.25)


def detect_arbitrage(odds):
    """
    Check if current odds create an arbitrage opportunity.
    Returns dict with is_arb, overround, and stakes for guaranteed profit.
    """
    if not all(odds.get(k, 0) > 1 for k in ["t1", "t2", "tie"]):
        return {"is_arb": False, "overround_pct": 0}

    impl_sum = sum(1.0 / odds[k] for k in ["t1", "t2", "tie"])
    overround_pct = (impl_sum - 1.0) * 100

    result = {"is_arb": impl_sum < 1.0, "overround_pct": overround_pct}
    if result["is_arb"]:
        # Calculate stakes per ₹1000 total investment for guaranteed profit
        total = 1000
        stakes = {}
        for k in ["t1", "t2", "tie"]:
            stakes[k] = round(total * (1.0 / odds[k]) / impl_sum, 2)
        guaranteed_return = min(stakes[k] * odds[k] for k in stakes)
        result["stakes_per_1000"] = stakes
        result["guaranteed_profit_per_1000"] = round(guaranteed_return - total, 2)
    return result


def compute_cashout_value(pnl, odds, capital):
    """
    What your current position is worth if you hedge right now.
    Returns the guaranteed P&L achievable = your 'cash-out' value.
    """
    sol = solve_optimal_hedge(pnl, odds, capital)
    if sol:
        return sol["min_pnl"]
    return min(pnl.values())


def compute_hedge_cost_projections(pnl, odds, capital):
    """
    Sensitivity analysis: how hedge cost changes if odds move 5%, 10%, 15%.
    Shows user whether to wait or act now.
    """
    if not all(odds[k] > 1 for k in odds):
        return []
    base_sol = solve_optimal_hedge(pnl, odds, capital)
    if not base_sol:
        return []
    base_cost = base_sol["total_stake"]
    base_min = base_sol["min_pnl"]

    projections = []
    # Find which outcome the user is most exposed to (biggest negative P&L)
    worst_sc = min(pnl, key=pnl.get)
    best_sc = max(pnl, key=pnl.get)

    for shift_pct in [-15, -10, -5, 5, 10, 15]:
        shifted_odds = dict(odds)
        # Shift the favourite (best_sc) odds — simulates match momentum
        factor = 1.0 + (shift_pct / 100.0)
        shifted_odds[best_sc] = max(1.01, odds[best_sc] * factor)
        # Counter-shift the underdog
        shifted_odds[worst_sc] = max(1.01, odds[worst_sc] / factor)

        shifted_sol = solve_optimal_hedge(pnl, shifted_odds, capital)
        if shifted_sol:
            projections.append({
                "shift_pct": shift_pct,
                "hedge_cost": shifted_sol["total_stake"],
                "min_pnl": shifted_sol["min_pnl"],
                "cost_diff": shifted_sol["total_stake"] - base_cost,
                "pnl_diff": shifted_sol["min_pnl"] - base_min,
            })

    return projections


def log_odds_snapshot():
    """Store current odds as a timestamped snapshot in match history."""
    odds = {"t1": st.session_state.odds_t1, "t2": st.session_state.odds_t2, "tie": st.session_state.odds_tie}
    if not all(odds[k] > 1 for k in odds):
        return
    history = st.session_state.get("odds_snapshots", [])
    # Avoid duplicate if odds haven't changed
    if history and history[-1]["odds"] == odds:
        return
    history.append({
        "odds": dict(odds),
        "timestamp": datetime.now().strftime("%H:%M:%S"),
        "phase": st.session_state.match_phase,
        "overround": sum(100 / odds[k] for k in odds) - 100,
    })
    st.session_state["odds_snapshots"] = history
    # After writing the new snapshot, check for volatility spike
    # (no-op if RISK GUARDS block hasn't initialized yet)
    try:
        _risk_detect_volatility_spike()
    except NameError:
        pass  # guards not yet loaded during cold start


def get_odds_momentum():
    """Compute odds momentum from snapshot history."""
    history = st.session_state.get("odds_snapshots", [])
    if len(history) < 2:
        return None
    first = history[0]["odds"]
    last = history[-1]["odds"]
    prev = history[-2]["odds"] if len(history) >= 2 else first

    momentum = {}
    for sc in ["t1", "t2", "tie"]:
        if first[sc] > 1 and last[sc] > 1:
            total_shift = (100.0 / last[sc]) - (100.0 / first[sc])
            recent_shift = (100.0 / last[sc]) - (100.0 / prev[sc])
            direction = "shortening" if last[sc] < first[sc] else ("drifting" if last[sc] > first[sc] else "stable")
            momentum[sc] = {
                "total_shift": total_shift,
                "recent_shift": recent_shift,
                "direction": direction,
                "odds_start": first[sc],
                "odds_now": last[sc],
            }
    return momentum


def extract_deep_patterns():
    """
    Advanced cross-match pattern mining.
    Detects behavioral patterns, timing leaks, sizing issues.
    """
    settled = get_settled_matches()
    if len(settled) < 3:
        return []

    patterns = []
    total = len(settled)

    # 1. Overtrading detection
    bet_counts = [len(m.get("bets", [])) + len(m.get("pre_bets", [])) for m in settled]
    avg_bets = np.mean(bet_counts) if bet_counts else 0
    high_bet_matches = [(m, c) for m, c in zip(settled, bet_counts) if c > avg_bets * 1.5]
    if high_bet_matches:
        high_pnl = np.mean([get_match_total_pnl(m) for m, _ in high_bet_matches])
        low_bet_matches = [m for m, c in zip(settled, bet_counts) if c <= avg_bets]
        low_pnl = np.mean([get_match_total_pnl(m) for m in low_bet_matches]) if low_bet_matches else 0
        if high_pnl < low_pnl:
            patterns.append({
                "type": "overtrading",
                "icon": "\u26a0\ufe0f",
                "insight": f"Matches with {int(avg_bets * 1.5)}+ bets average {fmt_inr(high_pnl)} vs {fmt_inr(low_pnl)} with fewer bets. Overtrading may be costing you.",
                "severity": "high" if high_pnl < 0 else "medium",
            })

    # 2. Phase timing pattern
    phase_results = {}
    for m in settled:
        ph = m.get("match_phase", "Middle")
        if ph not in phase_results:
            phase_results[ph] = []
        phase_results[ph].append(get_match_total_pnl(m))
    best_phase = max(phase_results, key=lambda p: np.mean(phase_results[p])) if phase_results else None
    worst_phase = min(phase_results, key=lambda p: np.mean(phase_results[p])) if phase_results else None
    if best_phase and worst_phase and best_phase != worst_phase:
        patterns.append({
            "type": "phase_timing",
            "icon": "\u23f0",
            "insight": f"Best results in '{best_phase}' (avg {fmt_inr(np.mean(phase_results[best_phase]))}). Worst in '{worst_phase}' (avg {fmt_inr(np.mean(phase_results[worst_phase]))}).",
            "severity": "medium",
        })

    # 3. Favourite vs underdog anchor analysis
    fav_results = []
    dog_results = []
    for m in settled:
        oo = m.get("opening_odds", {})
        bets = m.get("pre_bets", []) + m.get("bets", [])
        if not bets or not all(oo.get(k, 0) > 1 for k in ["t1", "t2"]):
            continue
        first = bets[0]
        fav = "t1" if oo.get("t1", 99) < oo.get("t2", 99) else "t2"
        if first.get("outcome") == fav:
            fav_results.append(get_match_total_pnl(m))
        else:
            dog_results.append(get_match_total_pnl(m))
    if len(fav_results) >= 2 and len(dog_results) >= 2:
        fav_avg = np.mean(fav_results)
        dog_avg = np.mean(dog_results)
        better = "favourites" if fav_avg > dog_avg else "underdogs"
        patterns.append({
            "type": "anchor_preference",
            "icon": "\U0001f3af",
            "insight": f"Anchoring on {better} works better for you. Favourites avg: {fmt_inr(fav_avg)}, Underdogs avg: {fmt_inr(dog_avg)}.",
            "severity": "low",
        })

    # 4. Capital utilization sweet spot
    utils_pnls = []
    for m in settled:
        oc = m.get("opening_capital", 50000)
        bets = m.get("pre_bets", []) + m.get("bets", []) + m.get("misc_bets", [])
        if oc > 0:
            util = compute_total_staked(bets) / oc
            utils_pnls.append((util, get_match_total_pnl(m)))
    if len(utils_pnls) >= 4:
        low_util = [p for u, p in utils_pnls if u < 0.3]
        high_util = [p for u, p in utils_pnls if u >= 0.3]
        if low_util and high_util:
            la = np.mean(low_util)
            ha = np.mean(high_util)
            patterns.append({
                "type": "utilization",
                "icon": "\U0001f4b0",
                "insight": f"Low utilization (<30%) avg P&L: {fmt_inr(la)}. High utilization (30%+) avg: {fmt_inr(ha)}. {'Conservative approach working.' if la > ha else 'Being more aggressive is paying off.'}",
                "severity": "low",
            })

    # 5. Streak-after-loss behavior
    prev_pnl = None
    chase_results = []
    for m in settled:
        rpnl = get_match_total_pnl(m)
        bets = m.get("pre_bets", []) + m.get("bets", [])
        if prev_pnl is not None and prev_pnl < 0 and bets:
            # Match after a loss — check if they sized up (chasing)
            oc = m.get("opening_capital", 50000)
            util = compute_total_staked(bets) / oc if oc > 0 else 0
            chase_results.append((util, rpnl))
        prev_pnl = rpnl
    if len(chase_results) >= 2:
        avg_chase_util = np.mean([u for u, _ in chase_results])
        avg_chase_pnl = np.mean([p for _, p in chase_results])
        if avg_chase_util > 0.3 and avg_chase_pnl < 0:
            patterns.append({
                "type": "loss_chasing",
                "icon": "\U0001f534",
                "insight": f"After losses, you stake {avg_chase_util:.0%} of capital and average {fmt_inr(avg_chase_pnl)}. Possible loss-chasing behavior.",
                "severity": "high",
            })

    # 6. Misc bet profitability
    misc_total = sum(len(m.get("misc_bets", [])) for m in settled)
    if misc_total >= 3:
        misc_pnl = sum(m.get("misc_realized_pnl", 0) for m in settled)
        misc_staked = sum(compute_total_staked(m.get("misc_bets", [])) for m in settled)
        roi = (misc_pnl / misc_staked * 100) if misc_staked > 0 else 0
        patterns.append({
            "type": "misc_performance",
            "icon": "\U0001f3b2",
            "insight": f"Misc bets: {misc_total} placed, {fmt_inr(misc_pnl)} total P&L, {roi:.1f}% ROI. {'Profitable side bets!' if roi > 0 else 'Side bets are a drag \u2014 consider reducing.'}",
            "severity": "medium" if roi < -10 else "low",
        })

    return patterns


# ── Excel Portfolio Export ────────────────────────────────────

def generate_portfolio_excel():
    """Generate a multi-sheet Excel workbook of the full portfolio."""
    if not OPENPYXL_AVAILABLE:
        return None
    wb = Workbook()
    h = st.session_state.history
    settled = get_settled_matches()
    learnings = h.get("learnings", {})

    # Styling
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0A0F1E", end_color="0A0F1E", fill_type="solid")
    green_font = Font(name="Calibri", color="00AA66")
    red_font = Font(name="Calibri", color="CC3333")
    num_fmt = '#,##0'
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header(ws, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Sheet 1: Portfolio Summary ──
    ws1 = wb.active
    ws1.title = "Portfolio Summary"
    summary_data = [
        ["Metric", "Value"],
        ["Username", st.session_state.get("active_user", "")],
        ["Current Capital", get_portfolio_capital()],
        ["Peak Capital", get_peak_capital()],
        ["Drawdown", f"{get_drawdown_pct():.1%}"],
        ["Portfolio Mode", get_portfolio_mode()[0]],
        ["Total Matches", learnings.get("total_matches", 0)],
        ["Wins", learnings.get("wins", 0)],
        ["Losses", learnings.get("losses", 0)],
        ["Win Rate", f"{learnings.get('win_rate', 0):.1%}"],
        ["Net P&L", learnings.get("total_pnl", 0)],
        ["Best Match P&L", learnings.get("best_pnl", 0)],
        ["Worst Match P&L", learnings.get("worst_pnl", 0)],
        ["Avg P&L per Match", learnings.get("avg_pnl", 0)],
        ["Current Streak", learnings.get("current_streak", 0)],
        ["Avg Bets per Match", f"{learnings.get('avg_bets_per_match', 0):.1f}"],
        ["Misc Bets Total", learnings.get("misc_bets_total", 0)],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
    ]
    for r, row in enumerate(summary_data, 1):
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if r == 1:
                cell.font = header_font; cell.fill = header_fill
            if c == 2 and isinstance(val, (int, float)) and r > 1:
                cell.number_format = num_fmt
    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 20

    # ── Sheet 2: Match History ──
    ws2 = wb.create_sheet("Match History")
    match_headers = ["Match", "Team 1", "Team 2", "Opening Capital", "Closing Capital", "Realized P&L", "Result", "Total Bets", "Settled At"]
    for c, h_val in enumerate(match_headers, 1):
        ws2.cell(row=1, column=c, value=h_val)
    style_header(ws2, len(match_headers))
    result_labels = {"t1_win": "T1 Won", "t2_win": "T2 Won", "tie": "Tie", "no_result": "N/R", "abandoned": "ABD", "manual": "Manual"}
    for r, m in enumerate(settled, 2):
        rpnl = get_match_total_pnl(m)
        nb = len(m.get("bets", [])) + len(m.get("pre_bets", [])) + len(m.get("misc_bets", []))
        vals = [m.get("label", ""), m.get("t1", ""), m.get("t2", ""),
                m.get("opening_capital", 0), m.get("closing_capital", 0), rpnl,
                result_labels.get(m.get("result", ""), "?"), nb,
                str(m.get("settled_at", ""))[:19]]
        for c, val in enumerate(vals, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = num_fmt
            if c == 6:
                cell.font = green_font if rpnl >= 0 else red_font
    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        ws2.column_dimensions[col_letter].width = 16

    # ── Sheet 3: All Bets (across all matches) ──
    ws3 = wb.create_sheet("All Bets")
    bet_headers = ["Match", "Type", "Label/Outcome", "Odds", "Stake", "Potential Profit"]
    for c, h_val in enumerate(bet_headers, 1):
        ws3.cell(row=1, column=c, value=h_val)
    style_header(ws3, len(bet_headers))
    row_idx = 2
    for m in settled:
        mlabel = m.get("label", "")
        for b in m.get("pre_bets", []) + m.get("bets", []) + m.get("misc_bets", []):
            src = b.get("source", "live")
            lbl = b.get("label", b.get("outcome", ""))
            pot = b["stake"] * (b["odds"] - 1)
            for c, val in enumerate([mlabel, src.upper(), lbl, b["odds"], b["stake"], pot], 1):
                cell = ws3.cell(row=row_idx, column=c, value=val)
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00' if c == 4 else num_fmt
            row_idx += 1
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws3.column_dimensions[col_letter].width = 18

    # ── Sheet 4: Fund Log ──
    ws4 = wb.create_sheet("Fund Log")
    fund_headers = ["Type", "Amount", "Timestamp", "Context"]
    for c, h_val in enumerate(fund_headers, 1):
        ws4.cell(row=1, column=c, value=h_val)
    style_header(ws4, len(fund_headers))
    for r, fl in enumerate(h.get("fund_log", []), 2):
        t = "Deposit" if fl.get("amount", 0) > 0 else "Withdrawal"
        for c, val in enumerate([t, fl.get("amount", 0), str(fl.get("timestamp", ""))[:19], fl.get("context", "")], 1):
            cell = ws4.cell(row=r, column=c, value=val)
            if c == 2:
                cell.number_format = num_fmt
                cell.font = green_font if fl.get("amount", 0) > 0 else red_font
    for col_letter in ["A", "B", "C", "D"]:
        ws4.column_dimensions[col_letter].width = 18

    # ── Sheet 5: Current Match (if active) ──
    if st.session_state.current_match_id:
        ws5 = wb.create_sheet("Current Match")
        cm = get_current_match()
        if cm:
            ws5.cell(row=1, column=1, value="Match"); ws5.cell(row=1, column=2, value=cm.get("label", ""))
            ws5.cell(row=2, column=1, value="Opening Capital"); ws5.cell(row=2, column=2, value=cm.get("opening_capital", 0))
            ws5.cell(row=3, column=1, value="Phase"); ws5.cell(row=3, column=2, value=st.session_state.match_phase)
            r = 5
            ws5.cell(row=r, column=1, value="Type"); ws5.cell(row=r, column=2, value="Outcome")
            ws5.cell(row=r, column=3, value="Odds"); ws5.cell(row=r, column=4, value="Stake")
            style_header(ws5, 4)
            for b in get_all_bets_including_misc():
                r += 1
                src = b.get("source", "live")
                lbl = b.get("label", b.get("outcome", ""))
                ws5.cell(row=r, column=1, value=src.upper())
                ws5.cell(row=r, column=2, value=lbl)
                ws5.cell(row=r, column=3, value=b["odds"])
                ws5.cell(row=r, column=4, value=b["stake"]).number_format = num_fmt
            # P&L scenarios
            pnl = compute_pnl(get_all_bets())
            r += 2
            ws5.cell(row=r, column=1, value="Scenario P&L").font = Font(bold=True)
            for sc, label in [("t1", st.session_state.t1_name + " Wins"), ("t2", st.session_state.t2_name + " Wins"), ("tie", "Tie")]:
                r += 1
                ws5.cell(row=r, column=1, value=label)
                cell = ws5.cell(row=r, column=2, value=pnl[sc])
                cell.number_format = num_fmt
                cell.font = green_font if pnl[sc] >= 0 else red_font
        ws5.column_dimensions["A"].width = 20
        ws5.column_dimensions["B"].width = 20

    # ── Sheet 6: Odds History (current match snapshots) ──
    snapshots = st.session_state.get("odds_snapshots", [])
    if snapshots:
        ws6 = wb.create_sheet("Odds History")
        oh_headers = ["Time", "T1 Odds", "T2 Odds", "Tie Odds", "Overround %", "Phase"]
        for c, h_val in enumerate(oh_headers, 1):
            ws6.cell(row=1, column=c, value=h_val)
        style_header(ws6, len(oh_headers))
        for r, snap in enumerate(snapshots, 2):
            o = snap.get("odds", {})
            ws6.cell(row=r, column=1, value=snap.get("timestamp", ""))
            ws6.cell(row=r, column=2, value=o.get("t1", 0))
            ws6.cell(row=r, column=3, value=o.get("t2", 0))
            ws6.cell(row=r, column=4, value=o.get("tie", 0))
            ws6.cell(row=r, column=5, value=snap.get("overround", 0))
            ws6.cell(row=r, column=6, value=snap.get("phase", ""))
        for col_letter in ["A", "B", "C", "D", "E", "F"]:
            ws6.column_dimensions[col_letter].width = 14

    # ── Sheet 7: Standalone Side Bets ──
    sm_bets = h.get("standalone_misc_bets", [])
    if sm_bets:
        ws7 = wb.create_sheet("Side Bets")
        sb_headers = ["Label", "Odds", "Stake", "Status", "P&L", "Created", "Settled"]
        for c, h_val in enumerate(sb_headers, 1):
            ws7.cell(row=1, column=c, value=h_val)
        style_header(ws7, len(sb_headers))
        for r, sb in enumerate(sm_bets, 2):
            rp = sb.get("realized_pnl", 0)
            ws7.cell(row=r, column=1, value=sb.get("label", ""))
            ws7.cell(row=r, column=2, value=sb.get("odds", 0))
            ws7.cell(row=r, column=3, value=sb.get("stake", 0)).number_format = num_fmt
            ws7.cell(row=r, column=4, value=sb.get("status", "active").title())
            cell = ws7.cell(row=r, column=5, value=rp)
            cell.number_format = num_fmt
            if sb.get("status") == "settled":
                cell.font = green_font if rp >= 0 else red_font
            ws7.cell(row=r, column=6, value=str(sb.get("created_at", ""))[:19])
            ws7.cell(row=r, column=7, value=str(sb.get("settled_at", "") or "")[:19])
        for col_letter in ["A", "B", "C", "D", "E", "F", "G"]:
            ws7.column_dimensions[col_letter].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ── Multi-User Persistence Layer ─────────────────────────────

def get_user_file(username):
    return f"user_{username}.json"


def default_history():
    return {
        "settings": {"gemini_key": ""},
        "matches": [],
        "fund_log": [],
        "learnings": {},
        "starting_capital": 50000.0,
        "standalone_misc_bets": [],
        "goals": {
            "season_target_pct": 300,  # 300% return target
            "total_matches": 70,       # matches in the season
            "per_match_style": "adaptive",  # adaptive = algo decides
            "risk_tolerance": "auto",  # auto = algorithm decides per match
        },
    }


def load_users_index():
    if os.path.exists(USERS_INDEX_FILE):
        try:
            with open(USERS_INDEX_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {"users": []}


def save_users_index(data=None):
    if data is None:
        data = st.session_state.get("users_index", {"users": []})
    try:
        with open(USERS_INDEX_FILE, "w") as f:
            json.dump(data, f, indent=2, default=str)
    except Exception:
        pass


def update_user_index_stats(username=None):
    """Refresh a user's summary stats in the master index."""
    if username is None:
        username = st.session_state.get("active_user")
    if not username:
        return
    idx = load_users_index()
    h = st.session_state.get("history", {})
    learnings = h.get("learnings", {})
    cap = get_portfolio_capital()
    health, _, _ = get_portfolio_health(learnings)
    for u in idx["users"]:
        if u["username"] == username:
            u["total_matches"] = learnings.get("total_matches", 0)
            u["current_capital"] = cap
            u["portfolio_health"] = health
            u["net_pnl"] = learnings.get("total_pnl", 0)
            break
    save_users_index(idx)
    st.session_state.users_index = idx


def load_history(username=None):
    if username is None:
        username = st.session_state.get("active_user")
    if not username:
        return default_history()
    fpath = get_user_file(username)
    if os.path.exists(fpath):
        try:
            with open(fpath) as f:
                data = json.load(f)
            d = default_history()
            for k in d:
                if k not in data:
                    data[k] = d[k]
            return data
        except Exception:
            pass
    return default_history()


def save_history(data=None):
    username = st.session_state.get("active_user")
    if not username:
        return
    if data is None:
        data = st.session_state.get("history", default_history())
    try:
        with open(get_user_file(username), "w") as f:
            json.dump(data, f, indent=2, default=str)
    except Exception:
        pass


def sync_match_to_history():
    """Sync current match session_state back into history object."""
    h = st.session_state.history
    mid = st.session_state.get("current_match_id")
    if not mid:
        return
    for m in h["matches"]:
        if m["id"] == mid:
            m["bets"] = copy.deepcopy(st.session_state.bets)
            m["pre_bets"] = copy.deepcopy(st.session_state.pre_bets)
            m["misc_bets"] = copy.deepcopy(st.session_state.get("misc_bets", []))
            m["current_odds"] = {
                "t1": st.session_state.odds_t1,
                "t2": st.session_state.odds_t2,
                "tie": st.session_state.odds_tie,
            }
            m["mode"] = st.session_state.mode
            m["match_phase"] = st.session_state.match_phase
            m["bet_counter"] = st.session_state.bet_counter
            m["pre_bet_counter"] = st.session_state.pre_bet_counter
            m["misc_bet_counter"] = st.session_state.get("misc_bet_counter", 0)
            m["odds_snapshots"] = copy.deepcopy(st.session_state.get("odds_snapshots", []))
            m["conviction"] = st.session_state.get("conviction")
            m["ponr_active"] = st.session_state.get("ponr_active", False)
            m["session_fund_adj"] = st.session_state.get("session_fund_adj", 0)
            break
    save_history(h)


def get_portfolio_capital():
    """
    Capital = starting_capital + fund_movements + in-app settled P&L + standalone misc P&L.
    Starting capital already includes all pre-app history.
    """
    h = st.session_state.get("history", default_history())
    base = h.get("starting_capital", 50000.0)
    for f in h.get("fund_log", []):
        base += f.get("amount", 0)
    for m in h.get("matches", []):
        if m.get("status") == "settled":
            base += m.get("realized_pnl", 0)
            base += m.get("misc_realized_pnl", 0)  # in-match misc bets P&L
    for sb in h.get("standalone_misc_bets", []):
        if sb.get("status") == "settled":
            base += sb.get("realized_pnl", 0)
    return base


def get_peak_capital():
    """Calculate all-time peak capital for drawdown calculation."""
    h = st.session_state.get("history", default_history())
    base = h.get("starting_capital", 50000.0)
    events = []
    for f in h.get("fund_log", []):
        events.append((f.get("timestamp", ""), f.get("amount", 0)))
    for m in h.get("matches", []):
        if m.get("status") == "settled":
            events.append((m.get("settled_at", m.get("created_at", "")), m.get("realized_pnl", 0) + m.get("misc_realized_pnl", 0)))
    for sb in h.get("standalone_misc_bets", []):
        if sb.get("status") == "settled":
            events.append((sb.get("settled_at", ""), sb.get("realized_pnl", 0)))
    events.sort(key=lambda x: str(x[0]))
    running = base
    peak = base
    for _, amt in events:
        running += amt
        if running > peak:
            peak = running
    return peak


def get_drawdown_pct():
    """Current drawdown from all-time peak as a fraction (0.0 to 1.0)."""
    peak = get_peak_capital()
    current = get_portfolio_capital()
    if peak <= 0:
        return 0.0
    return max(0.0, (peak - current) / peak)


def get_standalone_exposure():
    """Total capital locked in active standalone bets."""
    h = st.session_state.get("history", default_history())
    return sum(b.get("stake", 0) for b in h.get("standalone_misc_bets", []) if b.get("status") == "active")


def get_available_capital():
    """Portfolio capital minus active standalone bet exposure."""
    return get_portfolio_capital() - get_standalone_exposure()


def compute_goals_tracker():
    """
    Compute goal progress and per-match targets based on user's season goals.
    Returns dict with all goal metrics.
    """
    h = st.session_state.get("history", default_history())
    goals = h.get("goals", {})
    cap = get_portfolio_capital()
    starting = h.get("starting_capital", 50000)
    settled = get_settled_matches()
    matches_played = len(settled)

    target_pct = goals.get("season_target_pct", 300)
    total_matches = goals.get("total_matches", 70)
    matches_remaining = max(1, total_matches - matches_played)

    # Current progress — separate real betting profit from fund movements
    fund_total = sum(f.get("amount", 0) for f in h.get("fund_log", []))
    match_profit = sum(get_match_total_pnl(m) for m in settled)
    sm_settled_profit = sum(b.get("realized_pnl", 0) for b in h.get("standalone_misc_bets", []) if b.get("status") == "settled")
    betting_profit = match_profit + sm_settled_profit  # REAL profit from bets only
    profit_so_far = betting_profit  # excludes fund deposits
    profit_pct = (profit_so_far / starting * 100) if starting > 0 else 0
    target_profit = starting * target_pct / 100
    target_capital = starting + target_profit
    profit_needed = target_capital - cap  # cap already includes funds, so this accounts for everything
    progress_pct = (profit_so_far / target_profit * 100) if target_profit > 0 else 0

    # Per-match targets (compound growth model)
    # To go from current cap to target in remaining matches:
    # cap * (1+r)^remaining = target_capital
    # r = (target/cap)^(1/remaining) - 1
    if cap > 0 and profit_needed > 0:
        growth_rate = (target_capital / cap) ** (1.0 / matches_remaining) - 1
        per_match_profit = cap * growth_rate
        per_match_pct = growth_rate * 100
        target_achieved = False
    elif profit_needed <= 0:
        # Already hit target — suggest next milestone
        target_achieved = True
        surplus = abs(profit_needed)
        # Next target: double current capital from here
        next_target_pct = round((cap * 2 - starting) / starting * 100 / 100) * 100
        per_match_profit = cap * 0.03  # suggest 3% growth to keep compounding
        per_match_pct = 3.0
        growth_rate = 0.03
    else:
        per_match_profit = profit_needed / matches_remaining
        per_match_pct = (per_match_profit / cap * 100) if cap > 0 else 0
        growth_rate = per_match_pct / 100
        target_achieved = False

    # Recommended session stake
    avg_edge = 0.07
    if per_match_profit > 0:
        recommended_session_stake = min(per_match_profit / avg_edge, cap * 0.25)
    else:
        recommended_session_stake = cap * 0.15
    recommended_session_stake = round(max(recommended_session_stake, 500) / 10) * 10

    # Anchor sizing: fraction of session stake
    anchor_pct = min(0.25, max(0.10, growth_rate * 2))  # more aggressive if behind pace
    recommended_anchor = round(recommended_session_stake * anchor_pct / 10) * 10
    recommended_anchor = max(recommended_anchor, 300)

    # Pace check
    if matches_played > 0:
        avg_profit_per_match = profit_so_far / matches_played
        on_pace = avg_profit_per_match >= (target_profit / total_matches)
        projected_final = cap + avg_profit_per_match * matches_remaining
    else:
        avg_profit_per_match = 0
        on_pace = True  # no data yet
        projected_final = cap

    # Side bet potential
    sm_bets = [b for b in h.get("standalone_misc_bets", []) if b.get("status") == "active"]
    sm_best = sum(b["stake"] * (b["odds"] - 1) for b in sm_bets)
    sm_worst = -sum(b["stake"] for b in sm_bets)

    return {
        "target_pct": target_pct,
        "target_capital": target_capital,
        "target_profit": target_profit,
        "profit_so_far": profit_so_far,
        "betting_profit": betting_profit,
        "fund_total": fund_total,
        "match_profit": match_profit,
        "sm_settled_profit": sm_settled_profit,
        "profit_pct": profit_pct,
        "profit_needed": profit_needed,
        "target_achieved": target_achieved if 'target_achieved' in dir() else profit_needed <= 0,
        "next_target_pct": next_target_pct if 'next_target_pct' in locals() else target_pct * 2,
        "progress_pct": min(progress_pct, 100),
        "matches_played": matches_played,
        "matches_remaining": matches_remaining,
        "total_matches": total_matches,
        "per_match_profit": per_match_profit,
        "per_match_pct": per_match_pct,
        "recommended_session_stake": recommended_session_stake,
        "recommended_anchor": recommended_anchor,
        "anchor_pct": anchor_pct,
        "avg_profit_per_match": avg_profit_per_match,
        "on_pace": on_pace,
        "projected_final": projected_final,
        "sm_best_case": sm_best,
        "sm_worst_case": sm_worst,
        "cap": cap,
        "starting": starting,
    }


def get_portfolio_mode():
    """Determine portfolio mode based on cross-session state."""
    dd = get_drawdown_pct()
    if dd >= 0.20:
        return "PRESERVATION", RED, "Capital preservation \u2014 only hedge bets, no new anchors"
    learnings = st.session_state.get("history", {}).get("learnings", {})
    total_pnl = learnings.get("total_pnl", 0)
    current = get_portfolio_capital()
    peak = get_peak_capital()
    if current >= peak * 0.95 and total_pnl >= 0:
        return "GROWTH", GREEN, "Portfolio at peak \u2014 standard sizing permitted"
    if dd >= 0.10 or total_pnl < 0:
        return "STANDARD", AMBER, "Below peak \u2014 moderate caution applied"
    return "GROWTH", GREEN, "Portfolio healthy \u2014 full strategy available"


def get_settled_matches():
    return [m for m in st.session_state.history.get("matches", []) if m.get("status") == "settled"]


def get_current_match():
    mid = st.session_state.get("current_match_id")
    if not mid:
        return None
    for m in st.session_state.history.get("matches", []):
        if m["id"] == mid:
            return m
    return None


# ── Gemini AI Layer (google-genai SDK) ────────────────────────

def get_gemini_client():
    """Return a configured Gemini client, cached in session_state so it stays alive."""
    if not GENAI_AVAILABLE:
        return None
    key = st.session_state.history.get("settings", {}).get("gemini_key", "")
    if not key:
        return None
    # Check if we already have a cached client with the same key
    cached = st.session_state.get("gemini_client_cache")
    cached_key = st.session_state.get("gemini_client_key")
    if cached is not None and cached_key == key:
        return cached
    try:
        if genai_client_module:
            client = genai_client_module.Client(api_key=key)
        else:
            genai_legacy.configure(api_key=key)
            client = "legacy"
        st.session_state["gemini_client_cache"] = client
        st.session_state["gemini_client_key"] = key
        return client
    except Exception:
        return None


def gemini_generate(prompt, stream=False, use_search=False):
    """Unified generation with retry on 429 rate-limit errors. use_search=True enables Google Search grounding."""
    client = get_gemini_client()
    if client is None:
        return None

    # Throttle: ensure minimum gap between requests
    last_call = st.session_state.get("gemini_last_call_time", 0)
    elapsed = time.time() - last_call
    if elapsed < 4:
        time.sleep(4 - elapsed)

    for attempt in range(GEMINI_MAX_RETRIES):
        try:
            st.session_state["gemini_last_call_time"] = time.time()

            if client == "legacy":
                # Legacy google-generativeai library
                tools_arg = "google_search_retrieval" if use_search else None
                try:
                    model = genai_legacy.GenerativeModel(GEMINI_MODEL, system_instruction=GEMINI_SYSTEM, tools=tools_arg)
                except Exception:
                    # Fallback if tools param not supported in this version
                    model = genai_legacy.GenerativeModel(GEMINI_MODEL, system_instruction=GEMINI_SYSTEM)
                if stream:
                    chunks = []
                    for chunk in model.generate_content(prompt, stream=True):
                        if chunk.text:
                            chunks.append(chunk.text)
                    return "".join(chunks) if chunks else None
                else:
                    resp = model.generate_content(prompt)
                    return resp.text
            else:
                # New google-genai library
                tools_list = None
                if use_search:
                    try:
                        tools_list = [genai_types.Tool(google_search=genai_types.GoogleSearch())]
                    except Exception:
                        pass  # Search not available in this version, proceed without
                config = genai_types.GenerateContentConfig(
                    system_instruction=GEMINI_SYSTEM,
                    tools=tools_list,
                )
                if stream:
                    chunks = []
                    for chunk in client.models.generate_content_stream(
                        model=GEMINI_MODEL, contents=prompt, config=config,
                    ):
                        if hasattr(chunk, 'text') and chunk.text:
                            chunks.append(chunk.text)
                    return "".join(chunks) if chunks else None
                else:
                    resp = client.models.generate_content(
                        model=GEMINI_MODEL, contents=prompt, config=config,
                    )
                    return resp.text

        except Exception as e:
            err_str = str(e)
            st.session_state["gemini_last_error"] = err_str
            if "429" in err_str or "RESOURCE_EXHAUSTED" in err_str:
                wait = GEMINI_RETRY_DELAY * (attempt + 1)
                if attempt < GEMINI_MAX_RETRIES - 1:
                    time.sleep(wait)
                    continue
            return None
    return None


def test_gemini_key():
    """Quick test of the API key. Returns (success: bool, message: str)."""
    client = get_gemini_client()
    if client is None:
        return False, "No API key entered or Gemini package not installed."
    try:
        result = gemini_generate("Reply with exactly one word: OK", stream=False)
        if result and len(result.strip()) > 0:
            return True, f"\u2705 Model: {GEMINI_MODEL} \u00b7 Response: \"{result.strip()[:50]}\""
        err = st.session_state.get("gemini_last_error", "")
        if "429" in err or "RESOURCE_EXHAUSTED" in err:
            return False, f"Rate limited (429) after {GEMINI_MAX_RETRIES} retries. Wait 60 seconds and try again. Free tier allows ~15 requests/minute."
        return False, f"Empty response. Last error: {err[:150]}"
    except Exception as e:
        return False, f"Error: {str(e)[:200]}"


def gemini_live_insight(match_state_json):
    """Get live match insight from Gemini. Returns iterator for streaming."""
    prompt = (
        "Here is the current state of my IPL in-play betting session:\n\n"
        f"```json\n{match_state_json}\n```\n\n"
        "Give me a 2-3 sentence insight about my current position. "
        "Reference specific numbers. What should I pay attention to right now?"
    )
    return gemini_generate(prompt, stream=True)


def gemini_post_match_debrief(match_json):
    """Get post-match debrief from Gemini. Returns text string."""
    prompt = (
        "Here is the complete record of my settled IPL match:\n\n"
        f"```json\n{match_json}\n```\n\n"
        "Give me a 4-6 sentence debrief: what went well, what could have been "
        "timed better, and one specific lesson for my next match. "
        "Reference the actual P&L numbers and bet timings."
    )
    return gemini_generate(prompt, stream=False)


def gemini_portfolio_narrative(portfolio_json):
    """Get portfolio narrative from Gemini. Returns iterator for streaming."""
    prompt = (
        "Here is my complete IPL betting portfolio history:\n\n"
        f"```json\n{portfolio_json}\n```\n\n"
        "Analyse my portfolio: patterns you see, strengths, risks, and one "
        "actionable recommendation for my next match. "
        "Reference specific match results and capital numbers."
    )
    return gemini_generate(prompt, stream=True)


def gemini_strategy_advisor(context_json, phase="pre_match"):
    """
    AI Strategy Advisor — sharp, actionable, position-aware.
    No theory. Only decisions. References user's actual bets.
    """

    # Core system: force concise, decision-oriented output
    system_override = (
        "You are my personal IPL betting strategist. You've seen thousands of IPL matches. "
        "You think in probabilities, not opinions. You speak in numbers, not paragraphs.\n\n"
        "RULES YOU MUST FOLLOW:\n"
        "- Every prediction MUST have a specific number or range (e.g. '155-165, most likely 160')\n"
        "- Every recommendation MUST say DO or DON'T, never 'consider' or 'you might want to'\n"
        "- Reference MY ACTUAL BETS from the context and tell me what to do with each one\n"
        "- If I have active bets, assess each one: is it looking good, bad, or neutral right now?\n"
        "- For run predictions: factor wickets, pitch age, dew, matchups — NOT just run rate math\n"
        "- Compare your assessment vs the bookmaker odds — where is the VALUE gap?\n"
        "- Be contrarian when data supports it. Don't just agree with market odds.\n"
        "- Maximum 3 sentences per point. No filler. No disclaimers.\n"
        "- End with a DECISION TABLE: what to bet, at what odds, with what stake %\n\n"
    )

    # Position-aware context
    position_prompt = ""
    ctx = json.loads(context_json) if isinstance(context_json, str) else context_json

    if ctx.get("current_bets"):
        position_prompt += "\n\nMY ACTIVE POSITION:\n"
        for b in ctx["current_bets"]:
            outcome_name = ctx["team1"] if b["outcome"] == "t1" else (ctx["team2"] if b["outcome"] == "t2" else "Tie")
            position_prompt += f"  - Backed {outcome_name} @ {b['odds']:.2f} for ₹{b['stake']:,.0f}\n"
        pnl = ctx.get("pnl_scenarios", {})
        position_prompt += f"  P&L if {ctx['team1']} wins: ₹{pnl.get('t1', 0):,.0f}\n"
        position_prompt += f"  P&L if {ctx['team2']} wins: ₹{pnl.get('t2', 0):,.0f}\n"
        position_prompt += f"  Session target: ₹{ctx.get('session_target', 0):,.0f} | Floor: ₹{ctx.get('session_floor', 0):,.0f}\n"

    if ctx.get("misc_bets"):
        active_misc = [b for b in ctx["misc_bets"] if b.get("status") == "active"]
        if active_misc:
            position_prompt += "\nMY ACTIVE SIDE BETS:\n"
            for b in active_misc:
                position_prompt += f"  - {b['label']} @ {b['odds']:.2f} for ₹{b['stake']:,.0f}\n"
            position_prompt += "ASSESS each one: is it likely to WIN or LOSE based on current match state?\n"

    phase_prompts = {
        "pre_toss": f"""MATCH: {ctx.get('team1')} vs {ctx.get('team2')} — NOT STARTED YET

Give me a DECISION BRIEF (not an essay):

1. WIN PREDICTION: [Team] [X]% — one line why
2. TOSS: [Team] likely wins, will choose to [bat/bowl] — one line why
3. VENUE NUMBERS: Avg 1st innings [X], chase win% [X]%, pace/spin split
4. SCORE PREDICTIONS if batting first:
   | Phase | {ctx.get('team1')} | {ctx.get('team2')} |
   |-------|---------|---------|
   | Powerplay (6 ov) | XX-XX | XX-XX |
   | 10 overs | XX-XX | XX-XX |
   | 12 overs | XX-XX | XX-XX |
   | 15 overs | XX-XX | XX-XX |
   | Final total | XX-XX | XX-XX |
5. VALUE BETS: Compare your probabilities vs bookmaker odds ({ctx.get('live_odds', {})}). Where is the gap? Specific bets to take.
6. RISK FLAGS: 1-2 things that could upset predictions
7. STRENGTHS each team MUST capitalize on + weaknesses to exploit
8. CONVICTION: If I must pick ONE team — who and at what odds is it value?{position_prompt}""",

        "post_toss": f"""TOSS RESULT KNOWN. Update everything.
1. How does toss change win probability?
2. SCORE TABLE for batting team:
   | Milestone | Predicted Range | Most Likely |
   |-----------|----------------|-------------|
   | Powerplay | | |
   | 10 overs | | |
   | 12 overs | | |
   | 15 overs | | |
   | Total | | |
3. What bets to place NOW before first ball?
4. Are they playing to their strengths with this toss choice?{position_prompt}""",

        "powerplay": f"""POWERPLAY IS LIVE. Current situation: {ctx.get('live_score', 'unknown')}
1. vs PAR: Is this above/below par for this venue powerplay? By how much?
2. PROJECTIONS from current state:
   | Milestone | Projected | Confidence |
   |-----------|-----------|-----------|
   | End of PP | | |
   | 10 overs | | |
   | 12 overs | | |
   | 15 overs | | |
   | Total | | |
3. Has the total over/under line shifted? Which direction?
4. Win probability NOW vs pre-match — how much has it moved?
5. WHAT TO DO with my bets RIGHT NOW?{position_prompt}""",

        "middle": f"""MIDDLE OVERS. Current: {ctx.get('live_score', 'unknown')}
1. PAR CHECK: Ahead/behind par by how many runs?
2. PROJECTIONS:
   | Milestone | Projected | vs Pre-match |
   |-----------|-----------|-------------|
   | 15 overs | | |
   | Total | | |
3. Can they accelerate in death? Rate the death batting ability 1-10.
4. Momentum: Who has it? Is it about to shift?
5. HEDGE TIMING: Should I hedge NOW or wait? What odds movement do I expect in next 3-4 overs?
6. MY BETS — status check on each one{position_prompt}""",

        "death": f"""DEATH OVERS. Current: {ctx.get('live_score', 'unknown')}
1. FINAL TOTAL: [X] to [X], most likely [X]
2. CHASE-ABILITY: [X]% on this pitch. One line why.
3. HEDGE or HOLD? Specific instruction based on my position.
4. What will 2nd innings odds look like at this total?{position_prompt}""",

        "innings_break": f"""INNINGS BREAK. Score: {ctx.get('live_score', 'unknown')}
1. ABOVE or BELOW par? By how much?
2. CHASE SUCCESS: [X]% probability. Factor: pitch wear, dew, batting depth.
3. CHASE PHASE TABLE:
   | Phase | Expected for chasing team |
   |-------|--------------------------|
   | Powerplay | |
   | 10 overs | |
   | 15 overs | |
4. Where will the chase BREAK? Which over range is most dangerous?
5. BETS FOR 2ND INNINGS: Specific recommendations with odds ranges.
6. MY POSITION — what to do with each active bet?{position_prompt}""",

        "chase": f"""CHASE IS LIVE. Current: {ctx.get('live_score', 'unknown')}
1. vs DLS PAR: Ahead/behind/on track?
2. REQUIRED RATE: Is it gettable? Factor wickets in hand and who's left.
3. DECISIVE OVER: At what over will this be decided?
4. WIN PROBABILITY: [X]% right now.
5. MY BETS — HEDGE NOW or HOLD? Specific instruction.{position_prompt}""",
    }

    phase_content = phase_prompts.get(phase, phase_prompts["pre_toss"])

    search_instruction = ""
    if phase not in ("pre_toss",):
        search_instruction = (
            "\n\nIMPORTANT — FETCH LIVE DATA FIRST:\n"
            f"Before answering, search Google for the LIVE score of {ctx.get('team1')} vs {ctx.get('team2')} IPL 2026. "
            f"Get the current score, overs, wickets, and any recent events (wickets fallen, boundaries, etc). "
            f"Use this real-time data to make your predictions — do NOT guess or use only the user's input.\n"
        )
    else:
        search_instruction = (
            "\n\nIMPORTANT — FETCH MATCH DATA:\n"
            f"Search Google for: {ctx.get('team1')} vs {ctx.get('team2')} IPL 2026 preview, "
            f"venue stats, pitch report, team news, probable playing XI, head to head record, weather. "
            f"Use this data for your analysis — do NOT make up statistics.\n"
        )

    prompt = (
        f"{system_override}"
        f"CONTEXT:\n```json\n{context_json}\n```\n\n"
        f"{search_instruction}\n"
        f"{phase_content}\n\n"
        "FORMAT: Use tables where asked. Bold key numbers. No generic advice. "
        "If you don't have enough info for a specific prediction, say what you'd need to know. "
        "End with: ACTION ITEMS — numbered list of exactly what I should do right now."
    )
    return gemini_generate(prompt, stream=True, use_search=True)


def build_strategy_context():
    """Build rich context dict for the strategy advisor."""
    h = st.session_state.history
    cm = get_current_match()
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    odds = get_current_odds()
    all_bets = get_all_bets()
    pnl = compute_pnl(all_bets) if all_bets else {"t1": 0, "t2": 0, "tie": 0}
    g = h.get("goals", {})
    learnings = h.get("learnings", {})
    sl = compute_sizing_limits()

    settled = get_settled_matches()
    recent_matches = [{"label": m.get("label", "?"), "result": m.get("result", "?"),
                       "pnl": get_match_total_pnl(m)} for m in settled[-5:]]

    # Compute hedge info
    remaining_cap = max(0, st.session_state.get("total_capital", 0) + st.session_state.get("session_fund_adj", 0) - compute_total_staked(get_all_bets_including_misc()))
    hedge_info = {}
    if all_bets and all(odds[k] > 1 for k in odds):
        sol = solve_optimal_hedge(pnl, odds, remaining_cap)
        if sol:
            hedge_info = {"lock_in": sol["min_pnl"], "hedge_cost": sol["total_stake"],
                          "new_pnl": sol["new_pnl"]}

    context = {
        "team1": t1n, "team2": t2n,
        "match_label": cm.get("label", f"{t1n} vs {t2n}") if cm else f"{t1n} vs {t2n}",
        "live_odds": {"team1_win": odds.get("t1", 0), "team2_win": odds.get("t2", 0), "tie": odds.get("tie", 0)},
        "match_phase": st.session_state.match_phase,
        "mode": st.session_state.mode,
        "current_bets": [{"outcome": b.get("outcome", "?"), "odds": b["odds"], "stake": b["stake"]} for b in all_bets],
        "misc_bets": [{"label": b.get("label", ""), "odds": b["odds"], "stake": b["stake"],
                       "status": b.get("status", "active"), "realized_pnl": b.get("realized_pnl", 0)} for b in st.session_state.get("misc_bets", [])],
        "pnl_scenarios": pnl,
        "hedge_available": hedge_info,
        "session_target": g.get("session_target", 0),
        "session_floor": g.get("session_min_acceptable", 0),
        "capital": st.session_state.get("total_capital", 0),
        "remaining_capital": remaining_cap,
        "max_anchor": sl["max_anchor_amt"],
        "max_session": sl["max_session_amt"],
        "portfolio_mode": get_portfolio_mode()[0],
        "win_rate": learnings.get("win_rate", 0),
        "streak": learnings.get("current_streak", 0),
        "recent_matches": recent_matches,
        "season": "IPL 2026",
        "live_score": st.session_state.get("live_score_note", ""),
    }
    return context


# ── Cross-Match Learning Engine ──────────────────────────────

def extract_learnings():
    settled = get_settled_matches()
    if not settled:
        return {}
    learnings = {
        "total_matches": len(settled), "total_pnl": 0.0,
        "wins": 0, "losses": 0,
        "best_pnl": -float("inf"), "worst_pnl": float("inf"),
        "avg_pnl": 0.0, "win_rate": 0.0,
        "current_streak": 0, "streak_type": "none",
        "avg_bets_per_match": 0.0, "avg_overround": 0.0,
        "outcome_performance": {"t1": {"count": 0, "pnl": 0}, "t2": {"count": 0, "pnl": 0}, "tie": {"count": 0, "pnl": 0}},
        "phase_performance": {"Early": {"count": 0, "pnl": 0}, "Middle": {"count": 0, "pnl": 0}, "Last 5 Overs": {"count": 0, "pnl": 0}},
        "avg_capital_utilization": 0.0, "anchor_effectiveness": 0.0,
        "recommended_anchor_frac": 0.20, "hedge_urgency": "normal",
        "misc_bets_total": 0, "misc_bets_pnl": 0.0,
    }

    total_bets = 0; overrounds = []; cap_utils = []; anchor_contribs = []; streak = 0

    for m in settled:
        rpnl = get_match_total_pnl(m)
        learnings["total_pnl"] += rpnl
        if rpnl > learnings["best_pnl"]:
            learnings["best_pnl"] = rpnl
        if rpnl < learnings["worst_pnl"]:
            learnings["worst_pnl"] = rpnl
        if rpnl >= 0:
            learnings["wins"] += 1
            streak = streak + 1 if streak >= 0 else 1
        else:
            learnings["losses"] += 1
            streak = streak - 1 if streak <= 0 else -1

        all_bets = m.get("pre_bets", []) + m.get("bets", [])
        total_bets += len(all_bets)
        misc = m.get("misc_bets", [])
        learnings["misc_bets_total"] += len(misc)
        learnings["misc_bets_pnl"] += m.get("misc_realized_pnl", 0)

        oo = m.get("opening_odds", {})
        if all(oo.get(k, 0) > 1 for k in ["t1", "t2", "tie"]):
            overrounds.append(sum(100 / oo[k] for k in ["t1", "t2", "tie"]) - 100)

        oc = m.get("opening_capital", 50000)
        if oc > 0:
            cap_utils.append(compute_total_staked(all_bets + misc) / oc)

        result_key = {"t1_win": "t1", "t2_win": "t2", "tie": "tie"}.get(m.get("result", ""))
        if result_key:
            learnings["outcome_performance"][result_key]["count"] += 1
            learnings["outcome_performance"][result_key]["pnl"] += rpnl

        phase = m.get("match_phase", "Middle")
        learnings["phase_performance"].setdefault(phase, {"count": 0, "pnl": 0})
        learnings["phase_performance"][phase]["count"] += 1
        learnings["phase_performance"][phase]["pnl"] += rpnl

        if all_bets and rpnl != 0:
            fb = all_bets[0]
            anchor_contribs.append(abs(fb["stake"] * (fb["odds"] - 1) / max(abs(rpnl), 1)))

    n = len(settled)
    learnings["current_streak"] = streak
    learnings["streak_type"] = "winning" if streak > 0 else ("losing" if streak < 0 else "none")
    learnings["avg_pnl"] = learnings["total_pnl"] / n
    learnings["win_rate"] = learnings["wins"] / n
    learnings["avg_bets_per_match"] = total_bets / n
    learnings["avg_overround"] = float(np.mean(overrounds)) if overrounds else 0
    learnings["avg_capital_utilization"] = float(np.mean(cap_utils)) if cap_utils else 0
    learnings["anchor_effectiveness"] = float(np.mean(anchor_contribs)) if anchor_contribs else 0

    # Cross-session drawdown-aware adaptive sizing
    dd = get_drawdown_pct()
    pmode, _, _ = get_portfolio_mode()
    if pmode == "PRESERVATION":
        learnings["recommended_anchor_frac"] = 0.0
        learnings["hedge_urgency"] = "aggressive"
    elif streak <= -2 or dd >= 0.15:
        learnings["recommended_anchor_frac"] = 0.10
        learnings["hedge_urgency"] = "aggressive"
    elif dd >= 0.08 or learnings["total_pnl"] < 0:
        learnings["recommended_anchor_frac"] = 0.15
        learnings["hedge_urgency"] = "normal"
    elif streak >= 3 and dd < 0.05:
        learnings["recommended_anchor_frac"] = 0.28
        learnings["hedge_urgency"] = "normal"
    elif learnings["win_rate"] >= 0.7 and n >= 3:
        learnings["recommended_anchor_frac"] = 0.25
        learnings["hedge_urgency"] = "normal"
    elif learnings["win_rate"] < 0.4 and n >= 3:
        learnings["recommended_anchor_frac"] = 0.10
        learnings["hedge_urgency"] = "aggressive"
    else:
        learnings["recommended_anchor_frac"] = 0.20
        learnings["hedge_urgency"] = "normal"

    return learnings


def get_portfolio_health(learnings):
    if not learnings or learnings.get("total_matches", 0) == 0:
        return "HEALTHY", GREEN, "No history yet \u2014 starting fresh."
    streak = learnings.get("current_streak", 0)
    wr = learnings.get("win_rate", 0)
    total_pnl = learnings.get("total_pnl", 0)
    dd = get_drawdown_pct()

    if dd >= 0.20 or streak <= -3 or (wr < 0.3 and learnings["total_matches"] >= 5) or total_pnl < -50000:
        return "CRITICAL", RED, f"Drawdown {dd:.0%} \u00b7 Streak {streak} \u00b7 Win rate {wr:.0%} \u00b7 Net P&L {fmt_inr(total_pnl)}"
    elif dd >= 0.10 or streak <= -2 or (wr < 0.5 and learnings["total_matches"] >= 3) or total_pnl < -10000:
        return "CAUTION", AMBER, f"Drawdown {dd:.0%} \u00b7 Win rate {wr:.0%} \u00b7 Net P&L {fmt_inr(total_pnl)}"
    else:
        return "HEALTHY", GREEN, f"Win rate {wr:.0%} \u00b7 Net P&L {fmt_inr(total_pnl)} \u00b7 Streak: {'+' if streak > 0 else ''}{streak}"


def get_historical_insight_badge(learnings, rec_action, phase):
    if not learnings or learnings.get("total_matches", 0) < 2:
        return ""
    badges = []
    urgency = learnings.get("hedge_urgency", "normal")
    streak = learnings.get("current_streak", 0)
    pmode, _, _ = get_portfolio_mode()

    if pmode == "PRESERVATION":
        badges.append("\U0001f6d1 PRESERVATION MODE \u2014 hedge-only, no new anchor bets")
    elif urgency == "aggressive" and rec_action in ("ANCHOR", "HEDGE"):
        badges.append("\u26a0\ufe0f Portfolio on losing streak \u2014 tighter sizing applied")
    elif streak >= 3 and rec_action == "ANCHOR":
        badges.append("\U0001f525 Winning streak \u2014 slightly larger anchor allowed")

    if phase == "Early" and urgency == "aggressive":
        badges.append("\U0001f4ca History suggests hedging earlier in this situation")

    wr = learnings.get("win_rate", 0)
    if wr >= 0.75 and learnings["total_matches"] >= 4:
        badges.append(f"\u2705 Strong track record: {wr:.0%} win rate over {learnings['total_matches']} matches")

    if not badges:
        return ""
    html = ""
    for b in badges[:2]:
        html += (
            f'<span style="display:inline-block;padding:3px 10px;border-radius:12px;'
            f'font-family:Rajdhani,sans-serif;font-size:12px;font-weight:600;'
            f'background:rgba(255,179,71,0.15);color:{AMBER};border:1px solid rgba(255,179,71,0.3);'
            f'margin:4px 4px 4px 0;">{b}</span>'
        )
    return html


# ── Recommendation Engine (full v3 detail + cross-session overlay) ──

def generate_recommendation(all_bets, pnl, odds, capital, phase, mode, pre_bets, learnings=None):
    has_odds = all(odds[k] > 1.0 for k in odds)
    has_bets = len(all_bets) > 0
    has_pre = len(pre_bets) > 0
    pmode, _, _ = get_portfolio_mode()
    mp = compute_misc_pnl()
    misc_context = ""
    if mp["count"] > 0 and mp["active_count"] > 0:
        misc_context = f" (+{mp['active_count']} misc bets: best {fmt_inr(mp['best'])}, worst {fmt_inr(mp['worst'])})"

    if not has_odds:
        return {
            "action": "ENTER_ODDS", "icon": "\U0001f4e1",
            "headline": "Enter Current Live Odds",
            "detail": ("Input current decimal odds for all 3 outcomes to activate the engine."
                + (" Your pre-existing position is loaded \u2014 once odds are entered, the engine will "
                   "analyse your edge shift and compute optimal hedges." if has_pre else "")),
            "stakes": {}, "color": MUTED,
        }

    # Compute edge shift context for pre-existing bets
    edge_context = ""
    if has_pre and has_odds:
        shifts = [compute_edge_shift(b["odds"], odds, b["outcome"]) for b in pre_bets]
        shifts = [s for s in shifts if s]
        fav = sum(1 for s in shifts if s["direction"] == "favourable")
        against = sum(1 for s in shifts if s["direction"] == "against")
        if fav > against:
            edge_context = "Odds have moved in your favour since entry \u2014 hedging is now cheaper. "
        elif against > fav:
            edge_context = "Odds have moved against your position since entry \u2014 hedge cost is higher, but still recommended. "

    if not has_bets:
        # PRESERVATION MODE: block new anchor bets
        if pmode == "PRESERVATION":
            return {
                "action": "PRESERVATION", "icon": "\U0001f6d1",
                "headline": "Preservation Mode Active",
                "detail": "Portfolio drawdown exceeds 20%. No new anchor bets allowed. Only hedge existing positions or wait for recovery.",
                "stakes": {}, "color": RED,
            }

        # Goal-aware sizing
        gt = compute_goals_tracker()
        g_goals = st.session_state.history.get("goals", {})
        session_tgt = g_goals.get("session_target", gt["per_match_profit"])
        session_floor = g_goals.get("session_min_acceptable", 0)
        session_strat = g_goals.get("session_strategy", "conviction")

        # Compute sizing using the SAME source as Sizing Guide
        sl = compute_sizing_limits()
        suggested = sl["max_anchor_amt"]

        # Pattern mining adjustments (within the limit)
        pattern_advice = ""
        if learnings:
            total_m = learnings.get("total_matches", 0)
            avg_bets = learnings.get("avg_bets_per_match", 0)
            if avg_bets > 4 and total_m >= 3:
                pattern_advice += " Pattern: you avg {:.0f} bets/match \u2014 fewer, larger bets perform better.".format(avg_bets)
            wr = learnings.get("win_rate", 0)
            if wr >= 0.7 and total_m >= 3:
                pattern_advice += " Strong track record."
            elif wr < 0.4 and total_m >= 3:
                suggested = round(suggested * 0.7 / 10) * 10
                pattern_advice += " Win rate below 40% \u2014 reduced anchor."

        suggested = max(suggested, 300)
        anchor_pct = suggested / capital * 100 if capital > 0 else 0

        # Overround analysis
        overround = sum(100.0 / odds[k] for k in odds) - 100
        odds_quality = ""
        if overround < 3:
            odds_quality = "Overround is low ({:.1f}%) \u2014 good time to enter.".format(overround)
        elif overround > 8:
            odds_quality = "Overround is high ({:.1f}%) \u2014 consider waiting for better odds or using conviction mode to overcome margin.".format(overround)
        else:
            odds_quality = "Overround: {:.1f}%.".format(overround)

        # Sanity check session target
        target_warning = ""
        max_realistic = capital * 0.30 * max(odds[k] - 1 for k in odds)
        if session_tgt > max_realistic:
            target_warning = f" \u26a0 Session target of {fmt_inr(session_tgt)} is ambitious \u2014 would need {session_tgt/max_realistic:.0f}x your realistic max. Consider conviction mode or lowering target."
        elif session_tgt > capital * 0.15:
            target_warning = f" Session target ({fmt_inr(session_tgt)}) is {session_tgt/capital*100:.0f}% of capital \u2014 aggressive but achievable with conviction + good entry."

        # Strategy-specific guidance
        strat_text = ""
        if session_strat == "conviction":
            strat_text = "Pick your conviction team, enter at pre-match odds if possible, engine hedges the rest to break-even."
        elif session_strat == "hedge_first":
            strat_text = "Place anchor, then hedge immediately for guaranteed floor. Lower upside but safer."
        elif session_strat == "aggressive":
            strat_text = "Larger anchor, delay hedging for bigger upside. Higher risk."
        elif session_strat == "conservative":
            strat_text = "Smaller anchor, hedge early. Aim for small consistent gains."

        # Compute profit at different odds levels
        fav_odds = min(odds[k] for k in ["t1", "t2"])
        dog_odds = max(odds[k] for k in ["t1", "t2"])
        profit_at_fav = suggested * (fav_odds - 1)
        profit_at_dog = suggested * (dog_odds - 1)

        detail = (
            f"Recommended anchor: <strong>{fmt_inr(suggested)}</strong> ({anchor_pct:.0f}% of session capital). "
            f"You decide the team \u2014 here's what that stake returns:\n\n"
            f"\u2022 On favourite ({fav_odds:.2f}): profit {fmt_inr(profit_at_fav)}\n"
            f"\u2022 On underdog ({dog_odds:.2f}): profit {fmt_inr(profit_at_dog)}\n\n"
            f"Max session total: {fmt_inr(sl['max_session_amt'])} ({sl['max_session_pct']*100:.0f}%). "
            f"{odds_quality} {strat_text}{target_warning}{pattern_advice}"
        )

        return {
            "action": "STRATEGY", "icon": "\U0001f9e0",
            "headline": "Place Your Anchor Bet",
            "detail": detail,
            "stakes": {}, "color": GREEN,
        }

    min_current = min(pnl.values())
    max_current = max(pnl.values())

    # Session goals
    g_goals = st.session_state.history.get("goals", {})
    session_tgt = g_goals.get("session_target", 500)
    session_floor = g_goals.get("session_min_acceptable", 0)

    # Already guaranteed profit?
    if min_current >= 0:
        goal_status = ""
        if min_current >= session_tgt:
            goal_status = f" \u2705 Already exceeds session target ({fmt_inr(session_tgt)})."
        elif max_current >= session_tgt:
            goal_status = f" \U0001f3af Best case ({fmt_inr(max_current)}) hits target. Consider holding for upside."
        sol = solve_optimal_hedge(pnl, odds, capital)
        if sol and sol["min_pnl"] > min_current + 50:
            active_stakes = {k: v for k, v in sol["stakes"].items() if v >= 10}
            if active_stakes:
                return {
                    "action": "BOOST", "icon": "\U0001f4c8",
                    "headline": "Boost Guaranteed Profit",
                    "detail": (
                        f"{edge_context}You already guarantee {fmt_inr(min_current)} profit. "
                        f"With {fmt_inr(sol['total_stake'])} additional stake, "
                        f"raise the minimum to {fmt_inr(sol['min_pnl'])}.{goal_status}{misc_context}"
                    ),
                    "stakes": active_stakes, "new_pnl": sol["new_pnl"], "color": GREEN,
                }
        return {
            "action": "HOLD", "icon": "\U0001f6e1\ufe0f",
            "headline": "Portfolio Locked \u2014 Guaranteed Profit",
            "detail": f"Minimum profit across all outcomes: {fmt_inr(min_current)}.{goal_status}{misc_context} No further action needed.",
            "stakes": {}, "color": GREEN,
        }

    if capital < 50:
        return {
            "action": "NO_CAPITAL", "icon": "\u26a0\ufe0f",
            "headline": "Capital Exhausted",
            "detail": f"Remaining capital ({fmt_inr(capital)}) is too low to hedge effectively. Worst-case: {fmt_inr(min_current)}. Best-case: {fmt_inr(max_current)}.",
            "stakes": {}, "color": RED,
        }

    sol = solve_optimal_hedge(pnl, odds, capital)
    if sol is None:
        return {"action": "ERROR", "icon": "\u26a0\ufe0f", "headline": "Cannot Compute Hedge",
                "detail": "Check that all odds are valid decimal odds (> 1.0).", "stakes": {}, "color": RED}

    active_stakes = {k: v for k, v in sol["stakes"].items() if v >= 10}
    opt_min = sol["min_pnl"]

    # Session floor check
    floor_status = ""
    if opt_min >= session_floor:
        floor_status = f" After hedge, worst-case ({fmt_inr(opt_min)}) is above your floor ({fmt_inr(session_floor)})."
    else:
        floor_status = f" \u26a0 Even after hedge, worst-case ({fmt_inr(opt_min)}) is below your floor ({fmt_inr(session_floor)}). Entry odds may need to improve."

    if opt_min > 0:
        goal_note = f" \u2705 Target hit!" if opt_min >= session_tgt else f" Locks {fmt_inr(opt_min)} of your {fmt_inr(session_tgt)} target."
        return {
            "action": "LOCK_PROFIT", "icon": "\U0001f512",
            "headline": "Lock Guaranteed Profit NOW",
            "detail": (
                f"{edge_context}Place the hedge below to guarantee minimum {fmt_inr(opt_min)} profit.{goal_note} "
                f"Hedge stake: {fmt_inr(sol['total_stake'])}.{misc_context}"
            ),
            "stakes": active_stakes, "new_pnl": sol["new_pnl"], "color": GREEN,
        }

    if opt_min >= -10:
        return {
            "action": "LOCK_BREAKEVEN", "icon": "\u2696\ufe0f",
            "headline": "Lock Break-Even",
            "detail": (
                f"{edge_context}Eliminate virtually all loss. After hedging, worst-case: {fmt_inr(opt_min)}.{floor_status} "
                f"Hedge stake: {fmt_inr(sol['total_stake'])}.{misc_context}"
            ),
            "stakes": active_stakes, "new_pnl": sol["new_pnl"], "color": AMBER,
        }

    # Timing guidance
    phase_text = {
        "Early": "Match is early \u2014 odds will shift substantially. Wait for your conviction to play out unless your exposure exceeds 15% of capital.",
        "Middle": "Mid-match \u2014 good time to start hedging. If your team is winning, hedge now to lock a favourable position.",
        "Last 5 Overs": "Late match \u2014 hedge NOW. Odds are stabilizing and your window is closing.",
    }
    if learnings and learnings.get("hedge_urgency") == "aggressive" and phase == "Early":
        phase_text["Early"] = "Losing streak active \u2014 consider hedging earlier than usual even in early overs."

    return {
        "action": "HEDGE", "icon": "\U0001f504",
        "headline": "Reduce Downside Exposure",
        "detail": (
            f"{edge_context}Current worst-case: {fmt_inr(min_current)}. "
            f"After optimal hedge: {fmt_inr(opt_min)} "
            f"(improves by {fmt_inr(opt_min - min_current)}).{floor_status} "
            f"Hedge stake: {fmt_inr(sol['total_stake'])}.{misc_context}\n\n"
            f"*{phase_text.get(phase, '')}*"
        ),
        "stakes": active_stakes, "new_pnl": sol["new_pnl"], "color": AMBER,
    }


def build_strategy_text(all_bets, pnl, odds, capital, phase, pre_bets, live_bets):
    """Generate the full plain-English strategy explanation (restored from v3)."""
    sections = []

    sections.append("### How the Math Works\n")
    sections.append(
        "Each bet is a contract: if your chosen outcome occurs, you receive "
        "**stake \u00d7 (odds \u2212 1)** in profit. If any other outcome occurs, you lose the stake. "
        "With 3 possible outcomes (Team 1 Win, Team 2 Win, Tie), your portfolio's health is "
        "defined by 3 numbers: your net P&L under each scenario.\n"
    )

    if pre_bets:
        sections.append("### Pre-Existing Position\n")
        sections.append(
            f"You have **{len(pre_bets)}** pre-existing bet(s). "
            f"Total pre-staked: **{fmt_inr(compute_total_staked(pre_bets))}**.\n"
        )
        has_odds = all(odds[k] > 1.0 for k in odds)
        if has_odds:
            sections.append("**Odds Shift Analysis:**\n")
            t1n = st.session_state.t1_name or "Team 1"
            t2n = st.session_state.t2_name or "Team 2"
            name_map = {"t1": t1n, "t2": t2n, "tie": "Tie"}
            for b in pre_bets:
                es = compute_edge_shift(b["odds"], odds, b["outcome"])
                if es:
                    dir_emoji = "\U0001f7e2" if es["direction"] == "favourable" else ("\U0001f534" if es["direction"] == "against" else "\u26aa")
                    sections.append(
                        f"- {dir_emoji} **{name_map[b['outcome']]}** bet ({fmt_inr(b['stake'])} @ {b['odds']:.2f}): "
                        f"Entry implied {es['impl_entry']:.1f}% \u2192 Current {es['impl_now']:.1f}% "
                        f"(**{fmt_pct(es['edge_shift'])} shift**)."
                    )
                    if es["direction"] == "favourable":
                        sections.append(f"  Odds shortened {b['odds']:.2f} \u2192 {es['odds_now']:.2f}. Position **gained value** \u2014 hedge is cheaper.\n")
                    elif es["direction"] == "against":
                        sections.append(f"  Odds drifted {b['odds']:.2f} \u2192 {es['odds_now']:.2f}. Position **lost value** \u2014 engine finds optimal path.\n")
            sections.append("")
        sections.append("**Key Insight:** P&L uses your *entry odds*, not current. Current odds only matter for *new* hedge bets.\n")

    if len(all_bets) == 0:
        sections.append("**Current State:** No bets in portfolio. Import a pre-existing position or place a new bet.\n")
        return "\n".join(sections)

    sections.append("### Current Portfolio Analysis\n")
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    sc_names = {"t1": f"{t1n} Win", "t2": f"{t2n} Win", "tie": "Tie"}
    for sc in ["t1", "t2", "tie"]:
        val = pnl[sc]
        emoji = "\U0001f7e2" if val >= 0 else "\U0001f534"
        sections.append(f"- {emoji} If **{sc_names[sc]}**: {fmt_inr(val)}")
    sections.append("")

    min_pnl = min(pnl.values())
    max_pnl = max(pnl.values())
    worst_sc = min(pnl, key=pnl.get)
    best_sc = max(pnl, key=pnl.get)
    if min_pnl >= 0:
        sections.append(f"**Status: Guaranteed Profit** \u2014 worst case is {fmt_inr(min_pnl)} ({sc_names[worst_sc]}).\n")
    else:
        sections.append(f"**Exposure:** Exposed on **{sc_names[worst_sc]}** (worst: {fmt_inr(min_pnl)}). Best: **{sc_names[best_sc]}** at {fmt_inr(max_pnl)}.\n")

    has_odds = all(odds[k] > 1.0 for k in odds)
    if has_odds and min_pnl < 0 and capital >= 50:
        sections.append("### Hedge Strategy (Linear Optimization)\n")
        sections.append(
            "The engine solves a **linear program** to find stakes on any/all 3 outcomes "
            "that **maximize the worst-case P&L**, subject to remaining capital.\n\n"
            "Formally: maximize z = min(P\u2081+s\u2081(o\u2081\u22121)\u2212s\u2082\u2212s\u209c, P\u2082\u2212s\u2081+s\u2082(o\u2082\u22121)\u2212s\u209c, "
            "P\u209c\u2212s\u2081\u2212s\u2082+s\u209c(o\u209c\u22121)), where s\u2081+s\u2082+s\u209c \u2264 Capital.\n\n"
            "If z > 0 \u2192 **guaranteed profit lock**. z \u2248 0 \u2192 break-even. Otherwise \u2192 minimizes worst-case loss.\n"
        )
        if pre_bets and has_odds:
            sections.append("**Pre-Existing Impact:** Historical entry odds baked into P&L vector. Current odds used for new hedges.\n")
        if phase == "Early":
            sections.append("**Phase Note (Early):** Odds will shift substantially. Only hedge now if exposure is uncomfortably large.\n")
        elif phase == "Last 5 Overs":
            sections.append("**Phase Note (Last 5 Overs):** Odds are converging. Act promptly.\n")

    # Portfolio strategy overlay
    learnings = st.session_state.history.get("learnings", {})
    if learnings and learnings.get("total_matches", 0) >= 2:
        pmode, _, _ = get_portfolio_mode()
        dd = get_drawdown_pct()
        sections.append("### Portfolio Strategy Overlay\n")
        sections.append(f"**Mode:** {pmode} \u00b7 **Drawdown:** {dd:.1%} from peak\n")
        health, _, _ = get_portfolio_health(learnings)
        sections.append(f"**Health:** {health} \u00b7 Win rate: {learnings['win_rate']:.0%} over {learnings['total_matches']} matches.\n")
        if learnings.get("hedge_urgency") == "aggressive":
            sections.append("**Active adjustment:** Losing streak or high drawdown \u2014 tighter anchor sizing and earlier hedge triggers.\n")
        elif learnings.get("current_streak", 0) >= 3:
            sections.append("**Active adjustment:** Winning streak \u2014 slightly larger anchor allowed.\n")

    # Misc bets
    misc = st.session_state.get("misc_bets", [])
    if misc:
        mp = compute_misc_pnl()
        sections.append(f"### Miscellaneous Bets\n")
        sections.append(f"**{mp['count']}** side bet(s) active. Best case: {fmt_inr(mp['best'])}, Worst: {fmt_inr(mp['worst'])}. These are independent of match result.\n")

    tie_odds = odds.get("tie", 0)
    if tie_odds > 1:
        sections.append("### About the Tie Outcome\n")
        sections.append(
            f"Tie odds: **{tie_odds:.1f}** (implied: **{100 / tie_odds:.1f}%**). "
            f"T20 ties are very rare (<0.5%). The engine's tie hedge stake is typically tiny \u2014 mathematically sound tail-risk elimination.\n"
        )

    return "\n".join(sections)


# ── CSS / Theme (full v3 restoration) ────────────────────────

def inject_styles():
    css = (
        "<style>"
        "@import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;500;600;700&family=Orbitron:wght@400;500;600;700;800;900&display=swap');"
        ".stApp,.main .block-container{background-color:" + BG + ";color:" + TEXT + ";}"
        "header[data-testid=stHeader]{background:transparent;}"
        ".stApp>header{background-color:transparent;}"
        "section[data-testid=stSidebar]{background-color:#060B16;border-right:1px solid #1A2340;}"
        "section[data-testid=stSidebar] .stMarkdown p,"
        "section[data-testid=stSidebar] .stMarkdown label,"
        "section[data-testid=stSidebar] label{color:" + TEXT + " !important;font-family:'Rajdhani',sans-serif;}"
        "h1,h2,h3{font-family:'Orbitron',monospace !important;letter-spacing:1px;}"
        "p,li,span,div,label{font-family:'Rajdhani',sans-serif;font-weight:500;}"
        ".dash-card{background:" + CARD + ";border:1px solid " + CARD_BORDER + ";border-radius:12px;padding:20px 24px;margin-bottom:16px;}"
        ".metric-box{text-align:center;padding:12px;}"
        ".metric-label{font-family:'Rajdhani',sans-serif;font-size:13px;color:" + MUTED + ";text-transform:uppercase;letter-spacing:2px;margin-bottom:4px;}"
        ".metric-value{font-family:'Orbitron',monospace;font-size:26px;font-weight:700;}"
        ".metric-delta{font-family:'Rajdhani',sans-serif;font-size:13px;margin-top:2px;}"
        ".rec-card{background:linear-gradient(135deg,#0F1629 0%,#111D35 100%);border:2px solid;border-radius:16px;padding:28px 32px;margin:16px 0;}"
        ".rec-headline{font-family:'Orbitron',monospace;font-size:20px;font-weight:700;margin-bottom:8px;}"
        ".rec-detail{font-family:'Rajdhani',sans-serif;font-size:17px;color:" + TEXT + ";line-height:1.5;}"
        ".rec-stakes{margin-top:14px;padding:14px 18px;background:rgba(0,0,0,0.3);border-radius:10px;font-family:'Orbitron',monospace;font-size:15px;}"
        ".stProgress>div>div>div{background:linear-gradient(90deg," + GREEN + "," + CYAN + ");}"
        ".ledger-table{width:100%;border-collapse:separate;border-spacing:0 6px;font-family:'Rajdhani',sans-serif;}"
        ".ledger-table th{font-family:'Orbitron',monospace;font-size:11px;text-transform:uppercase;letter-spacing:2px;color:" + MUTED + ";padding:8px 14px;text-align:left;border-bottom:1px solid " + CARD_BORDER + ";}"
        ".ledger-table td{padding:10px 14px;font-size:15px;font-weight:600;}"
        ".ledger-row-profit{background:rgba(0,255,136,0.06);border-radius:8px;}"
        ".ledger-row-loss{background:rgba(255,75,75,0.06);border-radius:8px;}"
        ".ledger-row-pre{background:rgba(167,139,250,0.08);border-radius:8px;}"
        ".ledger-row-misc{background:rgba(123,108,246,0.08);border-radius:8px;}"
        ".stNumberInput label,.stSelectbox label,.stRadio label,.stTextInput label{font-family:'Rajdhani',sans-serif !important;font-weight:600 !important;color:" + TEXT + " !important;}"
        "input[type=number],.stTextInput input{font-family:'Orbitron',monospace !important;background-color:#111827 !important;color:" + TEXT + " !important;border-color:" + CARD_BORDER + " !important;}"
        "details summary{font-family:'Orbitron',monospace !important;font-weight:600;}"
        ".section-title{font-family:'Orbitron',monospace;font-size:15px;font-weight:600;color:" + MUTED + ";letter-spacing:3px;text-transform:uppercase;margin-bottom:12px;padding-bottom:8px;border-bottom:1px solid " + CARD_BORDER + ";}"
        ".edge-pill{display:inline-block;padding:2px 10px;border-radius:12px;font-family:'Orbitron',monospace;font-size:11px;font-weight:600;letter-spacing:1px;}"
        ".edge-fav{background:rgba(0,255,136,0.15);color:" + GREEN + ";border:1px solid rgba(0,255,136,0.3);}"
        ".edge-bad{background:rgba(255,75,75,0.15);color:" + RED + ";border:1px solid rgba(255,75,75,0.3);}"
        ".edge-flat{background:rgba(100,116,139,0.15);color:" + MUTED + ";border:1px solid rgba(100,116,139,0.3);}"
        ".pre-badge{display:inline-block;padding:1px 8px;border-radius:8px;font-family:'Orbitron',monospace;font-size:9px;font-weight:600;background:rgba(167,139,250,0.2);color:" + VIOLET + ";border:1px solid rgba(167,139,250,0.35);letter-spacing:1px;vertical-align:middle;margin-left:6px;}"
        ".live-badge{display:inline-block;padding:1px 8px;border-radius:8px;font-family:'Orbitron',monospace;font-size:9px;font-weight:600;background:rgba(0,255,136,0.15);color:" + GREEN + ";border:1px solid rgba(0,255,136,0.3);letter-spacing:1px;vertical-align:middle;margin-left:6px;}"
        ".misc-badge{display:inline-block;padding:1px 8px;border-radius:8px;font-family:'Orbitron',monospace;font-size:9px;font-weight:600;background:rgba(123,108,246,0.15);color:" + MISC_PURPLE + ";border:1px solid rgba(123,108,246,0.3);letter-spacing:1px;vertical-align:middle;margin-left:6px;}"
        ".gemini-card{background:linear-gradient(135deg,#0F1629 0%,#1A1505 100%);border-left:4px solid " + AMBER + ";border-radius:12px;padding:18px 22px;margin:12px 0;}"
        ".gemini-card p{font-style:italic;color:" + TEXT + ";}"
        ".health-badge{display:inline-block;padding:6px 18px;border-radius:20px;font-family:'Orbitron',monospace;font-size:13px;font-weight:700;letter-spacing:2px;}"
        ".user-card{background:" + CARD + ";border:2px solid " + CARD_BORDER + ";border-radius:16px;padding:24px;margin:8px 0;transition:border-color 0.2s;}"
        ".user-card:hover{border-color:" + GREEN + ";}"
        "</style>"
    )
    st.markdown(css, unsafe_allow_html=True)


# ── Session State ────────────────────────────────────────────

def init_state():
    if "users_index" not in st.session_state:
        st.session_state.users_index = load_users_index()
    if "active_user" not in st.session_state:
        st.session_state.active_user = None
    if "history" not in st.session_state:
        st.session_state.history = default_history()
    defaults = {
        "total_capital": 50000.0,
        "bets": [], "pre_bets": [], "misc_bets": [],
        "odds_t1": 0.0, "odds_t2": 0.0, "odds_tie": 0.0,
        "mode": "Pre-Match", "match_phase": "Early",
        "bet_counter": 0, "pre_bet_counter": 0, "misc_bet_counter": 0,
        "t1_name": "Team 1", "t2_name": "Team 2",
        "celebration_fired": False,
        "current_match_id": None,
        "app_view": "match",
        "odds_snapshots": [],  # within-match odds history
        "conviction": None,  # "t1", "t2", or None
        "ponr_active": False,  # Point of No Return activated
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v
    # Initialize live odds feed state (additive — does not touch any existing key)
    _live_feed_init_state()
    # Initialize risk guard state (additive)
    _risk_guard_init_state()


def load_user_session(username):
    """Load a user's data into session state."""
    st.session_state.active_user = username
    st.session_state.history = load_history(username)
    if not st.session_state.history.get("learnings"):
        st.session_state.history["learnings"] = extract_learnings()
    st.session_state.current_match_id = None
    st.session_state.bets = []
    st.session_state.pre_bets = []
    st.session_state.misc_bets = []
    st.session_state.celebration_fired = False


def load_match_into_session(match):
    """Load a match record from history into active session_state."""
    st.session_state.current_match_id = match["id"]
    st.session_state.t1_name = match.get("t1", "Team 1")
    st.session_state.t2_name = match.get("t2", "Team 2")
    st.session_state.total_capital = match.get("opening_capital", 50000)
    st.session_state.session_fund_adj = match.get("session_fund_adj", 0)
    st.session_state.bets = match.get("bets", [])
    st.session_state.pre_bets = match.get("pre_bets", [])
    st.session_state.misc_bets = match.get("misc_bets", [])
    co = match.get("current_odds", {})
    st.session_state.odds_t1 = co.get("t1", 0.0)
    st.session_state.odds_t2 = co.get("t2", 0.0)
    st.session_state.odds_tie = co.get("tie", 0.0)
    st.session_state.mode = match.get("mode", "Live")
    st.session_state.match_phase = match.get("match_phase", "Early")
    st.session_state.bet_counter = match.get("bet_counter", 0)
    st.session_state.pre_bet_counter = match.get("pre_bet_counter", 0)
    st.session_state.misc_bet_counter = match.get("misc_bet_counter", 0)
    st.session_state.odds_snapshots = match.get("odds_snapshots", [])
    st.session_state.conviction = match.get("conviction")
    st.session_state.ponr_active = match.get("ponr_active", False)
    st.session_state.celebration_fired = False
    st.session_state.app_view = "match"


# ── User Selection Screen ────────────────────────────────────

def render_user_selection():
    """Entry gate \u2014 first thing the app shows."""
    inject_styles()
    _, center, _ = st.columns([1, 2, 1])
    with center:
        st.markdown(f"""
        <div style="text-align:center; padding: 30px 0 10px;">
            <div style="font-family:'Orbitron',monospace; font-size:32px; font-weight:800;
                        background: linear-gradient(135deg, {GREEN}, {CYAN});
                        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
                        letter-spacing: 3px;">
                IPL HEDGE ENGINE
            </div>
            <div style="font-family:'Rajdhani',sans-serif; font-size:15px; color:{MUTED};
                        letter-spacing: 4px; text-transform:uppercase; margin-top:2px;">
                Multi-User Portfolio System
            </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown(f'<div class="section-title" style="text-align:center;">\U0001f464  Select User</div>', unsafe_allow_html=True)

    idx = st.session_state.users_index
    users = idx.get("users", [])

    if users:
        cols = st.columns(min(len(users), 3))
        for i, u in enumerate(users):
            with cols[i % 3]:
                hc = {"HEALTHY": GREEN, "CAUTION": AMBER, "CRITICAL": RED}.get(u.get("portfolio_health", "HEALTHY"), GREEN)
                st.markdown(f"""
                <div class="user-card">
                    <div style="font-family:'Orbitron',monospace;font-size:18px;font-weight:700;color:{GREEN};">
                        {u.get("display_name", u["username"])}
                    </div>
                    <div style="font-family:'Rajdhani';font-size:13px;color:{MUTED};margin-top:4px;">
                        @{u["username"]}
                    </div>
                    <div style="margin-top:12px;display:flex;justify-content:space-between;">
                        <div>
                            <div style="font-size:11px;color:{MUTED};text-transform:uppercase;letter-spacing:1px;">Capital</div>
                            <div style="font-family:'Orbitron',monospace;font-size:14px;color:{TEXT};">{fmt_inr(u.get("current_capital", 50000))}</div>
                        </div>
                        <div>
                            <div style="font-size:11px;color:{MUTED};text-transform:uppercase;letter-spacing:1px;">Matches</div>
                            <div style="font-family:'Orbitron',monospace;font-size:14px;color:{TEXT};">{u.get("total_matches", 0)}</div>
                        </div>
                        <div>
                            <div style="font-size:11px;color:{MUTED};text-transform:uppercase;letter-spacing:1px;">Health</div>
                            <div style="font-family:'Orbitron',monospace;font-size:12px;color:{hc};">{u.get("portfolio_health", "NEW")}</div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                ub1, ub2 = st.columns([3, 1])
                if ub1.button(f"\u25b6 Play as {u.get('display_name', u['username'])}", key=f"sel_{u['username']}", use_container_width=True):
                    load_user_session(u["username"])
                    st.rerun()
                if ub2.button("\U0001f5d1", key=f"delu_{u['username']}", help=f"Delete user {u['username']}"):
                    st.session_state[f"confirm_del_user_{u['username']}"] = True
                if st.session_state.get(f"confirm_del_user_{u['username']}"):
                    st.warning(f"Permanently delete **{u['username']}** and all their data?")
                    cd1, cd2 = st.columns(2)
                    if cd1.button("\u2713 Yes, delete", key=f"cdu_y_{u['username']}", type="primary"):
                        idx["users"] = [x for x in idx["users"] if x["username"] != u["username"]]
                        save_users_index(idx)
                        fp = get_user_file(u["username"])
                        if os.path.exists(fp):
                            os.remove(fp)
                        st.session_state[f"confirm_del_user_{u['username']}"] = False
                        st.rerun()
                    if cd2.button("\u2717 Cancel", key=f"cdu_n_{u['username']}"):
                        st.session_state[f"confirm_del_user_{u['username']}"] = False
                        st.rerun()

    st.markdown(f'<div style="margin-top:20px;"></div>', unsafe_allow_html=True)
    with st.expander("\u2795 Create New User", expanded=len(users) == 0):
        st.markdown(f'<div style="font-family:Rajdhani;font-size:15px;color:{MUTED};margin-bottom:12px;">'
            f'Your starting capital is what you have <strong style="color:{GREEN};">right now</strong> \u2014 '
            f'any past profits or losses are already included in this number.</div>', unsafe_allow_html=True)
        with st.form("create_user_form"):
            nu = st.text_input("Username (unique, no spaces)", placeholder="e.g. virat_k")
            nd = st.text_input("Display Name (optional)", placeholder="e.g. Virat")
            nc = st.number_input("Starting Capital (\u20b9) \u2014 your current total funds", min_value=1000.0, value=50000.0, step=5000.0, format="%.0f")
            if st.form_submit_button("\U0001f464 Create User", use_container_width=True, type="primary"):
                nu_clean = nu.strip().lower().replace(" ", "_")
                if not nu_clean:
                    st.error("Username required.")
                elif any(u["username"] == nu_clean for u in users):
                    st.error("Username already exists.")
                else:
                    new_user = {
                        "username": nu_clean, "display_name": nd or nu_clean,
                        "current_capital": nc, "total_matches": 0,
                        "portfolio_health": "HEALTHY", "net_pnl": 0,
                        "created_at": datetime.now().isoformat(),
                    }
                    idx["users"].append(new_user)
                    save_users_index(idx)
                    h = default_history()
                    h["starting_capital"] = nc
                    st.session_state.history = h
                    st.session_state.active_user = nu_clean
                    save_history(h)
                    load_user_session(nu_clean)
                    st.rerun()


# ── Render: Header, Fund Panel, Sidebar ──────────────────────

def render_header():
    st.markdown(f"""
    <div style="text-align:center; padding: 10px 0 6px 0;">
        <div style="font-family:'Orbitron',monospace; font-size:32px; font-weight:800;
                    background: linear-gradient(135deg, {GREEN}, {CYAN});
                    -webkit-background-clip: text; -webkit-text-fill-color: transparent;
                    letter-spacing: 3px;">
            IPL HEDGE ENGINE
        </div>
        <div style="font-family:'Rajdhani',sans-serif; font-size:15px; color:{MUTED};
                    letter-spacing: 4px; text-transform:uppercase; margin-top:2px;">
            In-Play Betting \u00b7 Risk Elimination \u00b7 Real-Time
        </div>
    </div>
    """, unsafe_allow_html=True)


def render_fund_sidebar_panel():
    """Compact fund management \u2014 always visible in sidebar."""
    h = st.session_state.history
    cap = get_portfolio_capital()
    st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{AMBER};letter-spacing:2px;margin:8px 0;">\U0001f4b0 FUNDS</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:16px;color:{GREEN};margin-bottom:8px;">{fmt_inr(cap)}</div>', unsafe_allow_html=True)
    fc1, fc2 = st.columns(2)
    with fc1:
        add_amt = st.number_input("Add \u20b9", min_value=0.0, step=1000.0, value=0.0, format="%.0f", key="sb_add_amt", label_visibility="collapsed")
    with fc2:
        if st.button("\u2795 Add", key="sb_add_btn", use_container_width=True):
            if add_amt > 0:
                ctx = st.session_state.match_phase if st.session_state.current_match_id else "pre-session"
                h["fund_log"].append({"type": "deposit", "amount": add_amt, "timestamp": datetime.now().isoformat(), "context": ctx})
                if st.session_state.get("current_match_id"):
                    st.session_state["session_fund_adj"] = st.session_state.get("session_fund_adj", 0) + add_amt
                save_history(h); update_user_index_stats(); st.rerun()
    wc1, wc2 = st.columns(2)
    with wc1:
        w_amt = st.number_input("Out \u20b9", min_value=0.0, step=1000.0, value=0.0, format="%.0f", key="sb_w_amt", label_visibility="collapsed")
    with wc2:
        if st.button("\U0001f4e4 Out", key="sb_w_btn", use_container_width=True):
            if w_amt > 0:
                exposure = compute_total_staked(get_all_bets_including_misc())
                available = cap - exposure
                if w_amt > available + 0.01:
                    st.error(f"Max withdrawal: {fmt_inr(available)}")
                else:
                    ctx = st.session_state.match_phase if st.session_state.current_match_id else "pre-session"
                    h["fund_log"].append({"type": "withdrawal", "amount": -w_amt, "timestamp": datetime.now().isoformat(), "context": ctx})
                    if st.session_state.get("current_match_id"):
                        st.session_state["session_fund_adj"] = st.session_state.get("session_fund_adj", 0) - w_amt
                    save_history(h); update_user_index_stats(); st.rerun()
    # Last 5 transactions (with delete)
    flog = h.get("fund_log", [])
    recent = flog[-5:] if flog else []
    if recent:
        for fl_display in reversed(recent):
            fl_idx = flog.index(fl_display)
            t = "ADD" if fl_display.get("amount", 0) > 0 else "OUT"
            c = GREEN if fl_display.get("amount", 0) > 0 else RED
            ts = str(fl_display.get("timestamp", ""))[:16]
            fc_row1, fc_row2 = st.columns([5, 1])
            fc_row1.markdown(f'<div style="font-family:Rajdhani;font-size:12px;color:{c};">{t} {fmt_inr(abs(fl_display.get("amount", 0)))} <span style="color:{MUTED};">{ts}</span></div>', unsafe_allow_html=True)
            if fc_row2.button("\u2717", key=f"del_fund_{fl_idx}", help="Delete this fund entry"):
                h["fund_log"].pop(fl_idx)
                save_history(h); update_user_index_stats(); st.rerun()


def render_sidebar():
    with st.sidebar:
        # Switch user \u2014 always at top
        if st.button("\U0001f464 Switch User", key="switch_user_btn", use_container_width=True, type="secondary"):
            sync_match_to_history(); update_user_index_stats()
            st.session_state.active_user = None
            st.session_state.current_match_id = None
            st.rerun()

        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:14px;font-weight:600;color:{GREEN};letter-spacing:2px;margin:8px 0 16px;">\u2699 CONTROL PANEL</div>', unsafe_allow_html=True)

        user = st.session_state.active_user
        if user:
            st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{CYAN};margin-bottom:4px;">\U0001f464 {user}</div>', unsafe_allow_html=True)
        cm = get_current_match()
        if cm:
            st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{AMBER};margin-bottom:8px;">\U0001f4cd {cm.get("label", "Match")}</div>', unsafe_allow_html=True)

        st.text_input("Team 1 Name", key="t1_name", placeholder="e.g. CSK")
        st.text_input("Team 2 Name", key="t2_name", placeholder="e.g. MI")

        # Live odds feed config (additive — see LIVE ODDS FEED block above)
        render_live_feed_config()

        # Risk guards config (additive — see RISK GUARDS block above)
        render_risk_guard_config()

        st.divider()

        st.markdown(f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-bottom:4px;">Tradable capital for this match</div>', unsafe_allow_html=True)
        # Pre-render: if sync was requested last run, update before widget renders
        if st.session_state.get("_sync_capital_requested"):
            st.session_state["_prev_total_capital"] = st.session_state.get("total_capital", 0)  # save for undo
            st.session_state["total_capital"] = get_available_capital()
            st.session_state["_sync_capital_requested"] = False
        if st.session_state.get("_undo_sync_requested"):
            prev = st.session_state.get("_prev_total_capital", 0)
            if prev > 0:
                st.session_state["total_capital"] = prev
            st.session_state["_undo_sync_requested"] = False
        sc1, sc2, sc3 = st.columns([3, 0.5, 0.5])
        with sc1:
            st.number_input("Session Capital (\u20b9)", min_value=1000.0, max_value=100_000_000.0, step=5000.0, key="total_capital", format="%.0f", label_visibility="collapsed")
        with sc2:
            if st.button("\U0001f504", key="sync_cap_btn", help="Sync to current available capital", use_container_width=True):
                st.session_state["_sync_capital_requested"] = True
                st.rerun()
        with sc3:
            if st.session_state.get("_prev_total_capital", 0) > 0:
                if st.button("\u21a9", key="undo_sync_btn", help="Undo last sync", use_container_width=True):
                    st.session_state["_undo_sync_requested"] = True
                    st.rerun()

        all_bets = get_all_bets_including_misc()
        staked = compute_total_staked(all_bets)
        sm_exp = get_standalone_exposure()
        misc_realized_pnl = sum(b.get("realized_pnl", 0) for b in st.session_state.get("misc_bets", []) if b.get("status") == "settled")
        fund_adj = st.session_state.get("session_fund_adj", 0)
        effective_capital = max(0, st.session_state.total_capital + misc_realized_pnl + fund_adj)
        remaining = max(0, effective_capital - staked)
        utilization = staked / effective_capital if effective_capital > 0 else 0

        st.markdown(f'<div class="metric-label">Capital Deployed</div>', unsafe_allow_html=True)
        st.progress(min(utilization, 1.0))

        # Show effective capital if adjusted
        if fund_adj != 0:
            adj_c = GREEN if fund_adj > 0 else RED
            st.markdown(f'<div style="font-size:11px;color:{adj_c};margin-top:-10px;margin-bottom:4px;">Funds {"added" if fund_adj > 0 else "withdrawn"} this session: {fmt_inr(abs(fund_adj))} \u2192 Effective: {fmt_inr(effective_capital)}</div>', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        c1.markdown(f'<div style="font-family:Orbitron,monospace;font-size:13px;color:{GREEN};">{fmt_inr(remaining)}</div><div style="font-size:11px;color:{MUTED};">REMAINING</div>', unsafe_allow_html=True)
        c2.markdown(f'<div style="font-family:Orbitron,monospace;font-size:13px;color:{AMBER};">{fmt_inr(staked)}</div><div style="font-size:11px;color:{MUTED};">DEPLOYED</div>', unsafe_allow_html=True)

        # Capital breakdown
        parts = []
        pre_total = compute_total_staked(st.session_state.pre_bets)
        live_total = compute_total_staked(st.session_state.bets)
        active_misc = [b for b in st.session_state.get("misc_bets", []) if b.get("status", "active") == "active"]
        misc_total = compute_total_staked(active_misc)
        if pre_total > 0:
            parts.append(f'<span style="color:{VIOLET};">\u25a0</span> Pre: {fmt_inr(pre_total)}')
        if live_total > 0:
            parts.append(f'<span style="color:{GREEN};">\u25a0</span> Live: {fmt_inr(live_total)}')
        if misc_total > 0:
            parts.append(f'<span style="color:{MISC_PURPLE};">\u25a0</span> Misc: {fmt_inr(misc_total)}')
        if misc_realized_pnl != 0:
            rc = GREEN if misc_realized_pnl > 0 else RED
            parts.append(f'<span style="color:{rc};">\u25a0</span> Misc P&L: {fmt_inr(misc_realized_pnl)}')
        if sm_exp > 0:
            parts.append(f'<span style="color:{AMBER};">\U0001f3b2</span> Side bets: {fmt_inr(sm_exp)} (separate)')
        if parts:
            st.markdown(f'<div style="font-size:12px;color:{MUTED};margin-top:6px;">{" \u00b7 ".join(parts)}</div>', unsafe_allow_html=True)

        st.divider()
        st.radio("Mode", ["Pre-Match", "Live"], key="mode", horizontal=True)
        st.radio("Match Phase", ["Early", "Middle", "Last 5 Overs"], key="match_phase", horizontal=True)

        # ── Session Goal (adjustable per match) ──
        st.divider()
        g = st.session_state.history.get("goals", {})
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:10px;color:{CYAN};letter-spacing:2px;margin-bottom:4px;">\U0001f3af THIS SESSION</div>', unsafe_allow_html=True)
        sg1, sg2 = st.columns(2)
        with sg1:
            session_tgt = st.number_input("Target \u20b9", min_value=0, max_value=500000, value=int(g.get("session_target", 500)), step=100, key="sidebar_session_target", label_visibility="collapsed")
        with sg2:
            session_floor = st.number_input("Floor \u20b9", min_value=-5000, max_value=50000, value=int(g.get("session_min_acceptable", 0)), step=100, key="sidebar_session_floor", label_visibility="collapsed")
        # Save if changed
        if session_tgt != g.get("session_target") or session_floor != g.get("session_min_acceptable"):
            g["session_target"] = session_tgt
            g["session_min_acceptable"] = session_floor
            st.session_state.history["goals"] = g
            save_history(st.session_state.history)
        st.markdown(
            f'<div style="font-family:Rajdhani;font-size:11px;color:{MUTED};">'
            f'Aim: <span style="color:{GREEN};">{fmt_inr(session_tgt)}</span> '
            f'\u00b7 Floor: <span style="color:{AMBER if session_floor >= 0 else RED};">{fmt_inr(session_floor)}</span>'
            f'</div>', unsafe_allow_html=True,
        )

        st.divider()
        # Fund management always accessible
        render_fund_sidebar_panel()
        st.divider()

        # Settle match button
        if st.session_state.current_match_id and cm and cm.get("status") != "settled":
            if st.button("\U0001f3c1 Settle Match", use_container_width=True, type="primary"):
                st.session_state["show_settle"] = True

        if st.session_state.current_match_id:
            if st.button("\u25c0 Back to Lobby", use_container_width=True, type="secondary"):
                sync_match_to_history(); update_user_index_stats()
                st.session_state.current_match_id = None
                st.rerun()

        if st.button("\U0001f5d1 Reset Match Bets", use_container_width=True, type="secondary"):
            st.session_state["confirm_reset"] = True
        if st.session_state.get("confirm_reset"):
            st.warning("Delete ALL bets in current match?")
            rc1, rc2 = st.columns(2)
            if rc1.button("\u2713 Yes", use_container_width=True, type="primary", key="cr_yes"):
                st.session_state.bets = []; st.session_state.pre_bets = []; st.session_state.misc_bets = []
                st.session_state.bet_counter = 0; st.session_state.pre_bet_counter = 0; st.session_state.misc_bet_counter = 0
                st.session_state.celebration_fired = False
                st.session_state["confirm_reset"] = False
                sync_match_to_history(); st.rerun()
            if rc2.button("\u2717 No", use_container_width=True, key="cr_no"):
                st.session_state["confirm_reset"] = False; st.rerun()

    return remaining


# ── Render: Odds, Pre-Existing, Edge Shift (full v3 detail) ──

def get_current_odds():
    return {"t1": st.session_state.odds_t1, "t2": st.session_state.odds_t2, "tie": st.session_state.odds_tie}


# ============================================================
# ── LIVE ODDS FEED ───────────────────────────────────────────
# Self-contained module for fetching live odds from The Odds
# API and (optionally) match state from CricketData.org. Also
# saves snapshots to ./captures/<match>.json in the same format
# data_feed.py uses, so we keep one canonical capture file.
#
# This block does NOT touch any existing dashboard logic. It
# only adds new functions called from render_odds_panel and
# render_sidebar. If the API key is missing or the request
# fails, manual odds entry continues to work exactly as before.
# ============================================================

import urllib.request as _u_request
import urllib.parse as _u_parse
import urllib.error as _u_error
import statistics as _statistics
from pathlib import Path as _Path
from datetime import timezone as _tz, timedelta as _td

LIVE_FEED_DEFAULTS = {
    "live_feed_odds_api_key":     "ae69ee92e01205d6df9221986d9ca548",
    "live_feed_cricket_api_key":  "015ec189-9cc1-4d3a-b0b6-c2fd4da5a013",
    "live_feed_enabled":          False,    # auto-refresh toggle
    "live_feed_interval_sec":     120,      # auto-refresh cadence
    "live_feed_last_fetch":       0.0,      # epoch seconds of last successful fetch
    "live_feed_last_error":       None,     # last error message
    "live_feed_quota_remaining":  None,     # int or None
    "live_feed_match_label":      None,     # which API fixture we matched
    "live_feed_book_count":       None,     # how many books in last snapshot
    "live_feed_overround":        None,     # last overround pct
    "live_feed_capture_path":     None,     # path to current capture file
    "live_feed_snapshot_count":   0,        # snapshots saved this session
    "live_feed_match_phase":      None,     # cricket data phase if available
    "live_feed_state_text":       None,     # "Innings 2, Over 12.3, MI 95/3" type string
}

LIVE_FEED_SPORT_KEY = "cricket_ipl"
LIVE_FEED_ODDS_BASE = "https://api.the-odds-api.com/v4"
LIVE_FEED_CRICKET_BASE = "https://api.cricapi.com/v1"
LIVE_FEED_CAPTURES_DIR = _Path("./captures")
LIVE_FEED_SGT = _tz(_td(hours=8))

# IPL franchise abbreviation → canonical full names. Used to expand the
# user's short input ("MI", "RR") into matchable substrings before sending
# to the fixture-matching logic. Each value is a list of accepted aliases.
LIVE_FEED_IPL_TEAMS = {
    "mi":   ["mumbai", "indians", "mumbai indians"],
    "csk":  ["chennai", "super kings", "chennai super kings"],
    "rcb":  ["bangalore", "bengaluru", "royal challengers", "royal challengers bangalore", "royal challengers bengaluru"],
    "kkr":  ["kolkata", "knight riders", "kolkata knight riders"],
    "dc":   ["delhi", "capitals", "delhi capitals"],
    "pbks": ["punjab", "kings", "punjab kings"],
    "kxip": ["punjab", "kings", "punjab kings"],   # legacy abbreviation
    "rr":   ["rajasthan", "royals", "rajasthan royals"],
    "srh":  ["sunrisers", "hyderabad", "sunrisers hyderabad"],
    "gt":   ["gujarat", "titans", "gujarat titans"],
    "lsg":  ["lucknow", "super giants", "lucknow super giants"],
    "pwi":  ["pune", "warriors", "pune warriors"],          # historical
    "rps":  ["rising pune", "supergiant", "rising pune supergiant"],  # historical
    "gl":   ["gujarat lions"],                               # historical
    "ktk":  ["kochi", "tuskers", "kochi tuskers kerala"],    # historical
}


def _live_feed_expand_team_aliases(user_name):
    """
    Given a user-entered team name (could be 'MI', 'mumbai', 'Mumbai Indians',
    etc.), return a list of lowercase substring aliases to try matching against
    The Odds API team names. Always includes the original input.
    """
    if not user_name:
        return []
    raw = user_name.strip().lower()
    aliases = {raw}
    # Direct abbreviation hit
    if raw in LIVE_FEED_IPL_TEAMS:
        aliases.update(LIVE_FEED_IPL_TEAMS[raw])
    # Also try collapsed (no spaces) version
    collapsed = raw.replace(" ", "")
    if collapsed in LIVE_FEED_IPL_TEAMS:
        aliases.update(LIVE_FEED_IPL_TEAMS[collapsed])
    # Also include any individual word ≥ 3 chars (so "Mumbai Indians" → "mumbai", "indians")
    for tok in raw.split():
        if len(tok) >= 3:
            aliases.add(tok)
    return list(aliases)


def _live_feed_init_state():
    """Idempotent — called from init_state()."""
    for k, v in LIVE_FEED_DEFAULTS.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _live_feed_http_get(url, params=None, timeout=12):
    """Plain GET → parsed JSON. Returns (data, headers_dict, error_msg)."""
    if params:
        url = url + "?" + _u_parse.urlencode(params)
    try:
        req = _u_request.Request(url, headers={"User-Agent": "ipl-dashboard/1.0"})
        with _u_request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read().decode("utf-8")
            data = json.loads(raw)
            headers = {k.lower(): v for k, v in resp.headers.items()}
            return data, headers, None
    except _u_error.HTTPError as e:
        body = ""
        try:
            body = e.read().decode("utf-8", errors="ignore")[:200]
        except Exception:
            pass
        return None, {}, f"HTTP {e.code}: {body or e.reason}"
    except _u_error.URLError as e:
        return None, {}, f"Network error: {e.reason}"
    except Exception as e:
        return None, {}, f"Unexpected error: {type(e).__name__}: {e}"


def _live_feed_match_fixture(api_data, t1_name, t2_name):
    """
    Find the fixture in The Odds API response whose home/away teams best
    match the user's t1_name and t2_name. Uses alias expansion so users can
    type abbreviations like "MI", "RR", "CSK" and they map to full names.
    Returns (fixture_dict, swapped_flag).
    swapped=True means the API's home_team corresponds to the user's t2.
    """
    if not api_data or not isinstance(api_data, list):
        return None, False

    t1_aliases = _live_feed_expand_team_aliases(t1_name)
    t2_aliases = _live_feed_expand_team_aliases(t2_name)

    if not t1_aliases or not t2_aliases:
        return None, False

    def team_matches_aliases(api_team, alias_list):
        """Return True if any alias is a substring of the API team name."""
        if not api_team:
            return False
        api_lower = api_team.lower()
        return any(alias in api_lower for alias in alias_list if alias)

    best = None
    for fx in api_data:
        home = fx.get("home_team", "")
        away = fx.get("away_team", "")
        # Direct: t1=home, t2=away
        if team_matches_aliases(home, t1_aliases) and team_matches_aliases(away, t2_aliases):
            return fx, False
        # Swapped: t1=away, t2=home
        if team_matches_aliases(home, t2_aliases) and team_matches_aliases(away, t1_aliases):
            return fx, True
        # Partial fallback (matches only one team)
        if best is None:
            t1_in_home = team_matches_aliases(home, t1_aliases)
            t1_in_away = team_matches_aliases(away, t1_aliases)
            t2_in_home = team_matches_aliases(home, t2_aliases)
            t2_in_away = team_matches_aliases(away, t2_aliases)
            if t1_in_home or t1_in_away or t2_in_home or t2_in_away:
                # Determine swap based on which side t1 was found on
                best = (fx, t1_in_away)
    return best if best else (None, False)


def _live_feed_extract_odds(fixture, swapped):
    """
    Extract median decimal odds across all bookmakers for h2h market.
    Returns dict {t1, t2, tie, n_books, raw_per_book, overround_pct}.
    """
    bookmakers = fixture.get("bookmakers") or []
    t1_quotes, t2_quotes, tie_quotes = [], [], []
    raw_per_book = []
    home_team = fixture.get("home_team", "")
    away_team = fixture.get("away_team", "")

    for bm in bookmakers:
        bm_key = bm.get("key", "?")
        markets = bm.get("markets") or []
        h2h = next((m for m in markets if m.get("key") == "h2h"), None)
        if not h2h:
            continue
        outcomes = h2h.get("outcomes") or []
        bm_t1 = bm_t2 = bm_tie = None
        for o in outcomes:
            name = o.get("name", "")
            price = o.get("price", 0)
            if not isinstance(price, (int, float)) or price <= 1:
                continue
            if name == home_team:
                if swapped:
                    bm_t2 = price
                else:
                    bm_t1 = price
            elif name == away_team:
                if swapped:
                    bm_t1 = price
                else:
                    bm_t2 = price
            elif name.lower() in ("draw", "tie"):
                bm_tie = price
        if bm_t1 and bm_t2:
            t1_quotes.append(bm_t1)
            t2_quotes.append(bm_t2)
            if bm_tie:
                tie_quotes.append(bm_tie)
            raw_per_book.append({
                "book": bm_key, "t1": bm_t1, "t2": bm_t2,
                "tie": bm_tie, "last_update": bm.get("last_update"),
            })

    if not t1_quotes or not t2_quotes:
        return None

    t1_med = float(_statistics.median(t1_quotes))
    t2_med = float(_statistics.median(t2_quotes))
    tie_med = float(_statistics.median(tie_quotes)) if tie_quotes else 50.0  # fallback
    overround = (1 / t1_med + 1 / t2_med + 1 / tie_med - 1) * 100

    return {
        "t1": t1_med, "t2": t2_med, "tie": tie_med,
        "n_books": len(t1_quotes), "raw_per_book": raw_per_book,
        "overround_pct": overround,
    }


def _live_feed_save_capture(fixture, extracted, cricket_state=None):
    """
    Append a snapshot to ./captures/<match>.json (same format as data_feed.py).
    Returns the path written, or None on error.
    """
    try:
        LIVE_FEED_CAPTURES_DIR.mkdir(parents=True, exist_ok=True)
        commence = fixture.get("commence_time", "")
        date_part = commence.split("T")[0] if commence else datetime.utcnow().strftime("%Y-%m-%d")
        home = (fixture.get("home_team") or "home").replace(" ", "_").lower()
        away = (fixture.get("away_team") or "away").replace(" ", "_").lower()
        fname = f"{date_part}__{home}__vs__{away}.json"
        fpath = LIVE_FEED_CAPTURES_DIR / fname

        # Load or initialize
        if fpath.exists():
            try:
                cap = json.loads(fpath.read_text(encoding="utf-8"))
            except Exception:
                cap = {"snapshots": []}
        else:
            cap = {
                "fixture_id": fixture.get("id"),
                "sport_key": fixture.get("sport_key"),
                "home_team": fixture.get("home_team"),
                "away_team": fixture.get("away_team"),
                "commence_time": fixture.get("commence_time"),
                "captured_via": "ipl_betting_dashboard",
                "snapshots": [],
            }

        snapshot = {
            "ts_utc": datetime.utcnow().isoformat() + "Z",
            "ts_sgt": datetime.now(LIVE_FEED_SGT).isoformat(),
            "median_odds": {
                "t1": extracted["t1"], "t2": extracted["t2"], "tie": extracted["tie"],
            },
            "n_books": extracted["n_books"],
            "overround_pct": extracted["overround_pct"],
            "raw_per_book": extracted["raw_per_book"],
            "cricket_state": cricket_state,
            "source": "dashboard_fetch",
        }
        cap.setdefault("snapshots", []).append(snapshot)
        fpath.write_text(json.dumps(cap, indent=2), encoding="utf-8")
        st.session_state.live_feed_capture_path = str(fpath)
        st.session_state.live_feed_snapshot_count = len(cap["snapshots"])
        return fpath
    except Exception as e:
        st.session_state.live_feed_last_error = f"Capture save failed: {e}"
        return None


def _live_feed_fetch_cricket_state(api_key):
    """
    Optional: fetch current match state from CricketData.org. Returns
    (phase_str, summary_str) or (None, None) on any failure. Best-effort.
    """
    if not api_key:
        return None, None
    try:
        # Get current matches
        url = f"{LIVE_FEED_CRICKET_BASE}/currentMatches"
        data, _, err = _live_feed_http_get(url, {"apikey": api_key, "offset": 0})
        if err or not data or not data.get("data"):
            return None, None
        # Find an IPL match
        t1_aliases = _live_feed_expand_team_aliases(st.session_state.t1_name or "")
        t2_aliases = _live_feed_expand_team_aliases(st.session_state.t2_name or "")
        for m in data.get("data", []):
            teams_str = " ".join(m.get("teams", [])).lower()
            t1_hit = any(a in teams_str for a in t1_aliases if a)
            t2_hit = any(a in teams_str for a in t2_aliases if a)
            if t1_hit or t2_hit:
                status = m.get("status", "")
                # Crude phase classification
                phase = "pre_match"
                if "innings" in status.lower() or "over" in status.lower():
                    phase = "in_play"
                elif "won" in status.lower() or "lost" in status.lower() or "ended" in status.lower():
                    phase = "post_match"
                return phase, status[:120]
        return None, None
    except Exception:
        return None, None


def live_feed_fetch_now(silent=False):
    """
    Fetch live odds and write them into session_state.odds_t1/t2/tie.
    Also saves a snapshot to captures/. Returns True on success.
    """
    api_key = st.session_state.get("live_feed_odds_api_key", "").strip()
    if not api_key:
        st.session_state.live_feed_last_error = "No Odds API key configured."
        return False

    t1n = st.session_state.get("t1_name", "").strip()
    t2n = st.session_state.get("t2_name", "").strip()
    if not t1n or not t2n or t1n == "Team 1" or t2n == "Team 2":
        st.session_state.live_feed_last_error = "Set Team 1 and Team 2 names first."
        return False

    url = f"{LIVE_FEED_ODDS_BASE}/sports/{LIVE_FEED_SPORT_KEY}/odds"
    params = {
        "apiKey": api_key,
        "regions": "eu,uk",
        "markets": "h2h",
        "oddsFormat": "decimal",
        "dateFormat": "iso",
    }
    data, headers, err = _live_feed_http_get(url, params)
    if err:
        st.session_state.live_feed_last_error = err
        return False

    # Quota tracking
    try:
        st.session_state.live_feed_quota_remaining = int(headers.get("x-requests-remaining", "0"))
    except Exception:
        pass

    fixture, swapped = _live_feed_match_fixture(data, t1n, t2n)
    if not fixture:
        st.session_state.live_feed_last_error = (
            f"No IPL fixture matched '{t1n}' vs '{t2n}'. "
            f"API returned {len(data) if data else 0} fixtures."
        )
        return False

    extracted = _live_feed_extract_odds(fixture, swapped)
    if not extracted:
        st.session_state.live_feed_last_error = "Fixture matched but no h2h odds available."
        return False

    # Optional: cricket state
    cricket_state = None
    cricket_key = st.session_state.get("live_feed_cricket_api_key", "").strip()
    if cricket_key:
        phase, summary = _live_feed_fetch_cricket_state(cricket_key)
        if phase or summary:
            cricket_state = {"phase": phase, "summary": summary}
            st.session_state.live_feed_match_phase = phase
            st.session_state.live_feed_state_text = summary

    # WRITE INTO DASHBOARD STATE — this is the integration point
    st.session_state.odds_t1 = round(extracted["t1"], 2)
    st.session_state.odds_t2 = round(extracted["t2"], 2)
    st.session_state.odds_tie = round(extracted["tie"], 2)

    # Update meta
    st.session_state.live_feed_last_fetch = time.time()
    st.session_state.live_feed_last_error = None
    st.session_state.live_feed_match_label = f"{fixture.get('home_team', '?')} vs {fixture.get('away_team', '?')}"
    st.session_state.live_feed_book_count = extracted["n_books"]
    st.session_state.live_feed_overround = extracted["overround_pct"]

    # Save snapshot to captures/
    _live_feed_save_capture(fixture, extracted, cricket_state)

    return True


def _live_feed_seconds_since_last_fetch():
    last = st.session_state.get("live_feed_last_fetch", 0)
    if last <= 0:
        return None
    return time.time() - last


def _live_feed_should_auto_refresh():
    if not st.session_state.get("live_feed_enabled", False):
        return False
    interval = st.session_state.get("live_feed_interval_sec", 120)
    elapsed = _live_feed_seconds_since_last_fetch()
    if elapsed is None:
        return True  # never fetched yet
    return elapsed >= interval


def render_live_feed_config():
    """
    Sidebar config block for the live odds feed. Render this from
    render_sidebar() inside an expander to keep the sidebar tidy.
    """
    with st.expander("\U0001f4e1 Live Odds Feed", expanded=False):
        st.text_input(
            "Odds API key",
            key="live_feed_odds_api_key",
            type="password",
            help="Get one free at the-odds-api.com (500 credits/month)",
        )
        st.text_input(
            "CricketData.org key (optional)",
            key="live_feed_cricket_api_key",
            type="password",
            help="Optional — used for live match phase / over info",
        )

        col_a, col_b = st.columns(2)
        with col_a:
            if st.button("\U0001f504 Fetch now", key="live_feed_fetch_btn", use_container_width=True):
                with st.spinner("Fetching..."):
                    ok = live_feed_fetch_now()
                if ok:
                    st.success(f"Updated. {st.session_state.live_feed_book_count} books.")
                    st.rerun()
                else:
                    st.error(st.session_state.live_feed_last_error or "Fetch failed.")

        with col_b:
            st.toggle(
                "Auto-refresh",
                key="live_feed_enabled",
                help="Automatically fetch new odds at the interval below",
            )

        st.slider(
            "Refresh interval (seconds)",
            min_value=30, max_value=600, step=30,
            key="live_feed_interval_sec",
        )

        # Status display
        last = st.session_state.get("live_feed_last_fetch", 0)
        if last > 0:
            elapsed = int(time.time() - last)
            quota = st.session_state.get("live_feed_quota_remaining")
            books = st.session_state.get("live_feed_book_count")
            overround = st.session_state.get("live_feed_overround")
            snaps = st.session_state.get("live_feed_snapshot_count", 0)
            label = st.session_state.get("live_feed_match_label", "?")
            phase_text = st.session_state.get("live_feed_state_text")

            st.markdown(
                f"<div style='font-family:Rajdhani;font-size:11px;color:{MUTED};line-height:1.6;'>"
                f"<b style='color:{GREEN};'>{label}</b><br/>"
                f"Last fetch: {elapsed}s ago<br/>"
                f"Books: {books or '?'} \u00b7 Overround: {overround:.2f}%<br/>"
                f"Snapshots saved: <b style='color:{CYAN};'>{snaps}</b><br/>"
                f"API credits left: <b>{quota if quota is not None else '?'}</b>"
                + (f"<br/>State: {phase_text}" if phase_text else "")
                + "</div>",
                unsafe_allow_html=True,
            )

        if st.session_state.get("live_feed_last_error"):
            st.markdown(
                f"<div style='font-family:Rajdhani;font-size:11px;color:{RED};margin-top:6px;'>"
                f"\u26a0 {st.session_state.live_feed_last_error}</div>",
                unsafe_allow_html=True,
            )

        # Auto-refresh tick: if due, fetch silently and rerun
        if _live_feed_should_auto_refresh():
            ok = live_feed_fetch_now(silent=True)
            if ok:
                st.rerun()


# ============================================================
# ── END LIVE ODDS FEED ───────────────────────────────────────
# ============================================================


# ============================================================
# ── RISK GUARDS ──────────────────────────────────────────────
# Hard-cap protection layer. Three mechanical rules that REFUSE
# to let bets through when they would breach pre-set limits:
#
#   1. SESSION LOSS FLOOR  — stop-loss on total session P&L
#   2. MISC BET CAP        — cap on total misc exposure as %
#                            of session capital
#   3. SINGLE-BET CAP      — max individual bet size as %
#                            of session capital
#
# Plus a soft guard:
#   4. VOLATILITY PAUSE    — any single snapshot odds move
#                            > threshold pp triggers a 5-min
#                            mandatory pause before any new bets
#
# All guards have per-session override (type OVERRIDE to bypass).
# The guards are FUSES, not intelligence. They blow on thresholds.
#
# Designed after the RR vs MI rain-break loss on Apr 7 2026:
# market moved MI from 1.67 → 3.25 in one snapshot (rain cut
# game from 20 overs to 11) and the user kept loading up on
# MI miscellaneous bets based on stale pre-match intuition.
# A 15pp volatility pause would have prevented that.
# ============================================================

RISK_GUARD_DEFAULTS = {
    "risk_loss_floor":              -3000.0,  # ₹ — stop adding bets below this realized+unrealized
    "risk_loss_floor_override":     False,
    "risk_misc_cap_pct":            0.10,     # 10% of session capital max on misc
    "risk_misc_cap_override":       False,
    "risk_single_bet_pct":          0.15,     # single bet cannot exceed 15% of session
    "risk_single_bet_override":     False,
    "risk_volatility_threshold_pp": 15.0,     # pp change in implied prob that triggers pause
    "risk_volatility_pause_sec":    300,      # 5 min mandatory cooldown
    "risk_volatility_override":     False,
    "risk_last_spike_ts":           0.0,      # epoch of last detected volatility spike
    "risk_last_spike_info":         None,     # dict describing the last spike
}


def _risk_guard_init_state():
    """Idempotent — called from init_state()."""
    for k, v in RISK_GUARD_DEFAULTS.items():
        if k not in st.session_state:
            st.session_state[k] = v


def _risk_get_current_session_pnl():
    """
    Compute realized + unrealized P&L for the current session.
    Realized: settled misc bets only (main bets are still open mid-match)
    Unrealized: use worst-case P&L of the main bet ledger as the conservative
    anchor for "how bad could things get from here before I place more bets"
    """
    misc_bets = st.session_state.get("misc_bets", [])
    realized_misc = sum(
        b.get("realized_pnl", 0) for b in misc_bets
        if b.get("status") == "settled"
    )
    # Worst-case scenario of existing main bets
    pnl_scenarios = compute_pnl(get_all_bets())
    worst_case_main = min(pnl_scenarios.values()) if pnl_scenarios else 0
    # Total staked on active misc bets (assume full loss in worst case)
    active_misc_stake = sum(
        b.get("stake", 0) for b in misc_bets
        if b.get("status", "active") == "active"
    )
    worst_case_total = realized_misc + worst_case_main - active_misc_stake
    return {
        "realized_misc": realized_misc,
        "worst_case_main": worst_case_main,
        "active_misc_stake": active_misc_stake,
        "worst_case_total": worst_case_total,
    }


def _risk_check_loss_floor():
    """Returns (allowed, reason, details) for placing any new bet."""
    if st.session_state.get("risk_loss_floor_override"):
        return True, None, None
    floor = float(st.session_state.get("risk_loss_floor", -3000))
    if floor >= 0:
        return True, None, None  # floor disabled
    pnl = _risk_get_current_session_pnl()
    if pnl["worst_case_total"] < floor:
        return False, (
            f"Session loss floor breached. Worst-case P&L "
            f"₹{pnl['worst_case_total']:,.0f} < floor ₹{floor:,.0f}."
        ), pnl
    return True, None, pnl


def _risk_check_misc_cap(new_stake):
    """Check whether adding `new_stake` to misc pushes total > cap."""
    if st.session_state.get("risk_misc_cap_override"):
        return True, None
    cap_pct = float(st.session_state.get("risk_misc_cap_pct", 0.10))
    if cap_pct <= 0:
        return True, None
    session_cap = float(st.session_state.get("total_capital", 0))
    cap_rupees = session_cap * cap_pct
    active_misc = [
        b for b in st.session_state.get("misc_bets", [])
        if b.get("status", "active") == "active"
    ]
    current_misc_stake = sum(b.get("stake", 0) for b in active_misc)
    projected = current_misc_stake + new_stake
    if projected > cap_rupees:
        return False, (
            f"Misc bet cap breached. Projected misc total ₹{projected:,.0f} "
            f"exceeds cap ₹{cap_rupees:,.0f} ({cap_pct*100:.0f}% of session). "
            f"Current misc exposure: ₹{current_misc_stake:,.0f}."
        )
    return True, None


def _risk_check_single_bet(stake):
    """Check whether a single bet exceeds the max-single-bet cap."""
    if st.session_state.get("risk_single_bet_override"):
        return True, None
    cap_pct = float(st.session_state.get("risk_single_bet_pct", 0.15))
    if cap_pct <= 0:
        return True, None
    session_cap = float(st.session_state.get("total_capital", 0))
    cap_rupees = session_cap * cap_pct
    if stake > cap_rupees:
        return False, (
            f"Single-bet cap breached. Stake ₹{stake:,.0f} exceeds "
            f"cap ₹{cap_rupees:,.0f} ({cap_pct*100:.0f}% of session capital)."
        )
    return True, None


def _risk_detect_volatility_spike():
    """
    Examine the last 2 odds snapshots (whichever source wrote them) and
    compute the implied probability change for each outcome. If any outcome's
    p-change exceeds the threshold, record a spike and return it.
    Called after every odds update.
    """
    snaps = st.session_state.get("odds_snapshots", [])
    if len(snaps) < 2:
        return None
    threshold = float(st.session_state.get("risk_volatility_threshold_pp", 15.0))

    def to_probs(snap):
        o = snap.get("odds") or snap
        try:
            t1, t2, tie = float(o.get("t1", 0)), float(o.get("t2", 0)), float(o.get("tie", 0))
            if t1 <= 1 or t2 <= 1:
                return None
            raw = [1/t1, 1/t2, 1/tie if tie > 1 else 0.02]
            Z = sum(raw)
            return [r/Z * 100 for r in raw]
        except Exception:
            return None

    latest = to_probs(snaps[-1])
    prior = to_probs(snaps[-2])
    if not latest or not prior:
        return None

    deltas = [abs(l - p) for l, p in zip(latest, prior)]
    max_delta = max(deltas)
    if max_delta >= threshold:
        labels = ["T1", "T2", "Tie"]
        worst_idx = deltas.index(max_delta)
        info = {
            "ts": time.time(),
            "max_delta_pp": max_delta,
            "worst_outcome": labels[worst_idx],
            "prior_probs": prior,
            "latest_probs": latest,
        }
        st.session_state.risk_last_spike_ts = info["ts"]
        st.session_state.risk_last_spike_info = info
        return info
    return None


def _risk_check_volatility_pause():
    """
    Returns (allowed, reason). If the most recent volatility spike is within
    the mandatory cooldown window AND not overridden, block new bets.
    """
    if st.session_state.get("risk_volatility_override"):
        return True, None
    last_spike = st.session_state.get("risk_last_spike_ts", 0)
    if last_spike <= 0:
        return True, None
    cooldown = float(st.session_state.get("risk_volatility_pause_sec", 300))
    elapsed = time.time() - last_spike
    if elapsed < cooldown:
        remaining = int(cooldown - elapsed)
        info = st.session_state.get("risk_last_spike_info") or {}
        return False, (
            f"Volatility pause active. Market moved "
            f"{info.get('max_delta_pp', 0):.1f}pp on "
            f"{info.get('worst_outcome', '?')} — "
            f"{remaining}s cooldown remaining. Re-evaluate before betting."
        )
    return True, None


def risk_check_all_gates(stake, is_misc=False):
    """
    Master gate check called from bet-submission handlers.
    Returns (allowed, list_of_blocking_reasons).
    """
    blockers = []

    ok_floor, reason_floor, _ = _risk_check_loss_floor()
    if not ok_floor:
        blockers.append(("loss_floor", reason_floor))

    ok_vol, reason_vol = _risk_check_volatility_pause()
    if not ok_vol:
        blockers.append(("volatility", reason_vol))

    ok_single, reason_single = _risk_check_single_bet(stake)
    if not ok_single:
        blockers.append(("single_bet", reason_single))

    if is_misc:
        ok_misc, reason_misc = _risk_check_misc_cap(stake)
        if not ok_misc:
            blockers.append(("misc_cap", reason_misc))

    return (len(blockers) == 0), blockers


def render_risk_guard_config():
    """Sidebar expander for risk guard settings."""
    with st.expander("\U0001f6e1 Risk Guards", expanded=False):
        st.markdown(
            f"<div style='font-family:Rajdhani;font-size:11px;color:{MUTED};margin-bottom:8px;'>"
            f"Hard caps that refuse to submit bets breaching thresholds. "
            f"Built as FUSES, not intelligence. Override requires typing OVERRIDE."
            f"</div>",
            unsafe_allow_html=True,
        )

        st.number_input(
            "Session loss floor (\u20b9, negative)",
            min_value=-1_000_000.0, max_value=0.0, step=500.0,
            key="risk_loss_floor",
            help="Stop all new bets when worst-case session P&L drops below this. "
                 "Set to 0 to disable.",
        )
        st.slider(
            "Misc bet cap (% of session)",
            min_value=0.0, max_value=1.0, step=0.05,
            key="risk_misc_cap_pct",
            help="Max total misc exposure as fraction of session capital. "
                 "Set to 0 to disable. Tonight this would have capped misc at "
                 "10% of your bankroll.",
        )
        st.slider(
            "Single-bet cap (% of session)",
            min_value=0.0, max_value=1.0, step=0.05,
            key="risk_single_bet_pct",
            help="Max single bet size as fraction of session capital.",
        )
        st.slider(
            "Volatility spike threshold (pp)",
            min_value=5.0, max_value=50.0, step=1.0,
            key="risk_volatility_threshold_pp",
            help="If implied probability moves by more than this in one snapshot, "
                 "a mandatory pause is triggered. Tonight's rain-break move was 29pp. "
                 "Default 15pp would have paused you.",
        )
        st.number_input(
            "Volatility pause duration (sec)",
            min_value=0, max_value=1800, step=30,
            key="risk_volatility_pause_sec",
        )

        # Live status
        pnl = _risk_get_current_session_pnl()
        floor = st.session_state.get("risk_loss_floor", 0)
        floor_pct = (pnl["worst_case_total"] / floor * 100) if floor < 0 else 0
        floor_color = RED if floor_pct > 80 else (AMBER if floor_pct > 50 else GREEN)

        active_misc_stake = sum(
            b.get("stake", 0) for b in st.session_state.get("misc_bets", [])
            if b.get("status", "active") == "active"
        )
        misc_cap = float(st.session_state.get("total_capital", 0)) * float(st.session_state.get("risk_misc_cap_pct", 0))
        misc_usage_pct = (active_misc_stake / misc_cap * 100) if misc_cap > 0 else 0
        misc_color = RED if misc_usage_pct > 90 else (AMBER if misc_usage_pct > 70 else GREEN)

        st.markdown(
            f"<div style='font-family:Rajdhani;font-size:11px;color:{MUTED};line-height:1.6;margin-top:8px;'>"
            f"<b>LIVE STATUS</b><br/>"
            f"Worst-case session P&L: <span style='color:{floor_color};'>\u20b9{pnl['worst_case_total']:,.0f}</span> "
            f"(floor \u20b9{floor:,.0f})<br/>"
            f"Misc exposure: <span style='color:{misc_color};'>\u20b9{active_misc_stake:,.0f}</span> / \u20b9{misc_cap:,.0f} "
            f"({misc_usage_pct:.0f}%)<br/>"
            f"</div>",
            unsafe_allow_html=True,
        )

        # Volatility spike indicator
        spike = st.session_state.get("risk_last_spike_info")
        last_spike_ts = st.session_state.get("risk_last_spike_ts", 0)
        if spike and last_spike_ts > 0:
            elapsed = int(time.time() - last_spike_ts)
            cooldown = int(st.session_state.get("risk_volatility_pause_sec", 300))
            if elapsed < cooldown:
                remaining = cooldown - elapsed
                st.markdown(
                    f"<div style='background:{RED};color:white;padding:8px;border-radius:4px;"
                    f"font-family:Rajdhani;font-size:12px;margin-top:8px;'>"
                    f"\u26a0 VOLATILITY PAUSE \u2014 {spike['max_delta_pp']:.1f}pp move on "
                    f"{spike['worst_outcome']}. {remaining}s remaining.</div>",
                    unsafe_allow_html=True,
                )

        # Override controls
        with st.container():
            override_text = st.text_input(
                "Type OVERRIDE to bypass all guards this session",
                key="risk_override_input",
                help="Emergency bypass — all caps disabled until next session reset",
            )
            if override_text.strip().upper() == "OVERRIDE":
                st.session_state.risk_loss_floor_override = True
                st.session_state.risk_misc_cap_override = True
                st.session_state.risk_single_bet_override = True
                st.session_state.risk_volatility_override = True
                st.markdown(
                    f"<div style='color:{RED};font-family:Rajdhani;font-size:11px;'>"
                    f"\u26a0 ALL GUARDS OVERRIDDEN FOR THIS SESSION</div>",
                    unsafe_allow_html=True,
                )

        # Reset overrides button
        if (st.session_state.get("risk_loss_floor_override") or
            st.session_state.get("risk_misc_cap_override") or
            st.session_state.get("risk_single_bet_override") or
            st.session_state.get("risk_volatility_override")):
            if st.button("\u21a9 Reset overrides", key="risk_reset_overrides", use_container_width=True):
                st.session_state.risk_loss_floor_override = False
                st.session_state.risk_misc_cap_override = False
                st.session_state.risk_single_bet_override = False
                st.session_state.risk_volatility_override = False
                st.rerun()


def render_risk_banner():
    """
    Top-of-page banner shown when any guard is actively blocking OR about
    to block. Called from main content area.
    """
    pnl = _risk_get_current_session_pnl()
    floor = float(st.session_state.get("risk_loss_floor", 0))

    messages = []

    # Loss floor
    if floor < 0:
        pct_to_floor = (pnl["worst_case_total"] / floor) if floor != 0 else 0
        if pnl["worst_case_total"] < floor and not st.session_state.get("risk_loss_floor_override"):
            messages.append((
                "critical",
                f"\U0001f6d1 SESSION LOSS FLOOR BREACHED \u2014 "
                f"worst-case P&L \u20b9{pnl['worst_case_total']:,.0f} < floor \u20b9{floor:,.0f}. "
                f"New bets BLOCKED."
            ))
        elif pct_to_floor > 0.8:
            messages.append((
                "warning",
                f"\u26a0 Approaching loss floor: worst-case "
                f"\u20b9{pnl['worst_case_total']:,.0f} vs floor \u20b9{floor:,.0f} "
                f"({pct_to_floor*100:.0f}%)"
            ))

    # Volatility pause
    ok_vol, reason_vol = _risk_check_volatility_pause()
    if not ok_vol:
        messages.append(("critical", f"\u23f8 {reason_vol}"))

    for severity, msg in messages:
        color = RED if severity == "critical" else AMBER
        st.markdown(
            f"<div style='background:{color};color:white;padding:12px 18px;"
            f"border-radius:6px;font-family:Rajdhani;font-size:14px;font-weight:600;"
            f"margin-bottom:12px;'>{msg}</div>",
            unsafe_allow_html=True,
        )


# ============================================================
# ── END RISK GUARDS ──────────────────────────────────────────
# ============================================================


def render_odds_panel():
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    st.markdown('<div class="section-title">\U0001f4e1  Live Odds</div>', unsafe_allow_html=True)

    # Live-feed quick action row (additive — manual entry below is unchanged)
    if st.session_state.get("live_feed_odds_api_key"):
        lc1, lc2 = st.columns([1, 3])
        with lc1:
            if st.button("\u26a1 Fetch", key="odds_panel_fetch_btn", help="Fetch live odds from bookmakers", use_container_width=True):
                with st.spinner("Fetching live odds..."):
                    ok = live_feed_fetch_now()
                if ok:
                    st.toast(f"Odds updated from {st.session_state.live_feed_book_count} books", icon="\u26a1")
                    st.rerun()
                else:
                    st.error(st.session_state.get("live_feed_last_error") or "Fetch failed")
        with lc2:
            last = st.session_state.get("live_feed_last_fetch", 0)
            if last > 0:
                age = int(time.time() - last)
                snaps = st.session_state.get("live_feed_snapshot_count", 0)
                quota = st.session_state.get("live_feed_quota_remaining")
                auto_on = "\U0001f7e2 AUTO" if st.session_state.get("live_feed_enabled") else "\u26ab MANUAL"
                st.markdown(
                    f"<div style='font-family:Rajdhani;font-size:11px;color:{MUTED};padding-top:8px;'>"
                    f"{auto_on} \u00b7 last {age}s ago \u00b7 {snaps} snapshots \u00b7 {quota or '?'} credits left"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    f"<div style='font-family:Rajdhani;font-size:11px;color:{MUTED};padding-top:8px;'>"
                    f"Not yet fetched. Configure in sidebar \u2192 Live Odds Feed."
                    f"</div>",
                    unsafe_allow_html=True,
                )

    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-weight:600;color:{GREEN};font-size:14px;">{t1n} Win</div>', unsafe_allow_html=True)
        st.number_input("Odds T1", min_value=0.0, step=0.05, key="odds_t1", format="%.2f", label_visibility="collapsed")
    with c2:
        st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-weight:600;color:{CYAN};font-size:14px;">{t2n} Win</div>', unsafe_allow_html=True)
        st.number_input("Odds T2", min_value=0.0, step=0.05, key="odds_t2", format="%.2f", label_visibility="collapsed")
    with c3:
        st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-weight:600;color:{AMBER};font-size:14px;">Tie</div>', unsafe_allow_html=True)
        st.number_input("Odds Tie", min_value=0.0, step=1.0, key="odds_tie", format="%.2f", label_visibility="collapsed")

    odds = get_current_odds()
    if all(odds[k] > 1 for k in odds):
        # Log snapshot on every render (deduped inside function)
        log_odds_snapshot()

        # Implied probabilities
        total_impl = sum(100 / odds[k] for k in odds)
        probs = {k: 100 / odds[k] for k in odds}
        margin = total_impl - 100
        labels = {"t1": t1n, "t2": t2n, "tie": "Tie"}
        prob_text = " \u00b7 ".join(f"{labels[k]}: {probs[k]:.1f}%" for k in ["t1", "t2", "tie"])
        st.markdown(
            f'<div style="text-align:center;font-family:Rajdhani;font-size:13px;color:{MUTED};">'
            f'Implied: {prob_text} \u00b7 Overround: {margin:.1f}%</div>',
            unsafe_allow_html=True,
        )

        # Arbitrage alert
        arb = detect_arbitrage(odds)
        if arb["is_arb"]:
            st.markdown(f"""
            <div class="dash-card" style="border-top:3px solid {GREEN};padding:14px 18px;">
                <div style="font-family:Orbitron,monospace;font-size:14px;font-weight:600;color:{GREEN};letter-spacing:1px;">
                    \U0001f4b0 ARBITRAGE DETECTED \u2014 Negative overround: {arb['overround_pct']:.2f}%
                </div>
                <div style="font-family:Rajdhani;font-size:15px;color:{TEXT};margin-top:6px;">
                    Per \u20b91,000 total stake: {labels['t1']} {fmt_inr(arb['stakes_per_1000']['t1'])} \u00b7
                    {labels['t2']} {fmt_inr(arb['stakes_per_1000']['t2'])} \u00b7
                    Tie {fmt_inr(arb['stakes_per_1000']['tie'])} \u2192
                    Guaranteed profit: <strong style="color:{GREEN};">{fmt_inr(arb['guaranteed_profit_per_1000'])}</strong>
                </div>
            </div>
            """, unsafe_allow_html=True)
        elif margin < 2.0:
            st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-size:12px;color:{AMBER};">\u26a1 Low overround ({margin:.1f}%) \u2014 close to fair odds, good entry point</div>', unsafe_allow_html=True)

        # Odds momentum (if we have snapshots)
        momentum = get_odds_momentum()
        if momentum:
            parts = []
            for sc, nm in [("t1", t1n), ("t2", t2n)]:
                if sc in momentum:
                    m = momentum[sc]
                    arrow = "\u25b2" if m["direction"] == "shortening" else ("\u25bc" if m["direction"] == "drifting" else "\u2013")
                    color = GREEN if m["direction"] == "shortening" else (RED if m["direction"] == "drifting" else MUTED)
                    parts.append(f'<span style="color:{color};">{nm}: {arrow}{abs(m["total_shift"]):.1f}pp ({m["odds_start"]:.2f}\u2192{m["odds_now"]:.2f})</span>')
            if parts:
                snap_count = len(st.session_state.get("odds_snapshots", []))
                st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:4px;">Momentum ({snap_count} snapshots): {" \u00b7 ".join(parts)}</div>', unsafe_allow_html=True)


def render_pre_existing_panel():
    """Full pre-existing panel with per-bet edge shift cards (restored from v3)."""
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"

    with st.expander(
        f"\U0001f4e6 Import Pre-Existing Bets ({len(st.session_state.pre_bets)} loaded)",
        expanded=len(st.session_state.pre_bets) == 0 and len(st.session_state.bets) == 0,
    ):
        st.markdown(
            f'<div style="font-family:Rajdhani;font-size:15px;color:{MUTED};margin-bottom:12px;">'
            f'Already placed bets before opening this app? Enter them at their '
            f'<strong style="color:{VIOLET};">original entry odds</strong>. '
            f'These are already part of your session capital above.</div>',
            unsafe_allow_html=True,
        )

        with st.form("pre_bet_form", clear_on_submit=True):
            pc1, pc2, pc3 = st.columns([1.2, 1, 1])
            with pc1:
                pre_outcome = st.selectbox("Outcome You Backed", [f"{t1n} Win", f"{t2n} Win", "Tie"], key="pre_o_sel")
            with pc2:
                pre_odds = st.number_input("Entry Odds (at time of bet)", min_value=1.01, step=0.05, format="%.2f", value=1.80, key="pre_o_inp")
            with pc3:
                pre_stake = st.number_input("Stake (\u20b9)", min_value=10.0, step=100.0, format="%.0f", value=5000.0, key="pre_s_inp")

            if st.form_submit_button("\U0001f4e6 Import This Bet", use_container_width=True, type="primary"):
                out_key = "t1" if pre_outcome == f"{t1n} Win" else ("t2" if pre_outcome == f"{t2n} Win" else "tie")
                staked = compute_total_staked(get_all_bets_including_misc())
                rem = max(0, st.session_state.total_capital - staked)
                if pre_stake > rem + 0.01:
                    st.error(f"Insufficient capital. Remaining: {fmt_inr(rem)}")
                elif pre_odds <= 1.0:
                    st.error("Entry odds must be > 1.0")
                else:
                    st.session_state.pre_bet_counter += 1
                    st.session_state.pre_bets.append({
                        "id": f"PRE-{st.session_state.pre_bet_counter}",
                        "time_label": f"Pre-Match #{st.session_state.pre_bet_counter}",
                        "outcome": out_key, "odds": pre_odds, "stake": pre_stake,
                        "source": "pre-existing", "timestamp": "Before app",
                    })
                    sync_match_to_history(); st.rerun()

        # Per-bet edge shift cards (restored from v3)
        if st.session_state.pre_bets:
            odds = get_current_odds()
            has_odds = all(odds[k] > 1.0 for k in odds)
            name_map = {"t1": t1n, "t2": t2n, "tie": "Tie"}

            st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{VIOLET};letter-spacing:2px;margin:12px 0 8px 0;">IMPORTED POSITIONS</div>', unsafe_allow_html=True)

            for b in st.session_state.pre_bets:
                out_color = OUTCOME_COLORS.get(b["outcome"], TEXT)
                shift_html = ""
                if has_odds:
                    es = compute_edge_shift(b["odds"], odds, b["outcome"])
                    if es:
                        if es["direction"] == "favourable":
                            pill_class = "edge-fav"; shift_text = f'\u25b2 {fmt_pct(es["edge_shift"])} in your favour'
                            detail = f'Odds: {b["odds"]:.2f} \u2192 {es["odds_now"]:.2f} (shortened)'
                        elif es["direction"] == "against":
                            pill_class = "edge-bad"; shift_text = f'\u25bc {fmt_pct(es["edge_shift"])} against you'
                            detail = f'Odds: {b["odds"]:.2f} \u2192 {es["odds_now"]:.2f} (drifted)'
                        else:
                            pill_class = "edge-flat"; shift_text = "No shift"
                            detail = f'Odds unchanged: {b["odds"]:.2f}'
                        shift_html = (
                            f'<div style="margin-top:6px;">'
                            f'<span class="edge-pill {pill_class}">{shift_text}</span>'
                            f'<span style="font-size:12px;color:{MUTED};margin-left:8px;">{detail}</span>'
                            f'</div>'
                        )

                st.markdown(f"""
                <div class="dash-card" style="padding:14px 18px;border-left:3px solid {VIOLET};">
                    <div style="display:flex;justify-content:space-between;align-items:center;">
                        <div>
                            <span style="color:{out_color};font-weight:700;font-size:16px;">
                                {name_map[b['outcome']]}
                            </span>
                            <span class="pre-badge">PRE-EXISTING</span>
                        </div>
                        <div style="text-align:right;font-family:Orbitron,monospace;">
                            <div style="font-size:14px;color:{TEXT};">{fmt_inr(b['stake'])}</div>
                            <div style="font-size:11px;color:{MUTED};">@ {b['odds']:.2f}</div>
                        </div>
                    </div>
                    {shift_html}
                </div>
                """, unsafe_allow_html=True)

            if st.button("\u21a9\ufe0f Remove Last Imported Bet", key="undo_pre_panel"):
                st.session_state.pre_bets.pop(); sync_match_to_history(); st.rerun()
            # Per-item delete for imported bets
            if len(st.session_state.pre_bets) > 1:
                with st.expander("\U0001f5d1 Delete a specific imported bet"):
                    for idx_pb, pb in enumerate(st.session_state.pre_bets):
                        nm_pb = name_map.get(pb.get("outcome", ""), "?")
                        dc1, dc2 = st.columns([4, 1])
                        dc1.write(f"{nm_pb} \u2022 {fmt_inr(pb['stake'])} @ {pb['odds']:.2f}")
                        if dc2.button("\U0001f5d1", key=f"del_imp_{idx_pb}"):
                            st.session_state.pre_bets.pop(idx_pb); sync_match_to_history(); st.rerun()


def render_edge_shift_summary():
    """Aggregate odds-shift advantage card (restored from v3)."""
    pre_bets = st.session_state.pre_bets
    odds = get_current_odds()
    if not pre_bets or not all(odds[k] > 1.0 for k in odds):
        return
    pre_pnl = compute_pnl(pre_bets)
    hypo_bets = [{"outcome": b["outcome"], "odds": odds[b["outcome"]], "stake": b["stake"]} for b in pre_bets]
    hypo_pnl = compute_pnl(hypo_bets)
    net_edge = sum(pre_pnl[sc] - hypo_pnl[sc] for sc in pre_pnl) / 3
    if abs(net_edge) < 1:
        return
    edge_color = GREEN if net_edge > 0 else RED
    edge_icon = "\U0001f4c8" if net_edge > 0 else "\U0001f4c9"
    direction = "gained" if net_edge > 0 else "lost"
    cheaper_text = (
        "Hedging now will be <strong>cheaper</strong> than if you'd tried at entry."
        if net_edge > 0 else
        "Hedging now costs <strong>more</strong> \u2014 but the engine finds the best path regardless."
    )
    st.markdown(f"""
    <div class="dash-card" style="border-top:3px solid {edge_color};padding:16px 22px;">
        <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;">
            <span style="font-size:22px;">{edge_icon}</span>
            <span style="font-family:Orbitron,monospace;font-size:14px;font-weight:600;color:{edge_color};letter-spacing:1px;">
                ODDS SHIFT ADVANTAGE
            </span>
        </div>
        <div style="font-family:Rajdhani;font-size:16px;color:{TEXT};">
            Since you placed your pre-existing bets, odds movement has
            <strong style="color:{edge_color};">{direction}</strong> you an average of
            <strong style="color:{edge_color};font-family:Orbitron,monospace;">{fmt_inr(abs(net_edge))}</strong>
            per scenario in edge value. {cheaper_text}
        </div>
    </div>
    """, unsafe_allow_html=True)


# ── Render: P&L, Recommendation, Bet Forms, Ledger, Chart ───

def render_pnl_metrics(pnl):
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    mp = compute_misc_pnl()
    has_misc = mp["count"] > 0

    st.markdown('<div class="section-title">\U0001f4ca  Scenario P&L</div>', unsafe_allow_html=True)
    labels = {"t1": t1n + " Wins", "t2": t2n + " Wins", "tie": "Tie"}
    colors = {"t1": GREEN, "t2": CYAN, "tie": AMBER}
    cols = st.columns(3)
    for i, sc in enumerate(["t1", "t2", "tie"]):
        main_val = pnl[sc]
        # Combined: main P&L + misc realized (already settled) + misc active best case
        combined_best = main_val + mp["best"] if has_misc else main_val
        combined_worst = main_val + mp["worst"] if has_misc else main_val
        # Show the main value prominently, with misc range if applicable
        val_color = GREEN if main_val > 0 else (RED if main_val < 0 else MUTED)
        delta_icon = "\u25b2" if main_val > 0 else ("\u25bc" if main_val < 0 else "\u2013")
        misc_line = ""
        if has_misc:
            cb_color = GREEN if combined_best > 0 else RED
            cw_color = GREEN if combined_worst > 0 else RED
            misc_line = (
                f'<div style="font-family:Rajdhani;font-size:11px;color:{MUTED};margin-top:4px;border-top:1px solid {CARD_BORDER};padding-top:4px;">'
                f'Main: {fmt_inr(main_val)} '
                f'+ Misc: {fmt_inr(mp["realized"])}'
                + (f' to {fmt_inr(mp["best"] if mp["active_count"] > 0 else mp["realized"])}' if mp["active_count"] > 0 else '')
                + f'<br>Total: <span style="color:{cw_color};">{fmt_inr(combined_worst)}</span>'
                + (f' to <span style="color:{cb_color};">{fmt_inr(combined_best)}</span>' if mp["active_count"] > 0 else '')
                + f'</div>'
            )
        with cols[i]:
            st.markdown(
                f'<div class="dash-card" style="border-top:3px solid {colors[sc]};">'
                f'<div class="metric-box">'
                f'<div class="metric-label">{labels[sc]}</div>'
                f'<div class="metric-value" style="color:{val_color};">{fmt_inr(main_val)}</div>'
                f'<div class="metric-delta" style="color:{val_color};">{delta_icon} {"Profit" if main_val >= 0 else "Loss"}</div>'
                f'{misc_line}'
                f'</div></div>',
                unsafe_allow_html=True,
            )

    # Misc summary line
    if has_misc:
        realized_text = f' \u00b7 Realized: <span style="color:{GREEN if mp["realized"]>=0 else RED};font-family:Orbitron,monospace;">{fmt_inr(mp["realized"])}</span>' if mp["realized"] != 0 else ""
        active_text = f'{mp["active_count"]} active' if mp["active_count"] > 0 else "all settled"
        st.markdown(
            f'<div style="text-align:center;font-family:Rajdhani;font-size:13px;color:{MISC_PURPLE};margin-top:-8px;margin-bottom:8px;">'
            f'\U0001f3b2 {mp["count"]} misc bet(s) ({active_text}): best {fmt_inr(mp["best"])}, worst {fmt_inr(mp["worst"])}{realized_text}</div>',
            unsafe_allow_html=True,
        )

    # Warning: all scenarios negative (use combined worst)
    combined_worsts = {sc: pnl[sc] + mp["worst"] for sc in pnl} if has_misc else pnl
    if all(combined_worsts[sc] < 0 for sc in combined_worsts) and any(pnl[sc] != 0 for sc in pnl):
        odds = get_current_odds()
        overround = sum(100.0 / odds[k] for k in odds) - 100 if all(odds[k] > 1 for k in odds) else 0
        st.markdown(
            f'<div class="dash-card" style="border-top:3px solid {RED};padding:14px 18px;">'
            f'<div style="font-family:Orbitron,monospace;font-size:13px;font-weight:600;color:{RED};letter-spacing:1px;">'
            f'\u26a0\ufe0f ALL SCENARIOS SHOW LOSS'
            f'</div>'
            f'<div style="font-family:Rajdhani;font-size:15px;color:{TEXT};margin-top:6px;">'
            f'Combined main + misc worst-case is negative in all outcomes. '
            f'Overround: <strong>{overround:.1f}%</strong>. '
            f'This is normal if building a position to hedge later at better odds.'
            f'</div></div>',
            unsafe_allow_html=True,
        )


def render_recommendation_card(rec, learnings=None):
    st.markdown('<div class="section-title">\U0001f9e0  Recommendation</div>', unsafe_allow_html=True)
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    name_map = {"t1": t1n, "t2": t2n, "tie": "Tie"}

    # ── Goal Context Card ──
    gt = compute_goals_tracker()
    g = st.session_state.history.get("goals", {})
    session_target = g.get("session_target", gt["per_match_profit"])
    session_floor = g.get("session_min_acceptable", 0)
    session_strat = g.get("session_strategy", "conviction")
    strat_labels = {"conviction": "\U0001f3af Conviction", "hedge_first": "\U0001f6e1 Hedge First", "aggressive": "\U0001f525 Aggressive", "conservative": "\U0001f9ca Conservative"}
    sl = compute_sizing_limits()

    if session_target > 0:
        season_text = ""
        if gt.get("target_achieved"):
            season_text = f'\U0001f3c6 Season target hit! Keep compounding.'
            pace_color = GREEN
        elif gt["on_pace"]:
            season_text = f'On pace \u2714 \u00b7 {gt["matches_played"]}/{gt["total_matches"]} done'
            pace_color = GREEN
        else:
            season_text = f'Behind pace \u2014 push harder'
            pace_color = AMBER

        st.markdown(
            f'<div class="dash-card" style="border-left:4px solid {CYAN};padding:14px 18px;">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;">'
            f'<div>'
            f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">SESSION TARGET</div>'
            f'<div style="font-family:Orbitron,monospace;font-size:24px;font-weight:700;color:{CYAN};">{fmt_inr(session_target)}</div>'
            f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};">'
            f'Floor: {fmt_inr(session_floor)} \u00b7 {strat_labels.get(session_strat, session_strat)}</div>'
            f'</div>'
            f'<div style="text-align:right;">'
            f'<div style="font-family:Rajdhani;font-size:13px;color:{pace_color};font-weight:600;">{season_text}</div>'
            f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};">Anchor max: {fmt_inr(sl["max_anchor_amt"])} ({sl["max_anchor_pct"]*100:.0f}%)</div>'
            f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};">Session max: {fmt_inr(sl["max_session_amt"])} ({sl["max_session_pct"]*100:.0f}%)</div>'
            f'</div></div></div>',
            unsafe_allow_html=True,
        )

    border_color = rec.get("color", MUTED)
    stakes_html = ""
    if rec.get("stakes"):
        stake_lines = []
        for outcome, stake in rec["stakes"].items():
            if stake >= 10:
                odds_val = get_current_odds().get(outcome, 0)
                potential_profit = stake * (odds_val - 1) if odds_val > 1 else 0
                stake_lines.append(
                    f'<div style="margin:4px 0;color:{OUTCOME_COLORS.get(outcome, TEXT)};">'
                    f'\u279c <strong>{name_map.get(outcome, outcome)}</strong>: '
                    f'{fmt_inr(stake)} @ {odds_val:.2f}'
                    f' \u2192 profit if wins: <strong style="color:{GREEN};">+{fmt_inr(potential_profit)}</strong></div>'
                )
        if stake_lines:
            stakes_html = f'<div class="rec-stakes">{"".join(stake_lines)}</div>'
    new_pnl_html = ""
    if "new_pnl" in rec:
        np_ = rec["new_pnl"]
        pnl_parts = [f'<span style="color:{GREEN if np_[sc] >= 0 else RED};">{name_map[sc]}: {fmt_inr(np_[sc])}</span>' for sc in ["t1", "t2", "tie"]]
        new_pnl_html = f'<div style="margin-top:10px;font-size:13px;color:{MUTED};">Projected P&L after hedge: {" \u00b7 ".join(pnl_parts)}</div>'
        mp = compute_misc_pnl()
        if mp["count"] > 0:
            min_hedge = min(np_[sc] for sc in np_)
            combined_worst = min_hedge + mp["worst"]
            combined_best = min_hedge + mp["best"]
            cw_c = GREEN if combined_worst >= 0 else RED
            cb_c = GREEN if combined_best >= 0 else RED
            new_pnl_html += (
                f'<div style="font-size:12px;color:{MISC_PURPLE};margin-top:4px;">'
                f'\U0001f3b2 With misc bets: worst <span style="color:{cw_c};">{fmt_inr(combined_worst)}</span>'
                f' to best <span style="color:{cb_c};">{fmt_inr(combined_best)}</span></div>'
            )
        # Goal context on profit
        if session_target > 0:
            min_pnl_val = min(np_[sc] for sc in np_)
            max_pnl_val = max(np_[sc] for sc in np_)
            if min_pnl_val >= session_target:
                new_pnl_html += f'<div style="font-size:12px;color:{GREEN};margin-top:4px;">\u2705 Guaranteed to hit session target of {fmt_inr(session_target)}</div>'
            elif max_pnl_val >= session_target:
                new_pnl_html += f'<div style="font-size:12px;color:{AMBER};margin-top:4px;">\U0001f3af Best case hits target ({fmt_inr(max_pnl_val)} vs {fmt_inr(session_target)} target)</div>'
            elif min_pnl_val >= session_floor:
                new_pnl_html += f'<div style="font-size:12px;color:{AMBER};margin-top:4px;">\U0001f6e1 Above floor ({fmt_inr(min_pnl_val)} vs {fmt_inr(session_floor)} floor) but below target. Conviction mode for more upside.</div>'
            else:
                new_pnl_html += f'<div style="font-size:12px;color:{RED};margin-top:4px;">\u26a0 Below floor ({fmt_inr(min_pnl_val)} vs {fmt_inr(session_floor)}). Need better entry or reduce position.</div>'
    badge_html = ""
    if learnings:
        badge_html = get_historical_insight_badge(learnings, rec.get("action", ""), st.session_state.match_phase)
        if badge_html:
            badge_html = f'<div style="margin-top:10px;">{badge_html}</div>'
    st.markdown(
        f'<div class="rec-card" style="border-color:{border_color};">'
        f'<div style="font-size:28px; margin-bottom:4px;">{rec.get("icon", "")}</div>'
        f'<div class="rec-headline" style="color:{border_color};">{rec.get("headline", "")}</div>'
        f'<div class="rec-detail">{rec.get("detail", "")}</div>'
        f'{stakes_html}{new_pnl_html}{badge_html}'
        f'</div>',
        unsafe_allow_html=True,
    )

    # ── Hedge Scenarios Table (when bets exist) ──
    all_bets = get_all_bets()
    odds = get_current_odds()
    if all_bets and all(odds[k] > 1 for k in odds):
        pnl = compute_pnl(all_bets)
        remaining_cap = max(0, st.session_state.get("total_capital", 10000) + st.session_state.get("session_fund_adj", 0) - compute_total_staked(get_all_bets_including_misc()))

        # Find user's side (best outcome)
        best_sc = max(pnl, key=pnl.get)
        worst_sc = min(pnl, key=pnl.get)
        best_name = name_map.get(best_sc, "?")

        with st.expander(f"\U0001f4ca Hedge Scenarios — What If You Hedge Now?", expanded=False):
            # Build scenarios at different hedge odds
            fav_odds = min(odds["t1"], odds["t2"])
            dog_odds = max(odds["t1"], odds["t2"])

            shifts = [
                ("Current odds", 1.0),
                (f"{best_name} pulls ahead (+10%)", 0.90),
                (f"{best_name} dominating (+20%)", 0.80),
                ("Odds tighten (-10%)", 1.10),
                ("Odds drift against (-20%)", 1.20),
            ]

            rows = ""
            for label, factor in shifts:
                shifted_odds = dict(odds)
                for k in ["t1", "t2"]:
                    if k == best_sc:
                        shifted_odds[k] = max(1.01, odds[k] * factor)
                    else:
                        shifted_odds[k] = max(1.01, odds[k] * (2 - factor))

                sol = solve_optimal_hedge(pnl, shifted_odds, remaining_cap)
                if sol:
                    lock_in = sol["min_pnl"]
                    hedge_cost = sol["total_stake"]
                    # Profit on each side after hedge
                    p_t1 = sol["new_pnl"]["t1"]
                    p_t2 = sol["new_pnl"]["t2"]
                    t1c = GREEN if p_t1 >= 0 else RED
                    t2c = GREEN if p_t2 >= 0 else RED
                    lc = GREEN if lock_in >= 0 else RED

                    # Goal check
                    if lock_in >= session_target:
                        action = f'<span style="color:{GREEN};">\u2705 Hits target</span>'
                    elif lock_in >= session_floor:
                        action = f'<span style="color:{AMBER};">Above floor</span>'
                    else:
                        action = f'<span style="color:{RED};">Below floor</span>'

                    rows += (
                        f'<tr>'
                        f'<td style="font-family:Rajdhani;font-size:13px;">{label}</td>'
                        f'<td style="font-family:Orbitron,monospace;font-size:12px;">{fmt_inr(hedge_cost)}</td>'
                        f'<td style="font-family:Orbitron,monospace;font-size:12px;color:{t1c};">{fmt_inr(p_t1)}</td>'
                        f'<td style="font-family:Orbitron,monospace;font-size:12px;color:{t2c};">{fmt_inr(p_t2)}</td>'
                        f'<td style="font-family:Orbitron,monospace;font-size:12px;color:{lc};">{fmt_inr(lock_in)}</td>'
                        f'<td>{action}</td>'
                        f'</tr>'
                    )

            if rows:
                st.markdown(
                    f'<div class="dash-card"><table class="ledger-table">'
                    f'<thead><tr><th>Scenario</th><th>Hedge Cost</th><th>{t1n} Wins</th><th>{t2n} Wins</th><th>Lock-in</th><th>vs Goal</th></tr></thead>'
                    f'<tbody>{rows}</tbody></table></div>',
                    unsafe_allow_html=True,
                )
                st.markdown(
                    f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:4px;">'
                    f'Lock-in = guaranteed minimum P&L regardless of result. '
                    f'Target: {fmt_inr(session_target)} \u00b7 Floor: {fmt_inr(session_floor)} \u00b7 '
                    f'Remaining capital: {fmt_inr(remaining_cap)}</div>',
                    unsafe_allow_html=True,
                )

    # ── Bet Sizing Guide ──
    render_sizing_guide(rec, learnings)

    # Timing guide
    render_timing_guide()


def compute_sizing_limits():
    """Single source of truth for all bet sizing limits based on portfolio mode."""
    cap = get_portfolio_capital()
    avail = get_available_capital()
    pmode, _, _ = get_portfolio_mode()
    odds = get_current_odds()

    if pmode == "PRESERVATION":
        max_session_pct, max_anchor_pct, max_total_pct = 0.08, 0.0, 0.15
    elif pmode == "STANDARD":
        max_session_pct, max_anchor_pct, max_total_pct = 0.15, 0.12, 0.30
    else:  # GROWTH
        max_session_pct, max_anchor_pct, max_total_pct = 0.25, 0.18, 0.40

    # Kelly adjustment
    if all(odds.get(k, 0) > 1 for k in ["t1", "t2", "tie"]):
        best_outcome = min(odds, key=odds.get)
        best_impl = 100.0 / odds[best_outcome]
        raw_impl_sum = sum(100.0 / odds[k] for k in odds)
        fair_prob = best_impl / raw_impl_sum * 100
        kelly = compute_kelly_fraction(fair_prob, odds[best_outcome])
        if kelly > 0.01 and kelly < max_anchor_pct:
            max_anchor_pct = kelly
    else:
        kelly = 0

    max_anchor_amt = round(avail * max_anchor_pct / 10) * 10
    max_session_amt = round(avail * max_session_pct / 10) * 10
    max_total_amt = round(cap * max_total_pct / 10) * 10

    return {
        "max_anchor_pct": max_anchor_pct, "max_anchor_amt": max_anchor_amt,
        "max_session_pct": max_session_pct, "max_session_amt": max_session_amt,
        "max_total_pct": max_total_pct, "max_total_amt": max_total_amt,
        "pmode": pmode, "kelly": kelly, "cap": cap, "avail": avail,
    }


def render_sizing_guide(rec, learnings):
    """Capital allocation and bet sizing guidance based on portfolio state."""
    sl = compute_sizing_limits()
    cap, avail = sl["cap"], sl["avail"]
    sm_exp = get_standalone_exposure()
    pmode = sl["pmode"]
    pmc = GREEN if pmode == "GROWTH" else (AMBER if pmode == "STANDARD" else RED)
    odds = get_current_odds()
    all_bets = get_all_bets()
    staked = compute_total_staked(all_bets)
    session_cap = st.session_state.get("total_capital", avail)

    if not all(odds[k] > 1 for k in odds):
        return

    session_util = staked / session_cap if session_cap > 0 else 0

    max_session_pct = sl["max_session_pct"]
    max_anchor_pct = sl["max_anchor_pct"]
    max_total_pct = sl["max_total_pct"]
    kelly = sl["kelly"]
    max_session_amt = sl["max_session_amt"]
    max_anchor_amt = sl["max_anchor_amt"]
    max_total_amt = sl["max_total_amt"]

    if pmode == "PRESERVATION":
        mode_desc = "Drawdown >20%. No new anchors. Hedge-only mode."
    elif pmode == "STANDARD":
        mode_desc = "Moderate drawdown. Conservative sizing."
    else:
        mode_desc = "At or near peak. Standard sizing allowed."

    # How many more bets can they take?
    remaining_session_budget = max(0, max_session_amt - staked)
    live_bets_count = len(st.session_state.bets)
    pre_bets_count = len(st.session_state.pre_bets)
    total_bets = live_bets_count + pre_bets_count

    # Recommended number of live bets
    if session_util < 0.5:
        live_bet_advice = "Room for 2-3 more bets at current sizing."
    elif session_util < 0.8:
        live_bet_advice = "Room for 1 more bet. Consider hedging soon."
    else:
        live_bet_advice = "Session near capacity. Focus on hedging, not new bets."

    # Build the sizing card
    rows = ""
    rows += (
        f'<tr><td style="color:{GREEN};">Anchor Bet (max)</td>'
        f'<td style="font-family:Orbitron,monospace;">{fmt_inr(max_anchor_amt)}</td>'
        f'<td>{max_anchor_pct*100:.0f}% of available</td>'
        f'<td style="color:{MUTED};">{"Kelly-adjusted" if kelly > 0.01 else "Mode-based"}</td></tr>'
    )
    rows += (
        f'<tr><td style="color:{CYAN};">Session Total (max)</td>'
        f'<td style="font-family:Orbitron,monospace;">{fmt_inr(max_session_amt)}</td>'
        f'<td>{max_session_pct*100:.0f}% of available</td>'
        f'<td style="color:{MUTED};">All bets this match</td></tr>'
    )
    rows += (
        f'<tr><td style="color:{AMBER};">Total Portfolio (max)</td>'
        f'<td style="font-family:Orbitron,monospace;">{fmt_inr(max_total_amt)}</td>'
        f'<td>{max_total_pct*100:.0f}% of capital</td>'
        f'<td style="color:{MUTED};">Session + side bets</td></tr>'
    )
    rows += (
        f'<tr style="border-top:2px solid {CARD_BORDER};">'
        f'<td style="font-weight:700;">Currently Deployed</td>'
        f'<td style="font-family:Orbitron,monospace;font-weight:700;">{fmt_inr(staked)}</td>'
        f'<td>{session_util*100:.0f}% of session limit</td>'
        f'<td style="color:{GREEN if session_util < 0.8 else RED};">{"OK" if session_util < 0.8 else "Near limit"}</td></tr>'
    )
    rows += (
        f'<tr><td style="font-weight:700;">Budget Remaining</td>'
        f'<td style="font-family:Orbitron,monospace;font-weight:700;color:{GREEN if remaining_session_budget > 0 else RED};">{fmt_inr(remaining_session_budget)}</td>'
        f'<td>for new bets</td>'
        f'<td style="color:{MUTED};">{live_bet_advice}</td></tr>'
    )

    # Pre-bet sizing advice
    pre_advice = ""
    if pre_bets_count > 0 and live_bets_count == 0:
        pre_advice = f'<div style="margin-top:8px;font-size:13px;color:{VIOLET};">\u25a0 You have {pre_bets_count} pre-match bet(s). Live bets should hedge or complement these \u2014 not duplicate the same outcome.</div>'
    elif pre_bets_count > 0 and live_bets_count > 0:
        pre_pct = compute_total_staked(st.session_state.pre_bets) / max(staked, 1) * 100
        live_pct = 100 - pre_pct
        pre_advice = f'<div style="margin-top:8px;font-size:13px;color:{MUTED};">Split: {pre_pct:.0f}% pre-match \u00b7 {live_pct:.0f}% live. {"Good balance." if 30 < pre_pct < 70 else "Consider diversifying entry timing."}</div>'

    with st.expander(f"\U0001f4b0 Sizing Guide \u2014 {pmode} Mode", expanded=False):
        st.markdown(
            f'<div style="font-family:Rajdhani;font-size:13px;color:{pmc};margin-bottom:8px;">{mode_desc}</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div class="dash-card"><table class="ledger-table">'
            '<thead><tr><th>Category</th><th>Amount</th><th>% Rule</th><th>Basis</th></tr></thead>'
            '<tbody>' + rows + '</tbody></table></div>',
            unsafe_allow_html=True,
        )
        if pre_advice:
            st.markdown(pre_advice, unsafe_allow_html=True)

        # Number of bets guidance
        st.markdown(
            f'<div style="font-family:Rajdhani;font-size:14px;color:{TEXT};margin-top:10px;">'
            f'<strong>Bet count guidance:</strong> '
            f'Optimal is 2-4 bets per session (1 anchor + 1-3 hedges). '
            f'You have {total_bets} bet(s). '
            f'{"Add hedges to balance your position." if total_bets == 1 else "Good count." if total_bets <= 4 else "Many bets \u2014 risk of overtrading. Each new bet should have clear purpose."}'
            f'</div>',
            unsafe_allow_html=True,
        )



def render_timing_guide():
    """Hedge timing guide — what to do at different odds scenarios."""
    odds = get_current_odds()
    if not all(odds[k] > 1 for k in odds):
        return
    all_bets = get_all_bets()
    if not all_bets:
        return

    pnl = compute_pnl(all_bets)
    phase = st.session_state.match_phase
    t1n = st.session_state.t1_name or "T1"
    t2n = st.session_state.t2_name or "T2"
    remaining = max(0, st.session_state.get("total_capital", 10000) - compute_total_staked(get_all_bets_including_misc()))

    # Find which side user is on
    best_sc = max(pnl, key=pnl.get)
    worst_sc = min(pnl, key=pnl.get)
    best_name = {"t1": t1n, "t2": t2n, "tie": "Tie"}[best_sc]
    worst_name = {"t1": t1n, "t2": t2n, "tie": "Tie"}[worst_sc]

    g_goals = st.session_state.history.get("goals", {})
    session_tgt = g_goals.get("session_target", 500)
    session_floor = g_goals.get("session_min_acceptable", 0)

    # Current hedge value
    sol = solve_optimal_hedge(pnl, odds, remaining)
    current_lock = sol["min_pnl"] if sol else min(pnl.values())

    with st.expander(f"\u23f0 Timing Guide \u2014 When to Hedge", expanded=False):
        # Phase-based guidance
        phase_rows = ""
        phases = [
            ("Early (Overs 1-6)", "Odds shift 20-40% in powerplay. Unless your team collapses, WAIT. Your entry edge grows as odds move.", MUTED),
            ("Middle (Overs 7-14)", "Best window to hedge. Odds are more settled. If your team is ahead, hedge NOW to lock value.", CYAN),
            ("Last 5 Overs (15-20)", "Odds nearly final. Hedge IMMEDIATELY if you haven't. Every over you wait, your window shrinks.", AMBER),
        ]
        for label, advice, color in phases:
            is_current = (label.startswith("Early") and phase == "Early") or (label.startswith("Middle") and phase == "Middle") or (label.startswith("Last") and phase == "Last 5 Overs")
            marker = "\u25b6 " if is_current else ""
            weight = "font-weight:700;" if is_current else ""
            phase_rows += (
                f'<tr><td style="color:{color};{weight}">{marker}{label}</td>'
                f'<td style="font-family:Rajdhani;{weight}">{advice}</td></tr>'
            )

        st.markdown(
            '<div class="dash-card"><table class="ledger-table">'
            '<thead><tr><th>Phase</th><th>Guidance</th></tr></thead>'
            '<tbody>' + phase_rows + '</tbody></table></div>',
            unsafe_allow_html=True,
        )

        # Odds scenario projections
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{CYAN};letter-spacing:2px;margin:12px 0 8px;">IF ODDS MOVE...</div>', unsafe_allow_html=True)

        scenarios = []
        for desc, best_shift, worst_shift in [
            (f"{best_name} dominating (+15%)", 0.85, 1.20),
            (f"{best_name} ahead (+10%)", 0.90, 1.12),
            (f"{best_name} slight edge (+5%)", 0.95, 1.06),
            ("No change", 1.0, 1.0),
            (f"{worst_name} fights back (-5%)", 1.06, 0.95),
            (f"{worst_name} taking over (-10%)", 1.12, 0.90),
            (f"{worst_name} dominating (-15%)", 1.20, 0.85),
        ]:
            shifted = dict(odds)
            shifted[best_sc] = max(1.01, odds[best_sc] * best_shift)
            shifted[worst_sc] = max(1.01, odds[worst_sc] * worst_shift)
            shifted_sol = solve_optimal_hedge(pnl, shifted, remaining)
            if shifted_sol:
                lock_val = shifted_sol["min_pnl"]
                diff = lock_val - current_lock
                action = ""
                if lock_val >= session_tgt:
                    action = f"\u2705 Hedge NOW \u2014 locks {fmt_inr(lock_val)} (hits target!)"
                elif lock_val >= session_floor and lock_val > current_lock:
                    action = f"\U0001f3af Hedge \u2014 locks {fmt_inr(lock_val)} (above floor)"
                elif lock_val >= session_floor:
                    action = f"\U0001f6e1 Can hedge \u2014 locks {fmt_inr(lock_val)}"
                elif lock_val < session_floor:
                    action = f"\u26a0 Wait \u2014 lock-in ({fmt_inr(lock_val)}) below floor"
                dc = GREEN if diff > 0 else (RED if diff < 0 else MUTED)
                scenarios.append((desc, lock_val, diff, action, dc))

        rows = ""
        for desc, lock_val, diff, action, dc in scenarios:
            lc = GREEN if lock_val >= 0 else RED
            rows += (
                f'<tr><td style="font-family:Rajdhani;font-size:13px;">{desc}</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:12px;color:{lc};">{fmt_inr(lock_val)}</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:12px;color:{dc};">{("+" if diff>=0 else "")}{fmt_inr(diff)}</td>'
                f'<td style="font-family:Rajdhani;font-size:12px;">{action}</td></tr>'
            )
        st.markdown(
            '<div class="dash-card"><table class="ledger-table">'
            '<thead><tr><th>Scenario</th><th>Lock-in</th><th>vs Now</th><th>Action</th></tr></thead>'
            '<tbody>' + rows + '</tbody></table></div>',
            unsafe_allow_html=True,
        )

        # Current position summary
        st.markdown(
            f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-top:8px;">'
            f'Current lock-in: <strong style="color:{GREEN if current_lock>=0 else RED};">{fmt_inr(current_lock)}</strong> '
            f'\u00b7 Target: {fmt_inr(session_tgt)} \u00b7 Floor: {fmt_inr(session_floor)} '
            f'\u00b7 Your side: {best_name} (best: {fmt_inr(pnl[best_sc])})'
            f'</div>',
            unsafe_allow_html=True,
        )


def render_add_bet_form(remaining_capital):
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    st.markdown('<div class="section-title">\u2795  Add Bet</div>', unsafe_allow_html=True)
    with st.form("add_bet_form", clear_on_submit=True):
        fc0, fc1, fc2, fc3, fc4 = st.columns([0.8, 1.2, 1, 1, 1])
        with fc0:
            bet_type = st.selectbox("Type", ["Live", "Pre-Match"], key="bet_type_sel")
        with fc1:
            time_label = st.text_input("Time / Over", placeholder="e.g. Over 6.2")
        with fc2:
            outcome = st.selectbox("Outcome", [f"{t1n} Win", f"{t2n} Win", "Tie"])
        with fc3:
            odds_input = st.number_input("Odds (decimal)", min_value=1.01, step=0.05, format="%.2f", value=1.50)
        with fc4:
            stake_input = st.number_input("Stake (\u20b9)", min_value=10.0, step=100.0, format="%.0f", value=1000.0)
        if st.form_submit_button("Place Bet", use_container_width=True, type="primary"):
            out_key = "t1" if outcome == f"{t1n} Win" else ("t2" if outcome == f"{t2n} Win" else "tie")
            # Risk guard check — FUSES before insufficient-capital check
            allowed, blockers = risk_check_all_gates(stake_input, is_misc=False)
            if not allowed:
                for _, msg in blockers:
                    st.error(f"\U0001f6e1 BLOCKED: {msg}")
                st.info("Adjust stake, wait for cooldown, or override in sidebar \u2192 Risk Guards.")
            elif stake_input > remaining_capital + 0.01:
                st.error(f"Insufficient capital. Remaining: {fmt_inr(remaining_capital)}")
            elif odds_input <= 1.0:
                st.error("Odds must be greater than 1.0")
            else:
                is_pre = bet_type == "Pre-Match"
                if is_pre:
                    st.session_state.pre_bet_counter += 1
                    st.session_state.pre_bets.append({
                        "id": f"PRE-{st.session_state.pre_bet_counter}",
                        "time_label": time_label or f"Pre-Match #{st.session_state.pre_bet_counter}",
                        "outcome": out_key, "odds": odds_input, "stake": stake_input,
                        "source": "pre-existing", "timestamp": "Pre-Match",
                    })
                else:
                    st.session_state.bet_counter += 1
                    st.session_state.bets.append({
                        "id": st.session_state.bet_counter,
                        "time_label": time_label or f"Bet #{st.session_state.bet_counter}",
                        "outcome": out_key, "odds": odds_input, "stake": stake_input,
                        "source": "live", "timestamp": datetime.now().strftime("%H:%M:%S"),
                    })
                sync_match_to_history(); st.rerun()


def render_misc_bet_form(remaining_capital):
    """Form for miscellaneous / side bets with live settlement."""
    st.markdown(f'<div class="section-title" style="color:{MISC_PURPLE};">\U0001f3b2  Misc / Side Bets</div>', unsafe_allow_html=True)
    with st.form("misc_bet_form", clear_on_submit=True):
        m1, m2, m3 = st.columns([2, 1, 1])
        with m1:
            ml = st.text_input("Bet Label", placeholder="e.g. Top Scorer Kohli, 50+ in PP")
        with m2:
            mo = st.number_input("Odds", min_value=1.01, step=0.5, format="%.2f", value=3.00, key="misc_odds")
        with m3:
            ms = st.number_input("Stake (\u20b9)", min_value=10.0, step=100.0, format="%.0f", value=500.0, key="misc_stake")
        if st.form_submit_button("\U0001f3b2 Add Misc Bet", use_container_width=True):
            # Risk guard check — includes misc cap
            allowed, blockers = risk_check_all_gates(ms, is_misc=True)
            if not allowed:
                for _, msg in blockers:
                    st.error(f"\U0001f6e1 BLOCKED: {msg}")
                st.info("Adjust stake, wait for cooldown, or override in sidebar \u2192 Risk Guards.")
            elif ms > remaining_capital + 0.01:
                st.error(f"Insufficient capital. Remaining: {fmt_inr(remaining_capital)}")
            else:
                st.session_state.misc_bet_counter += 1
                st.session_state.misc_bets.append({
                    "id": f"MISC-{st.session_state.misc_bet_counter}",
                    "label": ml or f"Misc #{st.session_state.misc_bet_counter}",
                    "odds": mo, "stake": ms, "source": "misc",
                    "status": "active",  # active / settled
                    "settled_result": None,  # won / lost / void
                    "realized_pnl": 0,
                    "timestamp": datetime.now().strftime("%H:%M:%S"),
                })
                sync_match_to_history(); st.rerun()

    misc_bets = st.session_state.get("misc_bets", [])
    active_misc = [b for b in misc_bets if b.get("status", "active") == "active"]
    settled_misc = [b for b in misc_bets if b.get("status") == "settled"]

    # ── Active misc bets with Won/Lost buttons ──
    if active_misc:
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{MISC_PURPLE};letter-spacing:2px;margin:8px 0 6px;">ACTIVE ({len(active_misc)})</div>', unsafe_allow_html=True)
        for idx_m, mb in enumerate(misc_bets):
            if mb.get("status", "active") != "active":
                continue
            pot = fmt_inr(mb["stake"] * (mb["odds"] - 1))
            mc1, mc2, mc3, mc4, mc5 = st.columns([3, 0.7, 0.7, 0.7, 0.5])
            mc1.markdown(f'<div style="color:{MISC_PURPLE};font-family:Rajdhani;font-size:14px;font-weight:600;">{mb.get("label","Misc")} \u00b7 {fmt_inr(mb["stake"])} @ {mb["odds"]:.2f} \u00b7 pot <span style="color:{GREEN};">+{pot}</span></div>', unsafe_allow_html=True)
            if mc2.button("\u2705 Won", key=f"mw_{idx_m}"):
                misc_bets[idx_m]["status"] = "settled"
                misc_bets[idx_m]["settled_result"] = "won"
                misc_bets[idx_m]["realized_pnl"] = round(mb["stake"] * (mb["odds"] - 1), 2)
                sync_match_to_history(); st.rerun()
            if mc3.button("\u274c Lost", key=f"ml_{idx_m}"):
                misc_bets[idx_m]["status"] = "settled"
                misc_bets[idx_m]["settled_result"] = "lost"
                misc_bets[idx_m]["realized_pnl"] = -mb["stake"]
                sync_match_to_history(); st.rerun()
            if mc4.button("\u23ed Void", key=f"mv_{idx_m}"):
                misc_bets[idx_m]["status"] = "settled"
                misc_bets[idx_m]["settled_result"] = "void"
                misc_bets[idx_m]["realized_pnl"] = 0
                sync_match_to_history(); st.rerun()
            if mc5.button("\U0001f5d1", key=f"del_mf_{idx_m}"):
                st.session_state.misc_bets.pop(idx_m); sync_match_to_history(); st.rerun()

    # ── Settled misc bets summary ──
    if settled_misc:
        settled_pnl = sum(b.get("realized_pnl", 0) for b in settled_misc)
        pc = GREEN if settled_pnl >= 0 else RED
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;margin:12px 0 4px;">SETTLED ({len(settled_misc)})</div>', unsafe_allow_html=True)
        parts = []
        for b in settled_misc:
            rp = b.get("realized_pnl", 0)
            rc = GREEN if rp > 0 else (RED if rp < 0 else MUTED)
            result_icon = "\u2705" if b.get("settled_result") == "won" else ("\u274c" if b.get("settled_result") == "lost" else "\u23ed")
            parts.append(f'{result_icon} {b.get("label","?")} <span style="color:{rc};font-family:Orbitron,monospace;font-size:12px;">{fmt_inr(rp)}</span>')
        st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{TEXT};line-height:1.8;">{"<br>".join(parts)}</div>', unsafe_allow_html=True)
        st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;margin-top:4px;">Misc P&L: <strong style="color:{pc};font-family:Orbitron,monospace;">{fmt_inr(settled_pnl)}</strong></div>', unsafe_allow_html=True)


def render_bet_ledger(pnl):
    """Unified bet ledger — shows active bets only, settled misc shown separately."""
    all_bets = get_all_bets()
    active_misc = [b for b in st.session_state.get("misc_bets", []) if b.get("status", "active") == "active"]
    combined = all_bets + active_misc
    if not combined:
        st.markdown(f'<div style="text-align:center;padding:30px;color:{MUTED};font-family:Rajdhani;">No bets placed yet. Import a pre-existing bet or add a new one.</div>', unsafe_allow_html=True)
        return
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    name_map = {"t1": t1n, "t2": t2n, "tie": "Tie"}
    st.markdown('<div class="section-title">\U0001f4cb  Bet Ledger</div>', unsafe_allow_html=True)
    worst_sc = min(pnl, key=pnl.get) if any(pnl[k] < 0 for k in pnl) else None
    odds = get_current_odds()
    has_odds = all(odds[k] > 1.0 for k in odds)
    rows_html = ""
    for b in combined:
        is_misc = b.get("source") == "misc"
        is_pre = b.get("source") == "pre-existing"
        potential = fmt_inr(b["stake"] * (b["odds"] - 1))
        risk = fmt_inr(b["stake"])

        if is_misc:
            row_class = "ledger-row-misc"
            source_badge = '<span class="misc-badge">MISC</span>'
            out_color = MISC_PURPLE
            out_name = b.get("label", "Misc")
        elif is_pre:
            row_class = "ledger-row-pre"
            source_badge = '<span class="pre-badge">PRE</span>'
            out_color = OUTCOME_COLORS.get(b.get("outcome", ""), TEXT)
            out_name = name_map.get(b.get("outcome", ""), "?")
        else:
            row_class = "ledger-row-profit" if worst_sc and b.get("outcome") == worst_sc else ("ledger-row-loss" if worst_sc else "ledger-row-profit")
            source_badge = '<span class="live-badge">LIVE</span>'
            out_color = OUTCOME_COLORS.get(b.get("outcome", ""), TEXT)
            out_name = name_map.get(b.get("outcome", ""), "?")

        # Per-bet edge shift indicator (restored from v3)
        shift_cell = ""
        if is_pre and has_odds:
            es = compute_edge_shift(b["odds"], odds, b["outcome"])
            if es:
                sc = GREEN if es["direction"] == "favourable" else (RED if es["direction"] == "against" else MUTED)
                arrow = "\u25b2" if es["direction"] == "favourable" else ("\u25bc" if es["direction"] == "against" else "\u2013")
                shift_cell = f'<span style="color:{sc};font-family:Orbitron,monospace;font-size:11px;margin-left:6px;">{arrow}{fmt_pct(abs(es["edge_shift"]), signed=False)}</span>'

        time_label = b.get("time_label", b.get("label", ""))
        # Compute EV using overround-adjusted fair probability
        ev_cell = ""
        outcome = b.get("outcome", "")
        if not is_misc and outcome in odds and has_odds:
            raw_impl = {k: 100.0 / odds[k] for k in ["t1", "t2", "tie"]}
            impl_sum = sum(raw_impl.values())
            fair_prob = raw_impl.get(outcome, 0) / impl_sum * 100 if impl_sum > 0 else 0
            ev_val = compute_bet_ev(b["odds"], fair_prob)
            ev_c = GREEN if ev_val > 1 else (RED if ev_val < -1 else MUTED)
            ev_cell = f'<span style="color:{ev_c};font-family:Orbitron,monospace;font-size:11px;">{ev_val:+.1f}%</span>'
        rows_html += (
            f'<tr class="{row_class}">'
            f'<td>{time_label} {source_badge}</td>'
            f'<td style="color:{out_color};font-weight:700;">{out_name}</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:13px;">{b["odds"]:.2f} {shift_cell}</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:13px;">{risk}</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{GREEN};">+{potential}</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:12px;">{ev_cell}</td>'
            f'</tr>'
        )
    table_html = (
        '<div class="dash-card">'
        '<table class="ledger-table">'
        '<thead><tr><th>Time / Source</th><th>Outcome</th><th>Odds</th><th>Stake</th><th>Potential</th><th>EV</th></tr></thead>'
        '<tbody>' + rows_html + '</tbody>'
        '</table></div>'
    )
    st.markdown(table_html, unsafe_allow_html=True)
    # Per-item delete (any bet can be removed, not just the last)
    with st.expander("\U0001f5d1 Manage / Delete Individual Bets", expanded=False):
        for idx_b, b in enumerate(st.session_state.pre_bets):
            nm_b = name_map.get(b.get("outcome", ""), "?")
            c1, c2 = st.columns([4, 1])
            c1.markdown(f'<span style="color:{VIOLET};font-family:Rajdhani;font-size:14px;">PRE \u2022 {nm_b} \u2022 {fmt_inr(b["stake"])} @ {b["odds"]:.2f}</span>', unsafe_allow_html=True)
            if c2.button("\U0001f5d1", key=f"del_pre_{idx_b}", help=f"Delete pre-bet #{idx_b+1}"):
                st.session_state.pre_bets.pop(idx_b); sync_match_to_history(); st.rerun()
        for idx_b, b in enumerate(st.session_state.bets):
            nm_b = name_map.get(b.get("outcome", ""), "?")
            c1, c2 = st.columns([4, 1])
            c1.markdown(f'<span style="color:{GREEN};font-family:Rajdhani;font-size:14px;">LIVE \u2022 {nm_b} \u2022 {fmt_inr(b["stake"])} @ {b["odds"]:.2f}</span>', unsafe_allow_html=True)
            if c2.button("\U0001f5d1", key=f"del_live_{idx_b}", help=f"Delete live bet #{idx_b+1}"):
                st.session_state.bets.pop(idx_b); sync_match_to_history(); st.rerun()
        for idx_b, b in enumerate(st.session_state.get("misc_bets", [])):
            c1, c2 = st.columns([4, 1])
            c1.markdown(f'<span style="color:{MISC_PURPLE};font-family:Rajdhani;font-size:14px;">MISC \u2022 {b.get("label","Misc")} \u2022 {fmt_inr(b["stake"])} @ {b["odds"]:.2f}</span>', unsafe_allow_html=True)
            if c2.button("\U0001f5d1", key=f"del_misc_{idx_b}", help=f"Delete misc bet #{idx_b+1}"):
                st.session_state.misc_bets.pop(idx_b); sync_match_to_history(); st.rerun()
        if not (st.session_state.pre_bets or st.session_state.bets or st.session_state.get("misc_bets")):
            st.markdown(f'<div style="color:{MUTED};font-family:Rajdhani;">No bets to manage.</div>', unsafe_allow_html=True)


def render_pnl_chart(pnl):
    """P&L chart with stacked bars for pre+live+misc."""
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    all_bets = get_all_bets()
    if not all_bets and not st.session_state.get("misc_bets"):
        return
    st.markdown('<div class="section-title">\U0001f4c8  P&L Scenarios</div>', unsafe_allow_html=True)
    scenarios = [f"{t1n} Wins", f"{t2n} Wins", "Tie"]
    total_values = [pnl["t1"], pnl["t2"], pnl["tie"]]
    mp = compute_misc_pnl()
    fig = go.Figure()

    has_pre = len(st.session_state.pre_bets) > 0
    has_live = len(st.session_state.bets) > 0

    if has_pre and has_live:
        pre_pnl = compute_pnl(st.session_state.pre_bets)
        live_pnl = compute_pnl(st.session_state.bets)
        fig.add_trace(go.Bar(name="Pre-Existing", x=scenarios, y=[pre_pnl["t1"], pre_pnl["t2"], pre_pnl["tie"]], marker_color=VIOLET, opacity=0.7))
        fig.add_trace(go.Bar(name="Live Bets", x=scenarios, y=[live_pnl["t1"], live_pnl["t2"], live_pnl["tie"]], marker_color=CYAN, opacity=0.7))
        fig.update_layout(barmode="relative")
        for i, val in enumerate(total_values):
            c = GREEN if val >= 0 else RED
            fig.add_annotation(x=scenarios[i], y=val, text=f"<b>Net: {fmt_inr(val)}</b>", showarrow=False, yshift=18 if val >= 0 else -18, font=dict(family="Rajdhani", size=14, color=c))
    elif all_bets:
        total_colors = [GREEN if v >= 0 else RED for v in total_values]
        fig.add_trace(go.Bar(x=scenarios, y=total_values, marker_color=total_colors, text=[fmt_inr(v) for v in total_values], textposition="outside", textfont=dict(family="Rajdhani", size=16, color=TEXT)))

    # Add misc P&L as horizontal band annotations
    if mp["count"] > 0 and (mp["realized"] != 0 or mp["active_count"] > 0):
        misc_best = mp["best"]
        misc_worst = mp["worst"]
        for i, sc in enumerate(["t1", "t2", "tie"]):
            combined_best = pnl[sc] + misc_best
            combined_worst = pnl[sc] + misc_worst
            cb_c = GREEN if combined_best >= 0 else RED
            cw_c = GREEN if combined_worst >= 0 else RED
            y_pos = min(pnl[sc], 0) - 15
            fig.add_annotation(x=scenarios[i], y=y_pos, text=f"<b>+Misc: {fmt_inr(combined_worst)} to {fmt_inr(combined_best)}</b>",
                showarrow=False, font=dict(family="Rajdhani", size=11, color=MISC_PURPLE), yshift=-12)

    fig.add_hline(y=0, line_dash="dot", line_color=MUTED, line_width=1)
    fig.update_layout(
        plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Rajdhani", color=TEXT),
        legend=dict(font=dict(family="Rajdhani", size=13, color=TEXT), bgcolor="rgba(0,0,0,0)"),
        xaxis=dict(showgrid=False, tickfont=dict(family="Orbitron", size=12, color=MUTED)),
        yaxis=dict(showgrid=True, gridcolor="#1A2340", gridwidth=1, tickfont=dict(family="Orbitron", size=11, color=MUTED), zeroline=False),
        height=380, margin=dict(l=20, r=20, t=20, b=40), bargap=0.35,
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


def render_gemini_insight_card():
    key = st.session_state.history.get("settings", {}).get("gemini_key", "")
    if not key or not GENAI_AVAILABLE:
        st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-size:13px;color:{MUTED};padding:8px;">\U0001f916 Add your Gemini API key in Settings to unlock AI Strategy Advisor</div>', unsafe_allow_html=True)
        return

    st.markdown(f'<div class="section-title">\U0001f9e0  AI Strategy Advisor</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:-8px;margin-bottom:8px;">\U0001f310 Auto-fetches live scores, pitch reports, team news via Google Search</div>', unsafe_allow_html=True)

    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    has_bets = len(get_all_bets()) > 0

    # Optional extra context
    live_note = st.text_input("\U0001f4dd Extra context (optional)", placeholder=f"e.g. Dew heavy, {t1n} dropped 2 catches, key player injured", key="live_score_note")

    # Phase selector
    phase_options = {
        "pre_toss": "\U0001f3cf Pre-Toss (match hasn't started)",
        "post_toss": "\U0001f3b2 Post-Toss (toss done, innings starting)",
        "powerplay": "\u26a1 Powerplay (overs 1-6)",
        "middle": "\U0001f3af Middle Overs (7-14)",
        "death": "\U0001f4a5 Death Overs (15-20)",
        "innings_break": "\u23f8 Innings Break",
        "chase": "\U0001f3c3 Chase (2nd innings live)",
    }
    phase_keys = list(phase_options.keys())

    # Auto-suggest phase
    if st.session_state.mode == "Pre-Match":
        default_idx = 0
    elif st.session_state.match_phase == "Early":
        default_idx = 2
    elif st.session_state.match_phase == "Middle":
        default_idx = 3
    else:
        default_idx = 4

    selected_phase = st.selectbox("Match Phase", phase_keys, index=default_idx, format_func=lambda x: phase_options[x], key="strategy_phase_sel")

    ac1, ac2 = st.columns(2)
    with ac1:
        if st.button("\U0001f9e0 Get Strategy Advice", key="strategy_advisor_btn", use_container_width=True, type="primary"):
            context = build_strategy_context()
            context_json = json.dumps(context, default=str)
            st.markdown(
                f'<div class="gemini-card"><div style="font-family:Orbitron,monospace;font-size:12px;color:{AMBER};letter-spacing:2px;margin-bottom:8px;">'
                f'\U0001f9e0 STRATEGY — {phase_options[selected_phase]}</div>',
                unsafe_allow_html=True,
            )
            with st.spinner("Analysing match..."):
                response = gemini_strategy_advisor(context_json, selected_phase)
            if response:
                st.markdown(response)
            else:
                err = st.session_state.get("gemini_last_error", "No response")
                if "429" in err or "RESOURCE_EXHAUSTED" in err:
                    st.warning("\U0001f551 Rate limited. Wait 60s and try again.")
                else:
                    st.warning(f"Error: {err[:200]}")
            st.markdown('</div>', unsafe_allow_html=True)

    with ac2:
        if has_bets and st.button("\U0001f916 Quick P&L Insight", key="gemini_live_btn", use_container_width=True):
            all_b = get_all_bets()
            pnl = compute_pnl(all_b)
            mp = compute_misc_pnl()
            state = {
                "match": get_current_match().get("label", "Unknown") if get_current_match() else "Unknown",
                "bets": [{"outcome": b.get("outcome", "misc"), "odds": b["odds"], "stake": b["stake"]} for b in all_b],
                "misc_bets": [{"label": b.get("label", ""), "odds": b["odds"], "stake": b["stake"]} for b in st.session_state.get("misc_bets", [])],
                "live_odds": get_current_odds(), "pnl_scenarios": pnl, "misc_pnl": mp,
                "remaining_capital": max(0, st.session_state.total_capital - compute_total_staked(get_all_bets_including_misc())),
                "session_target": st.session_state.history.get("goals", {}).get("session_target", 0),
                "session_floor": st.session_state.history.get("goals", {}).get("session_min_acceptable", 0),
            }
            st.markdown(f'<div class="gemini-card"><div style="font-family:Orbitron,monospace;font-size:12px;color:{AMBER};letter-spacing:2px;margin-bottom:8px;">\U0001f916 P&L INSIGHT</div>', unsafe_allow_html=True)
            with st.spinner("Analysing..."):
                response = gemini_live_insight(json.dumps(state, default=str))
            if response:
                st.markdown(f"*{response}*")
            else:
                err = st.session_state.get("gemini_last_error", "No response")
                st.warning(f"Error: {err[:150]}")
            st.markdown('</div>', unsafe_allow_html=True)


def render_settle_match_dialog():
    if not st.session_state.get("show_settle"):
        return
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    misc_bets = st.session_state.get("misc_bets", [])

    st.markdown(f'<div class="dash-card" style="border-top:3px solid {AMBER};">', unsafe_allow_html=True)
    st.markdown(f'<div class="section-title">\U0001f3c1  Settle Match</div>', unsafe_allow_html=True)
    result = st.selectbox("Actual Match Result", [f"{t1n} Won", f"{t2n} Won", "Tie", "No Result", "Abandoned"], key="settle_result_sel")

    # ── Per-misc-bet settlement ──
    unsettled_misc = [b for b in misc_bets if b.get("status", "active") == "active"]
    presettled_misc = [b for b in misc_bets if b.get("status") == "settled"]

    if misc_bets:
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:11px;color:{MISC_PURPLE};letter-spacing:2px;margin:12px 0 8px;">\U0001f3b2 SETTLE SIDE BETS ({len(misc_bets)})</div>', unsafe_allow_html=True)
        # Show already-settled bets as read-only
        for b in presettled_misc:
            rp = b.get("realized_pnl", 0)
            rc = GREEN if rp > 0 else (RED if rp < 0 else MUTED)
            icon = "\u2705" if b.get("settled_result") == "won" else ("\u274c" if b.get("settled_result") == "lost" else "\u23ed")
            st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};">{icon} {b.get("label","Misc")} \u2022 {fmt_inr(b["stake"])} @ {b["odds"]:.2f} \u2192 <span style="color:{rc};font-family:Orbitron,monospace;">{fmt_inr(rp)}</span> (already settled)</div>', unsafe_allow_html=True)
        # Show unsettled bets with radio buttons
        for i, mb in enumerate(unsettled_misc):
            mc1, mc2 = st.columns([3, 1])
            mc1.markdown(f'<span style="font-family:Rajdhani;font-size:14px;color:{MISC_PURPLE};">{mb.get("label","Misc")} \u2022 {fmt_inr(mb["stake"])} @ {mb["odds"]:.2f}</span>', unsafe_allow_html=True)
            misc_result = mc2.radio("Result", ["Won", "Lost", "Void"], key=f"misc_settle_{i}", horizontal=True, label_visibility="collapsed")
            st.session_state[f"_misc_result_{i}"] = misc_result

    # ── Preview P&L ──
    result_map = {f"{t1n} Won": "t1_win", f"{t2n} Won": "t2_win", "Tie": "tie", "No Result": "no_result", "Abandoned": "abandoned"}
    result_key = result_map.get(result, "no_result")
    all_b = get_all_bets()
    pnl = compute_pnl(all_b)
    if result_key in ("t1_win", "t2_win", "tie"):
        main_pnl = pnl[{"t1_win": "t1", "t2_win": "t2", "tie": "tie"}[result_key]]
    else:
        main_pnl = 0.0
    # Misc P&L: sum of presettled + unsettled selections
    misc_pnl_preview = sum(b.get("realized_pnl", 0) for b in presettled_misc)
    for i, mb in enumerate(unsettled_misc):
        mr = st.session_state.get(f"_misc_result_{i}", "Lost")
        if mr == "Won":
            misc_pnl_preview += mb["stake"] * (mb["odds"] - 1)
        elif mr == "Lost":
            misc_pnl_preview -= mb["stake"]
    total_preview = main_pnl + misc_pnl_preview
    pc = GREEN if total_preview >= 0 else RED
    st.markdown(f'<div style="font-family:Rajdhani;font-size:15px;margin:12px 0;">Preview: Main P&L <strong>{fmt_inr(main_pnl)}</strong>'
        + (f' + Misc P&L <strong style="color:{GREEN if misc_pnl_preview>=0 else RED};">{fmt_inr(misc_pnl_preview)}</strong>' if misc_bets else '')
        + f' = Total <strong style="color:{pc};font-family:Orbitron,monospace;">{fmt_inr(total_preview)}</strong></div>', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    if c1.button("\u2713 Confirm Settlement", use_container_width=True, type="primary", key="settle_confirm"):
        # Calculate misc P&L: presettled already have realized_pnl, settle the rest
        misc_total_pnl = sum(b.get("realized_pnl", 0) for b in presettled_misc)
        for i, mb in enumerate(unsettled_misc):
            mr = st.session_state.get(f"_misc_result_{i}", "Lost")
            if mr == "Won":
                mb["settled_result"] = "won"
                mb["realized_pnl"] = round(mb["stake"] * (mb["odds"] - 1), 2)
            elif mr == "Lost":
                mb["settled_result"] = "lost"
                mb["realized_pnl"] = -mb["stake"]
            else:  # Void
                mb["settled_result"] = "void"
                mb["realized_pnl"] = 0
            mb["status"] = "settled"
            misc_total_pnl += mb.get("realized_pnl", 0)

        realized = main_pnl
        mid = st.session_state.current_match_id
        h = st.session_state.history
        for m in h["matches"]:
            if m["id"] == mid:
                m["status"] = "settled"; m["result"] = result_key; m["realized_pnl"] = realized
                m["misc_realized_pnl"] = misc_total_pnl
                m["closing_capital"] = m.get("opening_capital", 50000) + realized + misc_total_pnl
                m["settled_at"] = datetime.now().isoformat()
                m["bets"] = copy.deepcopy(st.session_state.bets)
                m["pre_bets"] = copy.deepcopy(st.session_state.pre_bets)
                m["misc_bets"] = copy.deepcopy(st.session_state.get("misc_bets", []))
                m["current_odds"] = get_current_odds()
                gk = h.get("settings", {}).get("gemini_key", "")
                if gk and GENAI_AVAILABLE:
                    debrief = gemini_post_match_debrief(json.dumps(m, default=str))
                    if debrief:
                        m["debrief"] = debrief
                break
        h["learnings"] = extract_learnings()
        save_history(h); update_user_index_stats()
        st.session_state["show_settle"] = False
        total_settled = realized + misc_total_pnl
        st.success(f"Match settled! Main: {fmt_inr(realized)}" + (f" + Misc: {fmt_inr(misc_total_pnl)}" if misc_bets else "") + f" = Total: {fmt_inr(total_settled)}")
        time.sleep(1); st.rerun()
    if c2.button("\u2717 Cancel", use_container_width=True, key="settle_cancel"):
        st.session_state["show_settle"] = False; st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ── Lobby & Tabs ─────────────────────────────────────────────

def render_standalone_misc_bets():
    """Standalone misc bets with full analytics — scenario analysis, EV, recommendations."""
    h = st.session_state.history
    if "standalone_misc_bets" not in h:
        h["standalone_misc_bets"] = []
    smbets = h["standalone_misc_bets"]
    cap = get_portfolio_capital()
    sm_exposure = get_standalone_exposure()
    available = cap - sm_exposure

    st.markdown(f'<div class="section-title">\U0001f3b2  Season Side Bets</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};margin-bottom:12px;">Long-term bets outside live matches \u2014 playoff qualifiers, tournament winners, player awards, etc. Settle manually when the result is known.</div>', unsafe_allow_html=True)

    # ── Capital overview ──
    if sm_exposure > 0:
        exp_pct = (sm_exposure / cap * 100) if cap > 0 else 0
        st.markdown(f"""
        <div class="dash-card" style="padding:14px 18px;">
            <div style="display:flex;justify-content:space-between;align-items:center;">
                <div>
                    <div style="font-family:Rajdhani;font-size:12px;color:{MUTED};text-transform:uppercase;letter-spacing:1px;">Deployed in Side Bets</div>
                    <div style="font-family:Orbitron,monospace;font-size:20px;font-weight:700;color:{AMBER};">{fmt_inr(sm_exposure)} <span style="font-size:13px;color:{MUTED};">({exp_pct:.1f}% of capital)</span></div>
                </div>
                <div style="text-align:right;">
                    <div style="font-family:Rajdhani;font-size:12px;color:{MUTED};text-transform:uppercase;letter-spacing:1px;">Available</div>
                    <div style="font-family:Orbitron,monospace;font-size:20px;font-weight:700;color:{GREEN};">{fmt_inr(available)}</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Add new bet ──
    with st.form("standalone_misc_form", clear_on_submit=True):
        sm1, sm2, sm3 = st.columns([2.5, 1, 1])
        with sm1:
            sm_label = st.text_input("Bet Label", placeholder="e.g. MI to qualify, Orange Cap Kohli")
        with sm2:
            sm_odds = st.number_input("Odds", min_value=1.01, step=0.5, format="%.2f", value=3.00, key="sm_odds")
        with sm3:
            sm_stake = st.number_input("Stake (\u20b9)", min_value=10.0, step=100.0, format="%.0f", value=500.0, key="sm_stake")
        if st.form_submit_button("\U0001f3b2 Place Side Bet", use_container_width=True, type="primary"):
            if sm_stake > available + 0.01:
                st.error(f"Exceeds available capital ({fmt_inr(available)})")
            elif not sm_label.strip():
                st.error("Enter a label for the bet.")
            else:
                smbets.append({
                    "id": str(uuid.uuid4())[:8], "label": sm_label.strip(),
                    "odds": sm_odds, "stake": sm_stake,
                    "status": "active", "realized_pnl": 0,
                    "created_at": datetime.now().isoformat(), "settled_at": None,
                })
                save_history(h); update_user_index_stats(); st.rerun()

    # ── Active bets with analytics ──
    active = [b for b in smbets if b.get("status") == "active"]
    settled_sm = [b for b in smbets if b.get("status") == "settled"]

    if active:
        total_exposure = sum(b["stake"] for b in active)
        total_potential = sum(b["stake"] * (b["odds"] - 1) for b in active)
        worst_case = -total_exposure

        # ── Scenario Analysis ──
        st.markdown(f'<div class="section-title">\U0001f4ca  Scenario Analysis</div>', unsafe_allow_html=True)

        # Best case / worst case / expected
        n = len(active)
        sc1, sc2, sc3 = st.columns(3)
        with sc1:
            st.markdown(f'<div class="dash-card" style="border-top:3px solid {GREEN};"><div class="metric-box"><div class="metric-label">All {n} Win</div><div class="metric-value" style="color:{GREEN};">{fmt_inr(total_potential)}</div></div></div>', unsafe_allow_html=True)
        with sc2:
            st.markdown(f'<div class="dash-card" style="border-top:3px solid {RED};"><div class="metric-box"><div class="metric-label">All {n} Lose</div><div class="metric-value" style="color:{RED};">{fmt_inr(worst_case)}</div></div></div>', unsafe_allow_html=True)
        with sc3:
            # Expected value: sum of individual EVs
            # For each bet: EV = (1/odds)*profit - (1-1/odds)*stake
            total_ev = 0
            for b in active:
                p = 1.0 / b["odds"]  # implied prob (rough)
                total_ev += p * b["stake"] * (b["odds"] - 1) - (1 - p) * b["stake"]
            ev_color = GREEN if total_ev > 0 else (RED if total_ev < 0 else MUTED)
            st.markdown(f'<div class="dash-card" style="border-top:3px solid {AMBER};"><div class="metric-box"><div class="metric-label">Expected Value</div><div class="metric-value" style="color:{ev_color};">{fmt_inr(total_ev)}</div></div></div>', unsafe_allow_html=True)

        # ── Per-combination scenario table ──
        if n <= 6:
            with st.expander(f"\U0001f4cb Detailed Scenarios ({2**n} combinations)", expanded=n <= 3):
                scenarios = []
                for mask in range(2**n):
                    pnl_total = 0
                    wins = []
                    for i, b in enumerate(active):
                        if mask & (1 << i):
                            pnl_total += b["stake"] * (b["odds"] - 1)
                            wins.append(b["label"][:20])
                        else:
                            pnl_total -= b["stake"]
                    scenarios.append((len(wins), pnl_total, wins))
                scenarios.sort(key=lambda x: -x[1])
                rows = ""
                for nw, pnl_val, win_list in scenarios:
                    pc = GREEN if pnl_val >= 0 else RED
                    win_text = ", ".join(win_list) if win_list else "None"
                    rows += (f'<tr><td style="font-family:Rajdhani;font-size:13px;">{win_text}</td>'
                        f'<td style="text-align:center;">{nw}/{n}</td>'
                        f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{pc};">{fmt_inr(pnl_val)}</td></tr>')
                st.markdown(
                    '<div class="dash-card"><table class="ledger-table">'
                    '<thead><tr><th>Winners</th><th>Hits</th><th>P&L</th></tr></thead>'
                    '<tbody>' + rows + '</tbody></table></div>',
                    unsafe_allow_html=True,
                )

        # ── Portfolio Impact ──
        exp_ratio = total_exposure / cap if cap > 0 else 0
        if exp_ratio < 0.15:
            risk_text = f'<span style="color:{GREEN};"> Healthy exposure level.</span>'
        elif exp_ratio < 0.25:
            risk_text = f'<span style="color:{RED};"> \u26a0 Above 15% \u2014 consider if this aligns with your no-loss strategy.</span>'
        else:
            risk_text = f'<span style="color:{RED};font-weight:600;"> \U0001f6a8 Over 25% of capital in side bets \u2014 high risk!</span>'
        st.markdown(
            f'<div class="dash-card" style="border-left:3px solid {AMBER};padding:14px 18px;">'
            f'<div style="font-family:Rajdhani;font-size:15px;color:{TEXT};">'
            f'\U0001f4b0 <strong>Portfolio Impact:</strong> '
            f'Side bets use <strong style="color:{AMBER};">{fmt_inr(total_exposure)}</strong> '
            f'({exp_ratio*100:.1f}% of capital). {risk_text}'
            f'</div></div>',
            unsafe_allow_html=True,
        )

        # ── Active bets list with per-bet EV ──
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:12px;color:{MISC_PURPLE};letter-spacing:2px;margin:16px 0 8px;">ACTIVE BETS ({len(active)})</div>', unsafe_allow_html=True)

        for b in active:
            pot = b["stake"] * (b["odds"] - 1)
            impl_p = 100.0 / b["odds"]
            ev_per_bet = (impl_p / 100) * pot - (1 - impl_p / 100) * b["stake"]
            ev_pct = ev_per_bet / b["stake"] * 100 if b["stake"] > 0 else 0
            ev_c = GREEN if ev_pct > 1 else (RED if ev_pct < -1 else MUTED)
            bc1, bc2, bc3, bc4 = st.columns([3, 1, 1, 1])
            bc1.markdown(
                f'<div style="font-family:Rajdhani;font-size:15px;color:{TEXT};">'
                f'<span style="color:{MISC_PURPLE};font-weight:600;">{b["label"]}</span>'
                f' \u00b7 {fmt_inr(b["stake"])} @ {b["odds"]:.2f}'
                f' \u00b7 pot <span style="color:{GREEN};">+{fmt_inr(pot)}</span>'
                f' \u00b7 EV <span style="color:{ev_c};">{ev_pct:+.1f}%</span>'
                f'</div>', unsafe_allow_html=True)
            if bc2.button("\u2705 Won", key=f"smw_{b['id']}"):
                b["status"] = "settled"
                b["realized_pnl"] = round(pot, 2)
                b["settled_at"] = datetime.now().isoformat()
                save_history(h); update_user_index_stats()
                st.success(f"Won! +{fmt_inr(b['realized_pnl'])}"); st.rerun()
            if bc3.button("\u274c Lost", key=f"sml_{b['id']}"):
                b["status"] = "settled"
                b["realized_pnl"] = -b["stake"]
                b["settled_at"] = datetime.now().isoformat()
                save_history(h); update_user_index_stats()
                st.warning(f"Lost: {fmt_inr(b['realized_pnl'])}"); st.rerun()
            if bc4.button("\U0001f5d1", key=f"smd_{b['id']}", help="Delete this bet"):
                h["standalone_misc_bets"] = [x for x in smbets if x["id"] != b["id"]]
                save_history(h); update_user_index_stats(); st.rerun()

        # Edit active bets
        with st.expander(f"\u270f\ufe0f Edit Active Side Bets", expanded=False):
            for b in active:
                st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MISC_PURPLE};font-weight:600;margin-top:8px;">{b["label"]}</div>', unsafe_allow_html=True)
                ec1, ec2, ec3, ec4 = st.columns([2, 1, 1, 0.5])
                with ec1:
                    new_lbl = st.text_input("Label", value=b["label"], key=f"sme_lbl_{b['id']}")
                with ec2:
                    new_odds = st.number_input("Odds", value=float(b["odds"]), min_value=1.01, step=0.05, format="%.2f", key=f"sme_od_{b['id']}")
                with ec3:
                    new_stk = st.number_input("Stake", value=float(b["stake"]), min_value=10.0, step=100.0, format="%.0f", key=f"sme_sk_{b['id']}")
                with ec4:
                    if st.button("\u2713", key=f"sme_sv_{b['id']}", use_container_width=True):
                        b["label"] = new_lbl; b["odds"] = new_odds; b["stake"] = new_stk
                        save_history(h); st.rerun()
    else:
        st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};text-align:center;padding:16px;">No active side bets. Place one above.</div>', unsafe_allow_html=True)

    # ── Settled history ──
    if settled_sm:
        st.markdown(f'<div style="font-family:Orbitron,monospace;font-size:12px;color:{MUTED};letter-spacing:2px;margin:20px 0 8px;">SETTLED ({len(settled_sm)})</div>', unsafe_allow_html=True)
        total_pnl = sum(b.get("realized_pnl", 0) for b in settled_sm)
        wins = sum(1 for b in settled_sm if b.get("realized_pnl", 0) > 0)
        tc = GREEN if total_pnl >= 0 else RED
        st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-bottom:8px;">Total P&L: <span style="color:{tc};font-family:Orbitron,monospace;">{fmt_inr(total_pnl)}</span> \u00b7 {wins}/{len(settled_sm)} won</div>', unsafe_allow_html=True)
        rows = ""
        for b in reversed(settled_sm):
            rp = b.get("realized_pnl", 0)
            pc = GREEN if rp >= 0 else RED
            rows += (
                f'<tr><td style="color:{MISC_PURPLE};font-weight:600;">{b["label"]}</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:13px;">{b["odds"]:.2f}</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:13px;">{fmt_inr(b["stake"])}</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{pc};">{fmt_inr(rp)}</td>'
                f'<td style="font-size:12px;color:{MUTED};">{str(b.get("settled_at",""))[:10]}</td></tr>'
            )
        st.markdown(
            '<div class="dash-card"><table class="ledger-table">'
            '<thead><tr><th>Bet</th><th>Odds</th><th>Stake</th><th>P&L</th><th>Settled</th></tr></thead>'
            '<tbody>' + rows + '</tbody></table></div>',
            unsafe_allow_html=True,
        )
        with st.expander("\U0001f5d1 Delete Settled Side Bets", expanded=False):
            for b in reversed(settled_sm):
                rp = b.get("realized_pnl", 0)
                dc1, dc2 = st.columns([4, 1])
                dc1.markdown(f'<span style="font-family:Rajdhani;font-size:13px;color:{TEXT};">{b["label"]} \u2022 {fmt_inr(rp)}</span>', unsafe_allow_html=True)
                if dc2.button("\U0001f5d1", key=f"smdel_{b['id']}"):
                    h["standalone_misc_bets"] = [x for x in h["standalone_misc_bets"] if x["id"] != b["id"]]
                    save_history(h); update_user_index_stats(); st.rerun()


def render_match_lobby():
    _, center, _ = st.columns([1, 2, 1])
    with center:
        render_header()
    with st.sidebar:
        if st.button("\U0001f464 Switch User", key="sw_u_lobby", use_container_width=True, type="secondary"):
            update_user_index_stats()
            st.session_state.active_user = None
            st.rerun()
        render_fund_sidebar_panel()

    learnings = st.session_state.history.get("learnings", {})
    health, hc, hdesc = get_portfolio_health(learnings)
    pmode, pmc, pmdesc = get_portfolio_mode()
    cap = get_portfolio_capital()

    st.markdown(f"""
    <div style="text-align:center;margin:16px 0;">
        <span class="health-badge" style="background:{hc}22;color:{hc};border:1px solid {hc}44;">{health}</span>
        <span class="health-badge" style="background:{pmc}22;color:{pmc};border:1px solid {pmc}44;">{pmode}</span>
        <div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-top:6px;">{hdesc}</div>
    </div>
    """, unsafe_allow_html=True)

    sm_exposure = get_standalone_exposure()
    free_capital = cap - sm_exposure

    st.markdown(
        f'<div style="text-align:center;margin-bottom:24px;">'
        f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};letter-spacing:2px;text-transform:uppercase;">Free Capital</div>'
        f'<div style="font-family:Orbitron,monospace;font-size:36px;font-weight:800;color:{GREEN};">{fmt_inr(free_capital)}</div>'
        f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:4px;">Ready to deploy in match sessions</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    if sm_exposure > 0:
        st.markdown(
            f'<div style="text-align:center;margin:-16px 0 20px;">'
            f'<span style="font-family:Rajdhani;font-size:13px;color:{MUTED};">'
            f'Total: <strong style="font-family:Orbitron,monospace;">{fmt_inr(cap)}</strong>'
            f'</span>'
            f'<span style="color:{MUTED};"> \u00b7 </span>'
            f'<span style="font-family:Rajdhani;font-size:13px;color:{AMBER};">'
            f'\U0001f3b2 Side bets: <strong style="font-family:Orbitron,monospace;">{fmt_inr(sm_exposure)}</strong>'
            f'</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    tabs = st.tabs(["\U0001f3cf Matches", "\U0001f3b2 Side Bets", "\U0001f4ca Portfolio", "\U0001f4dc History", "\u2699\ufe0f Settings"])

    # ── TAB 0: Matches ──
    with tabs[0]:
        # Reconcile previous session
        with st.expander("\U0001f504 Reconcile Off-App Session"):
            st.markdown(
                f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};margin-bottom:4px;">'
                f'Played a session <strong>outside this app</strong> after you started using it? Record it here so '
                f'your portfolio stays accurate.</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                f'<div style="font-family:Rajdhani;font-size:13px;color:{AMBER};margin-bottom:12px;">'
                f'\u26a0 Do NOT reconcile sessions from <strong>before</strong> you created your account \u2014 '
                f'those are already in your starting capital.</div>',
                unsafe_allow_html=True,
            )

            with st.form("reconcile_form"):
                rl = st.text_input("Session Label", placeholder="e.g. Manual bets Mar 28")

                st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-bottom:4px;">'
                    f'<strong>Capital BEFORE</strong> this off-app session (what the app currently shows):</div>', unsafe_allow_html=True)
                r_before = st.number_input(
                    "Capital Before (\u20b9)",
                    min_value=0.0, step=1000.0, format="%.0f", value=cap,
                    help="Pre-filled with your current portfolio capital. Edit only if the app's number doesn't match your actual starting point for that session.",
                )

                st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-bottom:4px;">'
                    f'<strong>Capital AFTER</strong> this off-app session (what you actually have now):</div>', unsafe_allow_html=True)
                r_after = st.number_input(
                    "Capital After (\u20b9)",
                    min_value=0.0, step=1000.0, format="%.0f", value=cap,
                    help="Your actual total funds right now, after the off-app session concluded.",
                )

                st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-bottom:4px;">'
                    f'<strong>Total Wagered</strong> in that session (for analytics \u2014 does not affect P&L calculation):</div>', unsafe_allow_html=True)
                r_wagered = st.number_input(
                    "Total Wagered (\u20b9)",
                    min_value=0.0, step=1000.0, format="%.0f", value=0.0,
                    help="How much you staked in total during that off-app session. Used for learning engine only.",
                )

                # Show live P&L preview
                preview_pnl = r_after - r_before
                preview_color = GREEN if preview_pnl >= 0 else RED
                st.markdown(
                    f'<div style="font-family:Rajdhani;font-size:15px;margin:12px 0 4px;">'
                    f'Calculated P&L: <strong style="font-family:Orbitron,monospace;color:{preview_color};">'
                    f'{fmt_inr(preview_pnl)}</strong>'
                    f' <span style="color:{MUTED};font-size:13px;">({fmt_inr(r_after)} \u2212 {fmt_inr(r_before)})</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                if st.form_submit_button("\U0001f504 Reconcile Session", use_container_width=True):
                    # If "before" differs from what app tracks, record the gap as a fund adjustment
                    # This handles cases where user deposited/withdrew off-app without telling us
                    capital_gap = r_before - cap
                    if abs(capital_gap) >= 1:
                        gap_type = "deposit" if capital_gap > 0 else "withdrawal"
                        st.session_state.history["fund_log"].append({
                            "type": gap_type,
                            "amount": capital_gap,
                            "timestamp": datetime.now().isoformat(),
                            "context": f"Auto-adjustment from reconciliation (capital was {fmt_inr(r_before)} not {fmt_inr(cap)})",
                        })

                    reconciled_pnl = r_after - r_before
                    rec_match = {
                        "id": str(uuid.uuid4())[:8],
                        "label": rl or "Off-App Session",
                        "t1": "Manual", "t2": "Session",
                        "opening_capital": r_before,
                        "closing_capital": r_after,
                        "realized_pnl": reconciled_pnl,
                        "misc_realized_pnl": 0,
                        "status": "settled", "result": "manual",
                        "bets": [], "pre_bets": [], "misc_bets": [],
                        "opening_odds": {"t1": 0, "t2": 0, "tie": 0},
                        "match_phase": "Manual",
                        "settled_at": datetime.now().isoformat(),
                        "created_at": datetime.now().isoformat(),
                        "reconciled": True,
                        "wagered": r_wagered,
                        "capital_before": r_before,
                        "capital_after": r_after,
                    }
                    st.session_state.history["matches"].append(rec_match)
                    st.session_state.history["learnings"] = extract_learnings()
                    save_history(st.session_state.history); update_user_index_stats()
                    pnl_color = "\U0001f7e2" if reconciled_pnl >= 0 else "\U0001f534"
                    st.success(f"{pnl_color} Recorded: P&L {fmt_inr(reconciled_pnl)} \u2014 Portfolio capital updated to {fmt_inr(get_portfolio_capital())}")
                    st.rerun()

        # Continue in-progress matches
        in_progress = [m for m in st.session_state.history["matches"] if m.get("status") == "in_progress"]
        if in_progress:
            st.markdown(f'<div class="section-title">\u25b6\ufe0f  Continue Match</div>', unsafe_allow_html=True)
            for m in in_progress:
                nb = len(m.get("bets", [])) + len(m.get("pre_bets", [])) + len(m.get("misc_bets", []))
                ip_c1, ip_c2 = st.columns([4, 1])
                if ip_c1.button(f"\u25b6 {m['label']} \u2014 {nb} bets", key=f"cont_{m['id']}", use_container_width=True):
                    load_match_into_session(m); st.rerun()
                if ip_c2.button("\U0001f5d1", key=f"del_ip_{m['id']}", help=f"Discard {m.get('label','')}"):
                    st.session_state[f"confirm_del_ip_{m['id']}"] = True
                if st.session_state.get(f"confirm_del_ip_{m['id']}"):
                    st.warning(f"Discard in-progress match **{m.get('label','')}**? All bets will be lost.")
                    dip1, dip2 = st.columns(2)
                    if dip1.button("\u2713 Discard", key=f"dip_y_{m['id']}", type="primary"):
                        st.session_state.history["matches"] = [x for x in st.session_state.history["matches"] if x["id"] != m["id"]]
                        save_history(st.session_state.history)
                        st.session_state[f"confirm_del_ip_{m['id']}"] = False; st.rerun()
                    if dip2.button("\u2717 Keep", key=f"dip_n_{m['id']}"):
                        st.session_state[f"confirm_del_ip_{m['id']}"] = False; st.rerun()

        # Start new match
        st.markdown(f'<div class="section-title">\U0001f195  Start New Match</div>', unsafe_allow_html=True)
        with st.form("new_match_form"):
            ml = st.text_input("Match Label", placeholder="e.g. MI vs CSK \u2014 Mar 29")
            nc1, nc2 = st.columns(2)
            with nc1:
                t1 = st.text_input("Team 1", placeholder="e.g. MI")
            with nc2:
                t2 = st.text_input("Team 2", placeholder="e.g. CSK")
            oc1, oc2, oc3 = st.columns(3)
            with oc1:
                oo1 = st.number_input("Opening Odds T1", min_value=1.01, value=1.80, format="%.2f")
            with oc2:
                oo2 = st.number_input("Opening Odds T2", min_value=1.01, value=2.10, format="%.2f")
            with oc3:
                oot = st.number_input("Opening Odds Tie", min_value=1.01, value=51.0, format="%.2f")
            avail_cap = get_available_capital()
            st.markdown(
                f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-bottom:4px;">'
                f'Your tradable capital for this match (portfolio minus active side bets)'
                f'</div>', unsafe_allow_html=True)
            opening_cap = st.number_input("Opening Capital (\u20b9)", min_value=1000.0, value=avail_cap, step=5000.0, format="%.0f")
            if st.form_submit_button("\U0001f3cf Start Match", use_container_width=True, type="primary"):
                new_id = str(uuid.uuid4())[:8]
                match_rec = {
                    "id": new_id,
                    "label": ml or f"{t1 or 'Team 1'} vs {t2 or 'Team 2'}",
                    "t1": t1 or "Team 1", "t2": t2 or "Team 2",
                    "opening_capital": opening_cap,
                    "opening_odds": {"t1": oo1, "t2": oo2, "tie": oot},
                    "current_odds": {"t1": oo1, "t2": oo2, "tie": oot},
                    "bets": [], "pre_bets": [], "misc_bets": [],
                    "status": "in_progress",
                    "mode": "Pre-Match", "match_phase": "Early",
                    "bet_counter": 0, "pre_bet_counter": 0, "misc_bet_counter": 0,
                    "created_at": datetime.now().isoformat(),
                }
                st.session_state.history["matches"].append(match_rec)
                save_history(st.session_state.history)
                load_match_into_session(match_rec)
                st.rerun()

    # ── TAB 1: Side Bets ──
    with tabs[1]:
        render_standalone_misc_bets()

    # ── TAB 2: Portfolio Dashboard ──
    with tabs[2]:
        render_portfolio_dashboard()

    # ── TAB 3: Match History ──
    with tabs[3]:
        render_match_history()

    # ── TAB 4: Settings ──
    with tabs[4]:
        render_settings()


def generate_portfolio_recommendation():
    """Generate comprehensive portfolio-level recommendation considering everything."""
    h = st.session_state.history
    learnings = h.get("learnings", {})
    settled = get_settled_matches()
    cap = get_portfolio_capital()
    avail = get_available_capital()
    sm_exp = get_standalone_exposure()
    dd = get_drawdown_pct()
    peak = get_peak_capital()
    pmode, pmc, _ = get_portfolio_mode()

    # Active side bets
    sm_bets = [b for b in h.get("standalone_misc_bets", []) if b.get("status") == "active"]
    sm_settled = [b for b in h.get("standalone_misc_bets", []) if b.get("status") == "settled"]
    sm_settled_pnl = sum(b.get("realized_pnl", 0) for b in sm_settled)

    # In-progress matches
    in_progress = [m for m in h.get("matches", []) if m.get("status") == "in_progress"]

    rec = {
        "capital": cap, "available": avail, "peak": peak, "drawdown": dd,
        "mode": pmode, "sm_exposure": sm_exp,
        "cards": [],  # list of recommendation cards
    }

    # ── 1. Portfolio Health Card ──
    if dd >= 0.20:
        rec["cards"].append({"icon": "\U0001f6a8", "title": "Capital Preservation Required",
            "detail": f"Portfolio down {dd:.1%} from peak ({fmt_inr(peak)} \u2192 {fmt_inr(cap)}). Reduce bet sizes, hedge aggressively, avoid new anchor bets until recovery.",
            "color": RED, "priority": "critical"})
    elif dd >= 0.10:
        rec["cards"].append({"icon": "\u26a0\ufe0f", "title": "Moderate Drawdown \u2014 Caution",
            "detail": f"Down {dd:.1%} from peak. Tighter sizing recommended. Focus on high-confidence bets only.",
            "color": AMBER, "priority": "high"})
    elif dd < 0.05 and cap >= peak * 0.95:
        rec["cards"].append({"icon": "\U0001f4c8", "title": "Portfolio at Peak \u2014 Growth Mode",
            "detail": f"Capital at {fmt_inr(cap)} ({('all-time high!' if cap >= peak else f'{100-dd*100:.0f}% of peak')}). Standard sizing allowed. Good position to take calculated risks.",
            "color": GREEN, "priority": "info"})

    # ── 2. Exposure Analysis Card ──
    exposure_pct = (sm_exp / cap * 100) if cap > 0 else 0
    if sm_exp > 0:
        if exposure_pct > 25:
            rec["cards"].append({"icon": "\U0001f6a8", "title": f"Side Bet Exposure: {exposure_pct:.0f}% of Capital",
                "detail": f"{fmt_inr(sm_exp)} locked in {len(sm_bets)} active side bets. This is above 25% \u2014 a single bad run could significantly impact your capital. Consider reducing positions.",
                "color": RED, "priority": "critical"})
        elif exposure_pct > 15:
            rec["cards"].append({"icon": "\U0001f4b0", "title": f"Side Bet Exposure: {exposure_pct:.0f}% of Capital",
                "detail": f"{fmt_inr(sm_exp)} in {len(sm_bets)} bets. Moderate exposure \u2014 ensure remaining capital ({fmt_inr(avail)}) is sufficient for {3 if avail > 5000 else 2} more live match sessions.",
                "color": AMBER, "priority": "medium"})
        else:
            rec["cards"].append({"icon": "\u2705", "title": f"Side Bets: Healthy Exposure ({exposure_pct:.0f}%)",
                "detail": f"{fmt_inr(sm_exp)} deployed, {fmt_inr(avail)} available. Good balance \u2014 enough runway for live sessions.",
                "color": GREEN, "priority": "low"})

    # ── 3. Side Bet Scenario Consolidation ──
    if sm_bets:
        all_win_pnl = sum(b["stake"] * (b["odds"] - 1) for b in sm_bets)
        all_lose_pnl = -sum(b["stake"] for b in sm_bets)
        # Find the single best value bet
        best_ev = None
        for b in sm_bets:
            ev_pct = ((1.0 / b["odds"]) * (b["odds"] - 1) - (1 - 1.0 / b["odds"])) * 100
            if best_ev is None or ev_pct > best_ev[1]:
                best_ev = (b, ev_pct)
        # Find highest risk (largest stake at longest odds)
        highest_risk = max(sm_bets, key=lambda b: b["stake"] * (1 - 1.0 / b["odds"]))

        scenarios_text = f"Best case (all win): <strong style='color:{GREEN};'>{fmt_inr(all_win_pnl)}</strong>. "
        scenarios_text += f"Worst case (all lose): <strong style='color:{RED};'>{fmt_inr(all_lose_pnl)}</strong>."
        rec["cards"].append({"icon": "\U0001f4ca", "title": "Side Bet P&L Scenarios",
            "detail": scenarios_text, "color": CYAN, "priority": "info"})

    # ── 4. Next Session Sizing Recommendation ──
    if learnings.get("total_matches", 0) >= 1:
        anchor_frac = learnings.get("recommended_anchor_frac", 0.20)
        if pmode == "PRESERVATION":
            next_anchor = 0
            rec["cards"].append({"icon": "\U0001f6d1", "title": "Next Session: Hedge Only",
                "detail": "Preservation mode active. No new anchor bets. Only enter a session if you have a pre-existing position to hedge.",
                "color": RED, "priority": "high"})
        else:
            next_anchor = round(avail * anchor_frac / 10) * 10
            next_anchor = min(next_anchor, avail * 0.30)  # never more than 30% of available
            sessions_left = max(1, int(avail / max(next_anchor, 500)))
            rec["cards"].append({"icon": "\U0001f3af", "title": f"Next Session: Anchor \u2264 {fmt_inr(next_anchor)}",
                "detail": f"Based on {pmode} mode and {anchor_frac:.0%} sizing. Available capital ({fmt_inr(avail)}) supports ~{sessions_left} more sessions at this level.",
                "color": GREEN if pmode == "GROWTH" else AMBER, "priority": "medium"})

    # ── 5. Streak & Momentum Analysis ──
    streak = learnings.get("current_streak", 0)
    if streak >= 3:
        rec["cards"].append({"icon": "\U0001f525", "title": f"Winning Streak: {streak} in a Row",
            "detail": "Momentum is with you. The algorithm has slightly increased anchor sizing. Don't let confidence become overconfidence \u2014 stick to the hedge strategy.",
            "color": GREEN, "priority": "info"})
    elif streak <= -2:
        rec["cards"].append({"icon": "\u2744\ufe0f", "title": f"Losing Streak: {abs(streak)} in a Row",
            "detail": "The algorithm has tightened sizing and set hedge urgency to aggressive. This is correct. Do NOT chase losses with bigger bets. The math works over time \u2014 trust the process.",
            "color": RED, "priority": "high"})

    # ── 6. Cash-Out Consideration ──
    if sm_bets and sm_settled_pnl > 0 and cap > peak * 0.9:
        rec["cards"].append({"icon": "\U0001f4b5", "title": "Consider Partial Cash-Out",
            "detail": f"You've realized {fmt_inr(sm_settled_pnl)} from settled side bets and portfolio is near peak. Consider withdrawing {fmt_inr(int(sm_settled_pnl * 0.5))} (50% of realized gains) to lock in real profit while keeping the rest working.",
            "color": GREEN, "priority": "medium"})
    elif cap < h.get("starting_capital", 50000) * 0.8:
        rec["cards"].append({"icon": "\U0001f6b0", "title": "Capital Below 80% of Starting",
            "detail": f"Started with {fmt_inr(h.get('starting_capital', 0))}, now at {fmt_inr(cap)}. Consider adding funds if you believe in the strategy, or reduce session sizing to protect remaining capital.",
            "color": RED, "priority": "high"})

    # ── 7. Diversification Check ──
    if sm_bets and len(sm_bets) >= 2:
        unique_labels = set(b["label"].lower().split()[0] for b in sm_bets if b.get("label"))
        if len(unique_labels) <= 1:
            rec["cards"].append({"icon": "\u26a0\ufe0f", "title": "Low Diversification",
                "detail": "Multiple side bets on similar outcomes. If one fails, others likely fail too. Consider spreading across uncorrelated bets.",
                "color": AMBER, "priority": "medium"})

    # ── 8. ROI Tracking ──
    total_matches = learnings.get("total_matches", 0)
    if total_matches >= 3:
        match_pnl = learnings.get("total_pnl", 0)
        total_invested = sum(
            compute_total_staked(m.get("bets", []) + m.get("pre_bets", []) + m.get("misc_bets", []))
            for m in settled
        )
        roi = (match_pnl / total_invested * 100) if total_invested > 0 else 0
        roi_color = GREEN if roi > 0 else RED
        rec["cards"].append({"icon": "\U0001f4b1", "title": f"Match ROI: {roi:+.1f}%",
            "detail": f"Total wagered across {total_matches} matches: {fmt_inr(total_invested)}. Net return: {fmt_inr(match_pnl)}. {'Profitable strategy \u2014 keep executing.' if roi > 0 else 'Negative ROI \u2014 review bet timing and sizing.'}",
            "color": roi_color, "priority": "info"})

    # ── 9. Active Positions Summary ──
    if in_progress:
        rec["cards"].append({"icon": "\U0001f534", "title": f"{len(in_progress)} Match(es) In Progress",
            "detail": f"You have {len(in_progress)} active match session(s). Settle or manage them before starting new ones to keep capital accounting clean.",
            "color": CYAN, "priority": "info"})

    return rec


def render_portfolio_recommendation():
    """Render the portfolio-level recommendation engine."""
    rec = generate_portfolio_recommendation()
    h = st.session_state.history
    learnings = h.get("learnings", {})

    st.markdown(f'<div class="section-title">\U0001f9e0  Portfolio Recommendation</div>', unsafe_allow_html=True)

    # Sort cards: critical first, then high, medium, low, info
    priority_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    cards = sorted(rec["cards"], key=lambda c: priority_order.get(c.get("priority", "info"), 5))

    if not cards:
        st.markdown(f'<div style="font-family:Rajdhani;font-size:15px;color:{MUTED};text-align:center;padding:20px;">No recommendations yet. Place some bets or complete a match session to activate the portfolio engine.</div>', unsafe_allow_html=True)
        return

    for card in cards:
        bc = card["color"]
        st.markdown(
            f'<div class="dash-card" style="border-left:4px solid {bc};padding:16px 20px;">'
            f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:6px;">'
            f'<span style="font-size:22px;">{card["icon"]}</span>'
            f'<span style="font-family:Orbitron,monospace;font-size:14px;font-weight:600;color:{bc};letter-spacing:1px;">{card["title"]}</span>'
            f'</div>'
            f'<div style="font-family:Rajdhani;font-size:15px;color:{TEXT};line-height:1.5;">{card["detail"]}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Consolidated P&L Scenario Table ──
    sm_bets = [b for b in h.get("standalone_misc_bets", []) if b.get("status") == "active"]
    settled_list = get_settled_matches()
    match_pnl = sum(m.get("realized_pnl", 0) + m.get("misc_realized_pnl", 0) for m in settled_list)
    sm_settled_pnl = sum(b.get("realized_pnl", 0) for b in h.get("standalone_misc_bets", []) if b.get("status") == "settled")
    fund_total = sum(f.get("amount", 0) for f in h.get("fund_log", []))

    if sm_bets or settled_list:
        st.markdown(f'<div class="section-title">\U0001f4ca  Consolidated P&L</div>', unsafe_allow_html=True)
        rows = ""
        # Realized P&L
        rows += (f'<tr><td>Match Sessions (settled)</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{GREEN if match_pnl>=0 else RED};">{fmt_inr(match_pnl)}</td>'
            f'<td style="color:{MUTED};">{len(settled_list)} matches</td></tr>')
        rows += (f'<tr><td>Side Bets (settled)</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{GREEN if sm_settled_pnl>=0 else RED};">{fmt_inr(sm_settled_pnl)}</td>'
            f'<td style="color:{MUTED};">{len([b for b in h.get("standalone_misc_bets",[]) if b.get("status")=="settled"])} bets</td></tr>')
        if fund_total != 0:
            rows += (f'<tr><td>Fund Movements</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{GREEN if fund_total>=0 else RED};">{fmt_inr(fund_total)}</td>'
                f'<td style="color:{MUTED};">{len(h.get("fund_log",[]))} entries</td></tr>')
        total_realized = match_pnl + sm_settled_pnl
        rows += (f'<tr style="border-top:2px solid {CARD_BORDER};"><td style="font-weight:700;">Total Realized P&L</td>'
            f'<td style="font-family:Orbitron,monospace;font-size:15px;font-weight:700;color:{GREEN if total_realized>=0 else RED};">{fmt_inr(total_realized)}</td>'
            f'<td></td></tr>')
        # Unrealized (active side bets)
        if sm_bets:
            best_unrealized = sum(b["stake"] * (b["odds"] - 1) for b in sm_bets)
            worst_unrealized = -sum(b["stake"] for b in sm_bets)
            rows += (f'<tr><td style="color:{MISC_PURPLE};">Active Side Bets (if all win)</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{GREEN};">{fmt_inr(best_unrealized)}</td>'
                f'<td style="color:{MUTED};">{len(sm_bets)} active</td></tr>')
            rows += (f'<tr><td style="color:{MISC_PURPLE};">Active Side Bets (if all lose)</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:13px;color:{RED};">{fmt_inr(worst_unrealized)}</td>'
                f'<td></td></tr>')
            best_total = total_realized + best_unrealized
            worst_total = total_realized + worst_unrealized
            rows += (f'<tr style="border-top:2px solid {CARD_BORDER};"><td style="font-weight:700;">Best Case Total</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:15px;font-weight:700;color:{GREEN};">{fmt_inr(best_total)}</td><td></td></tr>')
            rows += (f'<tr><td style="font-weight:700;">Worst Case Total</td>'
                f'<td style="font-family:Orbitron,monospace;font-size:15px;font-weight:700;color:{RED if worst_total<0 else AMBER};">{fmt_inr(worst_total)}</td><td></td></tr>')

        st.markdown(
            '<div class="dash-card"><table class="ledger-table">'
            '<thead><tr><th>Category</th><th>P&L</th><th>Detail</th></tr></thead>'
            '<tbody>' + rows + '</tbody></table></div>',
            unsafe_allow_html=True,
        )

    # ── Gemini AI Strategic Advice ──
    gk = h.get("settings", {}).get("gemini_key", "")
    if gk and GENAI_AVAILABLE:
        st.markdown(f'<div class="section-title">\U0001f916  AI Strategic Advisor</div>', unsafe_allow_html=True)
        if st.button("\U0001f916 Get AI Portfolio Strategy", key="gemini_strat_btn", use_container_width=True):
            strat_data = {
                "portfolio": {
                    "total_capital": get_portfolio_capital(),
                    "available_capital": get_available_capital(),
                    "peak_capital": get_peak_capital(),
                    "starting_capital": h.get("starting_capital", 0),
                    "drawdown_pct": get_drawdown_pct(),
                    "portfolio_mode": get_portfolio_mode()[0],
                },
                "match_history": {
                    "total_matches": learnings.get("total_matches", 0),
                    "win_rate": learnings.get("win_rate", 0),
                    "net_pnl": learnings.get("total_pnl", 0),
                    "current_streak": learnings.get("current_streak", 0),
                    "best_match_pnl": learnings.get("best_pnl", 0),
                    "worst_match_pnl": learnings.get("worst_pnl", 0),
                    "avg_bets_per_match": learnings.get("avg_bets_per_match", 0),
                },
                "active_side_bets": [{"label": b["label"], "odds": b["odds"], "stake": b["stake"],
                    "potential_profit": b["stake"] * (b["odds"] - 1)} for b in sm_bets],
                "settled_side_bets_pnl": sm_settled_pnl,
                "fund_movements": fund_total,
                "recommendations_generated": [c["title"] for c in rec["cards"]],
            }
            prompt = (
                "You are a portfolio strategist for IPL cricket betting. Here is the user's COMPLETE portfolio state:\n\n"
                f"```json\n{json.dumps(strat_data, default=str, indent=2)}\n```\n\n"
                "Give 5-7 specific, actionable strategic recommendations. Cover:\n"
                "1. Should they add more side bets or stop? Which types?\n"
                "2. What should their next live match session look like (anchor size, hedge timing)?\n"
                "3. Should they cash out any gains or stay fully deployed?\n"
                "4. Risk assessment: what's the single biggest threat to their capital?\n"
                "5. If they follow optimal strategy, what's a realistic season-end capital target?\n"
                "6. Any side bets they should consider exiting early (if possible) or doubling down on?\n"
                "7. One thing they're doing right and one thing to improve.\n\n"
                "Be specific with numbers. Reference their actual capital, bets, and P&L. No generic advice."
            )
            st.markdown(f'<div class="gemini-card"><div style="font-family:Orbitron,monospace;font-size:12px;color:{AMBER};letter-spacing:2px;margin-bottom:8px;">\U0001f916 STRATEGIC ADVISOR</div>', unsafe_allow_html=True)
            with st.spinner("Analysing portfolio strategy (may retry on rate limit)..."):
                response = gemini_generate(prompt, stream=False)
            if response:
                st.markdown(response)
            else:
                err = st.session_state.get("gemini_last_error", "No response")
                if "429" in err or "RESOURCE_EXHAUSTED" in err:
                    st.warning(f"\U0001f551 Rate limited. Wait 60 seconds and try again.")
                else:
                    st.warning(f"Gemini error: {err[:200]}")
            st.markdown('</div>', unsafe_allow_html=True)


def render_portfolio_dashboard():
    learnings = st.session_state.history.get("learnings", {})
    settled = get_settled_matches()
    h = st.session_state.history

    st.markdown('<div class="section-title">\U0001f4ca  Portfolio Stats</div>', unsafe_allow_html=True)
    total_m = learnings.get("total_matches", 0)
    wr = learnings.get("win_rate", 0)
    match_pnl = learnings.get("total_pnl", 0)
    sm_settled_pnl = sum(b.get("realized_pnl", 0) for b in h.get("standalone_misc_bets", []) if b.get("status") == "settled")
    total_betting_pnl = match_pnl + sm_settled_pnl
    dd = get_drawdown_pct()

    c1, c2, c3, c4 = st.columns(4)
    for col, label, val, color in [
        (c1, "Matches", str(total_m), CYAN),
        (c2, "Win Rate", f"{wr:.0%}" if total_m else "\u2014", GREEN if wr >= 0.5 else RED),
        (c3, "Betting P&L", fmt_inr(total_betting_pnl) if (total_m or sm_settled_pnl) else "\u2014", GREEN if total_betting_pnl >= 0 else RED),
        (c4, "Drawdown", f"{dd:.1%}", GREEN if dd < 0.1 else (AMBER if dd < 0.2 else RED)),
    ]:
        with col:
            st.markdown(f'<div class="dash-card"><div class="metric-box"><div class="metric-label">{label}</div><div style="font-family:Orbitron,monospace;font-size:20px;font-weight:700;color:{color};">{val}</div></div></div>', unsafe_allow_html=True)

    # Season Goals Progress
    gt = compute_goals_tracker()
    if gt["target_pct"] > 0:
        if gt.get("target_achieved"):
            st.markdown(f'<div class="section-title">\U0001f3c6  Season Goal — ACHIEVED!</div>', unsafe_allow_html=True)
            st.markdown(
                f'<div class="dash-card" style="border:2px solid {GREEN};padding:18px;">'
                f'<div style="display:flex;justify-content:space-between;margin-bottom:10px;">'
                f'<div><span style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">TARGET</span>'
                f'<div style="font-family:Orbitron,monospace;font-size:20px;color:{GREEN};text-decoration:line-through;">{fmt_inr(gt["target_capital"])}</div></div>'
                f'<div style="text-align:center;"><span style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">CURRENT</span>'
                f'<div style="font-family:Orbitron,monospace;font-size:20px;color:{GREEN};">{fmt_inr(gt["cap"])}</div></div>'
                f'<div style="text-align:right;"><span style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">RETURN</span>'
                f'<div style="font-family:Orbitron,monospace;font-size:20px;color:{GREEN};">{gt["profit_pct"]:.0f}%</div></div>'
                f'</div>'
                f'<div style="font-family:Rajdhani;font-size:14px;color:{TEXT};text-align:center;">'
                f'\U0001f389 You crushed your {gt["target_pct"]}% target with {gt["matches_remaining"]} matches still to go! '
                f'Set a higher target in Settings or keep compounding.'
                f'</div></div>',
                unsafe_allow_html=True,
            )
            st.progress(1.0)
        else:
            st.markdown(f'<div class="section-title">\U0001f3af  Season Goal Progress</div>', unsafe_allow_html=True)
            prog = min(gt["progress_pct"] / 100, 1.0)
            prog_color = GREEN if gt["on_pace"] else AMBER

            st.markdown(
                f'<div class="dash-card" style="padding:18px;">'
                f'<div style="display:flex;justify-content:space-between;margin-bottom:10px;">'
                f'<div><span style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">TARGET</span>'
                f'<div style="font-family:Orbitron,monospace;font-size:20px;color:{CYAN};">{fmt_inr(gt["target_capital"])}</div></div>'
                f'<div style="text-align:center;"><span style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">PROGRESS</span>'
                f'<div style="font-family:Orbitron,monospace;font-size:20px;color:{prog_color};">{gt["progress_pct"]:.1f}%</div></div>'
                f'<div style="text-align:right;"><span style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">PROJECTED</span>'
                f'<div style="font-family:Orbitron,monospace;font-size:20px;color:{GREEN if gt["projected_final"]>=gt["target_capital"] else RED};">{fmt_inr(gt["projected_final"])}</div></div>'
                f'</div></div>',
                unsafe_allow_html=True,
            )
            st.progress(prog)

        # Key metrics row
        mc1, mc2, mc3, mc4 = st.columns(4)
        if gt.get("target_achieved"):
            surplus = abs(gt["profit_needed"])
            for col, label, val, color in [
                (mc1, "Betting Profit", fmt_inr(gt["betting_profit"]), GREEN),
                (mc2, "Surplus", f"+{fmt_inr(surplus)}", GREEN),
                (mc3, "Suggested/Match", fmt_inr(gt["per_match_profit"]), CYAN),
                (mc4, "Matches Left", str(gt["matches_remaining"]), MUTED),
            ]:
                with col:
                    st.markdown(f'<div class="dash-card"><div class="metric-box"><div class="metric-label">{label}</div><div style="font-family:Orbitron,monospace;font-size:16px;font-weight:700;color:{color};">{val}</div></div></div>', unsafe_allow_html=True)
        else:
            for col, label, val, color in [
                (mc1, "Betting Profit", fmt_inr(gt["betting_profit"]), GREEN if gt["betting_profit"] >= 0 else RED),
                (mc2, "Still Needed", fmt_inr(gt["profit_needed"]), AMBER if gt["profit_needed"] > 0 else GREEN),
                (mc3, "Per Match", fmt_inr(gt["per_match_profit"]), CYAN),
                (mc4, "Matches Left", str(gt["matches_remaining"]), MUTED),
            ]:
                with col:
                    st.markdown(f'<div class="dash-card"><div class="metric-box"><div class="metric-label">{label}</div><div style="font-family:Orbitron,monospace;font-size:16px;font-weight:700;color:{color};">{val}</div></div></div>', unsafe_allow_html=True)

        # Breakdown: what makes up the profit
        breakdown_parts = []
        if gt["match_profit"] != 0:
            mc = GREEN if gt["match_profit"] >= 0 else RED
            breakdown_parts.append(f'Matches: <span style="color:{mc};font-family:Orbitron,monospace;">{fmt_inr(gt["match_profit"])}</span>')
        if gt["sm_settled_profit"] != 0:
            sc = GREEN if gt["sm_settled_profit"] >= 0 else RED
            breakdown_parts.append(f'Side bets: <span style="color:{sc};font-family:Orbitron,monospace;">{fmt_inr(gt["sm_settled_profit"])}</span>')
        if gt["fund_total"] != 0:
            breakdown_parts.append(f'Funds added: <span style="color:{MUTED};font-family:Orbitron,monospace;">{fmt_inr(gt["fund_total"])}</span> <span style="color:{MUTED};">(not profit)</span>')
        if breakdown_parts:
            st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:-8px;">{" \u00b7 ".join(breakdown_parts)}</div>', unsafe_allow_html=True)

        # Side bets impact on goal
        if gt["sm_best_case"] > 0:
            with_sides_best = gt["profit_so_far"] + gt["sm_best_case"]
            with_sides_pct = (with_sides_best / gt["target_profit"] * 100) if gt["target_profit"] > 0 else 0
            st.markdown(
                f'<div style="font-family:Rajdhani;font-size:13px;color:{MISC_PURPLE};text-align:center;margin-top:-8px;">'
                f'\U0001f3b2 If all side bets win: progress jumps to {with_sides_pct:.0f}% '
                f'(+{fmt_inr(gt["sm_best_case"])} from side bets)</div>',
                unsafe_allow_html=True,
            )

    # Portfolio Recommendation Engine
    render_portfolio_recommendation()

    # Equity curve — shows all events including fund deposits
    if settled or h.get("fund_log"):
        st.markdown('<div class="section-title">\U0001f4c8  Equity Curve</div>', unsafe_allow_html=True)

        # Build timeline of ALL events
        events = []
        for m in settled:
            events.append({
                "type": "match", "label": m.get("label", "Match")[:20],
                "pnl": get_match_total_pnl(m),
                "time": m.get("settled_at", m.get("created_at", "")),
            })
        for f in h.get("fund_log", []):
            events.append({
                "type": "fund", "label": f"{'+ ' if f.get('amount',0) > 0 else '- '}{fmt_inr(abs(f.get('amount',0)))}",
                "pnl": f.get("amount", 0),
                "time": f.get("timestamp", ""),
            })
        events.sort(key=lambda x: str(x["time"]))

        starting_cap = h.get("starting_capital", 50000)
        running = starting_cap
        # Two series: total capital and betting-only capital
        total_caps = [starting_cap]
        betting_caps = [starting_cap]
        labels_plot = ["Start"]
        marker_colors = [GREEN]
        marker_sizes = [8]
        running_betting = starting_cap

        for ev in events:
            running += ev["pnl"]
            total_caps.append(running)
            labels_plot.append(ev["label"])
            if ev["type"] == "fund":
                running_betting += 0  # funds don't count as betting profit
                betting_caps.append(running_betting)
                marker_colors.append(AMBER)
                marker_sizes.append(12)
            else:
                running_betting += ev["pnl"]
                betting_caps.append(running_betting)
                mc = GREEN if ev["pnl"] >= 0 else RED
                marker_colors.append(mc)
                marker_sizes.append(8)

        fig = go.Figure()
        # Total capital line (includes fund deposits)
        fig.add_trace(go.Scatter(
            x=labels_plot, y=total_caps, mode="lines+markers", name="Total Capital",
            line=dict(color=GREEN, width=3),
            marker=dict(size=marker_sizes, color=marker_colors),
        ))
        # Betting-only line (excludes fund deposits)
        if any(ev["type"] == "fund" for ev in events):
            fig.add_trace(go.Scatter(
                x=labels_plot, y=betting_caps, mode="lines+markers", name="Betting Only",
                line=dict(color=CYAN, width=2, dash="dot"),
                marker=dict(size=6, color=CYAN),
            ))
        fig.add_hline(y=starting_cap, line_dash="dot", line_color=MUTED, line_width=1,
                      annotation_text=f"Start: {fmt_inr(starting_cap)}", annotation_font=dict(size=10, color=MUTED))

        # Annotate fund deposits
        for i, ev in enumerate(events):
            if ev["type"] == "fund":
                idx = i + 1  # +1 because of "Start"
                fig.add_annotation(x=labels_plot[idx], y=total_caps[idx], text=f"\U0001f4b0{ev['label']}",
                    showarrow=True, arrowhead=2, arrowcolor=AMBER, ax=0, ay=-30,
                    font=dict(size=10, color=AMBER))

        fig.update_layout(
            plot_bgcolor="rgba(0,0,0,0)", paper_bgcolor="rgba(0,0,0,0)",
            font=dict(family="Rajdhani", color=TEXT),
            legend=dict(font=dict(family="Rajdhani", size=12, color=TEXT), bgcolor="rgba(0,0,0,0)"),
            xaxis=dict(showgrid=False, tickfont=dict(family="Rajdhani", size=11, color=MUTED)),
            yaxis=dict(showgrid=True, gridcolor="#1A2340", tickfont=dict(family="Orbitron", size=11, color=MUTED)),
            height=350, margin=dict(l=20, r=20, t=10, b=40),
        )
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    # Excel export in portfolio dashboard
    if settled and OPENPYXL_AVAILABLE:
        st.markdown(f'<div class="section-title">\U0001f4be  Download Portfolio Report</div>', unsafe_allow_html=True)
        xlsx_data = generate_portfolio_excel()
        if xlsx_data:
            st.download_button(
                "\U0001f4ca Download Full Portfolio Spreadsheet",
                data=xlsx_data,
                file_name=f"portfolio_{st.session_state.active_user}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown(f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:-8px;">5 sheets: Summary, Match History, All Bets, Fund Log, Current Match</div>', unsafe_allow_html=True)

    # Pattern Mining
    render_deep_patterns()

    # Edit fund entries
    render_edit_fund_dialog()

    # Gemini portfolio narrative (on-demand)
    st.markdown('<div class="section-title">\U0001f916  AI Portfolio Analysis</div>', unsafe_allow_html=True)
    gk = h.get("settings", {}).get("gemini_key", "")
    if not gk or not GENAI_AVAILABLE:
        st.markdown(f'<div style="color:{MUTED};font-family:Rajdhani;">Add Gemini API key in Settings to unlock AI analysis.</div>', unsafe_allow_html=True)
    elif not settled:
        st.markdown(f'<div style="color:{MUTED};font-family:Rajdhani;">Settle at least one match to enable portfolio analysis.</div>', unsafe_allow_html=True)
    else:
        if st.button("\U0001f916 Ask Gemini About My Portfolio", key="gemini_portfolio_btn", use_container_width=True):
            portfolio_data = {
                "matches": [{"label": m.get("label"), "result": m.get("result"), "realized_pnl": get_match_total_pnl(m),
                    "opening_capital": m.get("opening_capital"), "closing_capital": m.get("closing_capital"),
                    "total_bets": len(m.get("bets", [])) + len(m.get("pre_bets", [])),
                    "misc_bets": len(m.get("misc_bets", [])),
                    "debrief": m.get("debrief", "")} for m in settled],
                "fund_log": h.get("fund_log", []),
                "learnings": learnings,
                "current_capital": get_portfolio_capital(),
                "drawdown_pct": get_drawdown_pct(),
                "portfolio_mode": get_portfolio_mode()[0],
            }
            st.markdown(f'<div class="gemini-card"><div style="font-family:Orbitron,monospace;font-size:12px;color:{AMBER};letter-spacing:2px;margin-bottom:8px;">\U0001f916 PORTFOLIO ANALYSIS</div>', unsafe_allow_html=True)
            with st.spinner("Analysing portfolio..."):
                response = gemini_portfolio_narrative(json.dumps(portfolio_data, default=str))
            if response:
                st.markdown(response)
            else:
                err = st.session_state.get("gemini_last_error", "No response")
                if "429" in err or "RESOURCE_EXHAUSTED" in err:
                    st.warning(f"\U0001f551 Rate limited after {GEMINI_MAX_RETRIES} retries. Wait 60 seconds and try again.")
                else:
                    st.warning(f"Gemini error: {err[:200]}")
                    st.markdown(f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};">Settings \u2192 Test API Key to diagnose.</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)


def render_match_history():
    settled = get_settled_matches()
    h = st.session_state.history
    st.markdown('<div class="section-title">\U0001f4dc  Match History</div>', unsafe_allow_html=True)
    if not settled:
        st.markdown(f'<div style="color:{MUTED};font-family:Rajdhani;text-align:center;padding:30px;">No settled matches yet.</div>', unsafe_allow_html=True)
        return
    result_labels = {"t1_win": "T1 Won", "t2_win": "T2 Won", "tie": "Tie", "no_result": "N/R", "abandoned": "ABD", "manual": "Manual"}
    rows = ""
    for m in reversed(settled):
        rpnl = get_match_total_pnl(m)
        pc = GREEN if rpnl >= 0 else RED
        nb = len(m.get("bets", [])) + len(m.get("pre_bets", [])) + len(m.get("misc_bets", []))
        rl = result_labels.get(m.get("result", ""), "?")
        rows += f'<tr><td>{m.get("label", "\u2014")}</td><td style="font-family:Orbitron,monospace;font-size:13px;">{fmt_inr(m.get("opening_capital", 0))}</td><td>{nb}</td><td style="font-family:Orbitron,monospace;font-size:13px;color:{pc};">{fmt_inr(rpnl)}</td><td style="font-family:Orbitron,monospace;font-size:13px;">{fmt_inr(m.get("closing_capital", 0))}</td><td>{rl}</td></tr>'
    st.markdown(f'<div class="dash-card"><table class="ledger-table"><thead><tr><th>Match</th><th>Opening</th><th>Bets</th><th>P&L</th><th>Closing</th><th>Result</th></tr></thead><tbody>{rows}</tbody></table></div>', unsafe_allow_html=True)

    # Per-match delete
    with st.expander("\U0001f5d1 Delete a Settled Match", expanded=False):
        for m in reversed(settled):
            rpnl = get_match_total_pnl(m)
            pc = GREEN if rpnl >= 0 else RED
            dm1, dm2 = st.columns([4, 1])
            dm1.markdown(f'<span style="font-family:Rajdhani;font-size:14px;color:{TEXT};">{m.get("label", "?")} \u2022 <span style="color:{pc};">{fmt_inr(rpnl)}</span></span>', unsafe_allow_html=True)
            if dm2.button("\U0001f5d1", key=f"del_match_{m['id']}", help=f"Delete {m.get('label','')}"):
                st.session_state[f"confirm_del_match_{m['id']}"] = True
            if st.session_state.get(f"confirm_del_match_{m['id']}"):
                st.warning(f"Delete **{m.get('label','')}** ({fmt_inr(rpnl)} P&L)? This will change your portfolio capital.")
                cdm1, cdm2 = st.columns(2)
                if cdm1.button("\u2713 Yes", key=f"cdm_y_{m['id']}", type="primary"):
                    st.session_state.history["matches"] = [x for x in st.session_state.history["matches"] if x["id"] != m["id"]]
                    st.session_state.history["learnings"] = extract_learnings()
                    save_history(st.session_state.history); update_user_index_stats()
                    st.session_state[f"confirm_del_match_{m['id']}"] = False; st.rerun()
                if cdm2.button("\u2717 No", key=f"cdm_n_{m['id']}"):
                    st.session_state[f"confirm_del_match_{m['id']}"] = False; st.rerun()

    for m in reversed(settled):
        db = m.get("debrief", "")
        if db:
            with st.expander(f"\U0001f4cb Debrief: {m.get('label', 'Match')}"):
                st.markdown(db)

    # Data Management: Export JSON + Export Excel + Import
    st.markdown('<div class="section-title">\U0001f4be  Data Management</div>', unsafe_allow_html=True)
    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        st.download_button("\U0001f4e5 Export JSON", data=json.dumps(h, indent=2, default=str), file_name=f"user_{st.session_state.active_user}.json", mime="application/json", use_container_width=True)
    with ec2:
        if OPENPYXL_AVAILABLE:
            xlsx_data = generate_portfolio_excel()
            if xlsx_data:
                st.download_button(
                    "\U0001f4ca Export Excel",
                    data=xlsx_data,
                    file_name=f"portfolio_{st.session_state.active_user}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        else:
            st.button("\U0001f4ca Excel (install openpyxl)", disabled=True, use_container_width=True)
    with ec3:
        uploaded = st.file_uploader("\U0001f4e4 Import History", type="json", key="import_hist")
        if uploaded:
            try:
                imported = json.load(uploaded)
                st.session_state.history = imported
                save_history(imported)
                st.success("History imported!"); st.rerun()
            except Exception:
                st.error("Invalid JSON file.")


def render_settings():
    h = st.session_state.history
    if "goals" not in h:
        h["goals"] = {"season_target_pct": 300, "total_matches": 70, "per_match_style": "adaptive", "risk_tolerance": "auto", "session_target": 500, "session_min_acceptable": 0, "session_strategy": "conviction"}

    g = h["goals"]

    # ── Session Goals (per match) ──
    st.markdown(f'<div class="section-title">\U0001f3af  Session Goals (Per Match)</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};margin-bottom:12px;">What you want from each individual match. Adjustable inside a live session too.</div>', unsafe_allow_html=True)

    sc1, sc2, sc3 = st.columns(3)
    with sc1:
        new_session_target = st.number_input("Target Profit (\u20b9)", min_value=0, max_value=500000, value=int(g.get("session_target", 500)), step=100, key="goal_session_target", help="How much profit you aim for each match")
    with sc2:
        new_session_min = st.number_input("Min Acceptable (\u20b9)", min_value=-5000, max_value=50000, value=int(g.get("session_min_acceptable", 0)), step=100, key="goal_session_min", help="0 = break-even minimum (no-loss). Negative = willing to accept small loss")
    with sc3:
        strategies = ["conviction", "hedge_first", "aggressive", "conservative"]
        strategy_labels = {"conviction": "\U0001f3af Conviction (pick team, hedge others)", "hedge_first": "\U0001f6e1 Hedge First (break-even priority)", "aggressive": "\U0001f525 Aggressive (maximize upside)", "conservative": "\U0001f9ca Conservative (small safe gains)"}
        current_strat = g.get("session_strategy", "conviction")
        new_strat = st.selectbox("Strategy", strategies, index=strategies.index(current_strat) if current_strat in strategies else 0, format_func=lambda x: strategy_labels.get(x, x), key="goal_session_strat")

    if new_session_target != g.get("session_target") or new_session_min != g.get("session_min_acceptable") or new_strat != g.get("session_strategy"):
        g["session_target"] = new_session_target
        g["session_min_acceptable"] = new_session_min
        g["session_strategy"] = new_strat
        save_history(h)

    # Preview
    cap = get_portfolio_capital()
    if cap > 0:
        st.markdown(
            f'<div class="dash-card" style="border-left:3px solid {GREEN};padding:12px 18px;">'
            f'<div style="font-family:Rajdhani;font-size:14px;color:{TEXT};">'
            f'Each match: aim for <strong style="color:{GREEN};">{fmt_inr(new_session_target)}</strong> profit '
            f'({new_session_target/cap*100:.1f}% of capital) \u00b7 '
            f'Floor: <strong style="color:{AMBER if new_session_min >= 0 else RED};">{fmt_inr(new_session_min)}</strong> '
            f'{"(no-loss)" if new_session_min >= 0 else "(willing to risk)"} \u00b7 '
            f'Strategy: {strategy_labels.get(new_strat, new_strat)}'
            f'</div></div>',
            unsafe_allow_html=True,
        )

    st.divider()

    # ── Season Goals (long term) ──
    st.markdown(f'<div class="section-title">\U0001f4c8  Season Goals (Long Term)</div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};margin-bottom:12px;">Your overall season target. The engine calculates the per-match pace needed.</div>', unsafe_allow_html=True)

    gc1, gc2 = st.columns(2)
    with gc1:
        new_target = st.number_input("Season Return Target (%)", min_value=50, max_value=10000, value=int(g.get("season_target_pct", 300)), step=50, key="goal_target_pct")
    with gc2:
        new_matches = st.number_input("Total Matches in Season", min_value=10, max_value=100, value=int(g.get("total_matches", 70)), step=1, key="goal_total_matches")
    if new_target != g.get("season_target_pct") or new_matches != g.get("total_matches"):
        g["season_target_pct"] = new_target
        g["total_matches"] = new_matches
        save_history(h)

    gt = compute_goals_tracker()
    if gt.get("target_achieved"):
        surplus = abs(gt["profit_needed"])
        st.markdown(
            f'<div class="dash-card" style="border-left:3px solid {GREEN};padding:14px 18px;">'
            f'<div style="font-family:Rajdhani;font-size:15px;color:{TEXT};">'
            f'\U0001f3c6 <strong style="color:{GREEN};">TARGET ACHIEVED!</strong> '
            f'{gt["target_pct"]}% return on {fmt_inr(gt["starting"])} — surpassed by {fmt_inr(surplus)}'
            f'<br>\U0001f4b0 Current capital: <strong style="color:{GREEN};">{fmt_inr(gt["cap"])}</strong>'
            f'<br>\U0001f4c8 Betting profit: <strong style="color:{GREEN};">{fmt_inr(gt["betting_profit"])}</strong> '
            f'({gt["profit_pct"]:.0f}% return)'
            f'<br>\U0001f680 Next milestone: Set a new target above or keep compounding at '
            f'{fmt_inr(gt["per_match_profit"])}/match (3% growth)'
            f'</div></div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="dash-card" style="border-left:3px solid {CYAN};padding:14px 18px;">'
            f'<div style="font-family:Rajdhani;font-size:15px;color:{TEXT};">'
            f'\U0001f3af Season target: <strong style="color:{GREEN};">{fmt_inr(gt["target_capital"])}</strong> '
            f'({gt["target_pct"]}% on {fmt_inr(gt["starting"])})'
            f'<br>Need: <strong style="color:{AMBER};">{fmt_inr(gt["profit_needed"])}</strong> more across {gt["matches_remaining"]} matches'
            f'<br>Pace: <strong style="color:{CYAN};">{fmt_inr(gt["per_match_profit"])}</strong>/match '
            f'({gt["per_match_pct"]:.1f}%)'
            f'<br>Session goal is <strong style="color:{GREEN if g["session_target"] >= gt["per_match_profit"] else AMBER};">'
            f'{"aligned" if g["session_target"] >= gt["per_match_profit"] else "below pace"}</strong> '
            f'({fmt_inr(g["session_target"])} vs {fmt_inr(gt["per_match_profit"])} needed)'
            f'</div></div>',
            unsafe_allow_html=True,
        )

    st.divider()
    st.markdown(f'<div style="font-family:Rajdhani;font-size:16px;color:{TEXT};margin-bottom:8px;">\U0001f916 <strong>Gemini API Key</strong></div>', unsafe_allow_html=True)
    st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-bottom:8px;">Get your free key from <a href="https://aistudio.google.com/apikey" style="color:{CYAN};">aistudio.google.com/apikey</a></div>', unsafe_allow_html=True)
    current_key = h.get("settings", {}).get("gemini_key", "")
    new_key = st.text_input("Enter Gemini API Key", value=current_key, type="password", key="gemini_key_input")
    if new_key != current_key:
        h.setdefault("settings", {})["gemini_key"] = new_key
        save_history(h)
        st.success("API key saved!")

    # Test button
    tc1, tc2 = st.columns([1, 2])
    with tc1:
        if st.button("\U0001f916 Test API Key", key="test_gemini_btn", use_container_width=True):
            if not new_key and not current_key:
                st.error("Enter an API key first.")
            elif not GENAI_AVAILABLE:
                st.error("Gemini package not installed.")
            else:
                with st.spinner("Testing connection..."):
                    ok, msg = test_gemini_key()
                    if ok:
                        st.success(msg)
                    else:
                        st.error(msg)
                        last_err = st.session_state.get("gemini_last_error", "")
                        if last_err:
                            st.markdown(f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:4px;">Detail: {last_err[:300]}</div>', unsafe_allow_html=True)
    with tc2:
        if GENAI_AVAILABLE:
            pkg = "google-genai (new)" if genai_client_module else "google-generativeai (legacy)"
            st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{GREEN};padding-top:8px;">\u2713 Package: {pkg}</div>', unsafe_allow_html=True)
        else:
            st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{RED};padding-top:8px;">\u2717 Not installed. Run: <code>pip install google-genai</code></div>', unsafe_allow_html=True)

    st.divider()
    st.markdown(f'<div style="font-family:Rajdhani;font-size:16px;color:{RED};margin-bottom:8px;">\U0001f5d1 <strong>Danger Zone</strong></div>', unsafe_allow_html=True)
    if st.button("\U0001f5d1 Reset ALL Data (matches, portfolio, learnings)", type="secondary", key="reset_all_btn"):
        st.session_state["confirm_full_reset"] = True
    if st.session_state.get("confirm_full_reset"):
        st.error("This will permanently delete ALL match history, portfolio data, and learnings for this user.")
        rc1, rc2 = st.columns(2)
        if rc1.button("\u2713 Delete Everything", type="primary", key="full_reset_yes"):
            st.session_state.history = {"settings": h.get("settings", {}), "matches": [], "fund_log": [], "learnings": {}, "starting_capital": h.get("starting_capital", 50000)}
            save_history(st.session_state.history); update_user_index_stats()
            st.session_state.current_match_id = None
            st.session_state["confirm_full_reset"] = False
            st.rerun()
        if rc2.button("\u2717 Cancel", key="full_reset_no"):
            st.session_state["confirm_full_reset"] = False; st.rerun()



# ── New: Cash-Out, Hedge Projection, EV, Patterns, Edit ─────

def render_cashout_and_hedge_projection(pnl, odds, remaining):
    """Show cash-out equivalent value and hedge cost sensitivity, including misc bets."""
    if not all(odds[k] > 1 for k in odds) or not get_all_bets():
        return

    cashout = compute_cashout_value(pnl, odds, remaining)
    mp = compute_misc_pnl()
    # Combined cash-out: main hedge value + misc realized + misc active worst (conservative)
    combined_cashout = cashout + mp.get("realized", 0) - sum(b["stake"] for b in st.session_state.get("misc_bets", []) if b.get("status", "active") == "active")
    # Best case combined: cashout + misc best
    combined_best = cashout + mp["best"] if mp["count"] > 0 else cashout

    display_val = combined_cashout if mp["count"] > 0 else cashout
    cashout_color = GREEN if display_val >= 0 else RED

    misc_detail = ""
    if mp["count"] > 0:
        misc_detail = (
            f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};margin-top:4px;">'
            f'Main hedge: {fmt_inr(cashout)} + Misc worst: {fmt_inr(mp["worst"])} = Combined: {fmt_inr(combined_cashout)}'
            f'{f" | Best: {fmt_inr(combined_best)}" if mp["active_count"] > 0 else ""}'
            f'</div>'
        )

    st.markdown(
        f'<div class="dash-card" style="border-left:3px solid {cashout_color};padding:14px 18px;">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;">'
        f'<div>'
        f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};text-transform:uppercase;letter-spacing:1px;">Cash-Out Value (hedge now)</div>'
        f'<div style="font-family:Orbitron,monospace;font-size:22px;font-weight:700;color:{cashout_color};">{fmt_inr(display_val)}</div>'
        f'{misc_detail}'
        f'</div>'
        f'<div style="text-align:right;">'
        f'<div style="font-family:Rajdhani;font-size:12px;color:{MUTED};">Lock in if you hedge right now</div>'
        f'</div></div></div>',
        unsafe_allow_html=True,
    )

    # Hedge cost sensitivity
    projections = compute_hedge_cost_projections(pnl, odds, remaining)
    if projections:
        with st.expander("\U0001f4ca Hedge Cost Sensitivity (what if odds move?)", expanded=False):
            t1n = st.session_state.t1_name or "T1"
            t2n = st.session_state.t2_name or "T2"
            best_sc = max(pnl, key=pnl.get)
            label = {"t1": t1n, "t2": t2n, "tie": "Tie"}[best_sc]
            st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{MUTED};margin-bottom:8px;">If <strong>{label}</strong> odds shift (your strong side):</div>', unsafe_allow_html=True)
            for p in projections:
                shift = p["shift_pct"]
                pnl_diff = p["pnl_diff"]
                dc = GREEN if pnl_diff > 0 else (RED if pnl_diff < 0 else MUTED)
                arrow = "\u25b2" if shift > 0 else "\u25bc"
                direction = "shorten" if shift < 0 else "drift"
                st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;padding:2px 0;">{arrow} {abs(shift)}% {direction}: lock-in becomes <span style="color:{dc};font-family:Orbitron,monospace;">{fmt_inr(p["min_pnl"])}</span> (<span style="color:{dc};">{("+" if pnl_diff>=0 else "")}{fmt_inr(pnl_diff)}</span>)</div>', unsafe_allow_html=True)


def render_ev_summary(pnl, odds):
    """Show EV summary for all active bets."""
    if not all(odds[k] > 1 for k in odds):
        return
    all_b = get_all_bets()
    if not all_b:
        return
    evs = compute_all_bets_ev(all_b, odds)
    pos_ev = [e for e in evs if e["ev_pct"] > 0]
    neg_ev = [e for e in evs if e["ev_pct"] < 0]
    avg_ev = np.mean([e["ev_pct"] for e in evs]) if evs else 0

    avg_color = GREEN if avg_ev > 0 else (RED if avg_ev < 0 else MUTED)
    st.markdown(f"""
    <div style="text-align:center;font-family:Rajdhani;font-size:13px;margin:4px 0;">
        <span style="color:{avg_color};font-weight:600;">Avg EV: {avg_ev:+.1f}%</span>
        <span style="color:{MUTED};"> \u00b7 </span>
        <span style="color:{GREEN};">{len(pos_ev)} +EV bets</span>
        <span style="color:{MUTED};"> \u00b7 </span>
        <span style="color:{RED};">{len(neg_ev)} \u2212EV bets</span>
    </div>
    """, unsafe_allow_html=True)


def render_deep_patterns():
    """Render pattern mining insights in portfolio dashboard."""
    patterns = extract_deep_patterns()
    if not patterns:
        st.markdown(f'<div style="color:{MUTED};font-family:Rajdhani;">Need 3+ settled matches for pattern analysis.</div>', unsafe_allow_html=True)
        return

    st.markdown(f'<div class="section-title">\U0001f50d  Pattern Mining</div>', unsafe_allow_html=True)
    for p in patterns:
        sev_color = RED if p["severity"] == "high" else (AMBER if p["severity"] == "medium" else MUTED)
        st.markdown(f"""
        <div class="dash-card" style="border-left:3px solid {sev_color};padding:14px 18px;">
            <div style="font-family:Rajdhani;font-size:15px;color:{TEXT};">
                {p['icon']} {p['insight']}
            </div>
        </div>
        """, unsafe_allow_html=True)


def render_edit_bet_dialog():
    """Expandable dialog to modify any bet's odds or stake."""
    t1n = st.session_state.t1_name or "T1"
    t2n = st.session_state.t2_name or "T2"
    name_map = {"t1": t1n, "t2": t2n, "tie": "Tie"}
    all_items = []
    for i, b in enumerate(st.session_state.pre_bets):
        all_items.append(("pre", i, b))
    for i, b in enumerate(st.session_state.bets):
        all_items.append(("live", i, b))
    for i, b in enumerate(st.session_state.get("misc_bets", [])):
        all_items.append(("misc", i, b))
    if not all_items:
        return

    with st.expander("\u270f\ufe0f Edit / Modify Bets", expanded=False):
        for src, idx, b in all_items:
            is_misc = src == "misc"
            lbl = b.get("label", b.get("time_label", f"{src} #{idx+1}"))
            out = b.get("label", name_map.get(b.get("outcome", ""), "?")) if is_misc else name_map.get(b.get("outcome", ""), "?")
            tag_color = {
                "pre": VIOLET, "live": GREEN, "misc": MISC_PURPLE
            }[src]

            key_prefix = f"edit_{src}_{idx}"
            st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{tag_color};font-weight:600;margin-top:8px;">{src.upper()} \u2022 {out} \u2022 {lbl}</div>', unsafe_allow_html=True)
            ec1, ec2, ec3 = st.columns([1, 1, 1])
            with ec1:
                new_odds = st.number_input("Odds", value=float(b["odds"]), min_value=1.01, step=0.05, format="%.2f", key=f"{key_prefix}_odds")
            with ec2:
                new_stake = st.number_input("Stake", value=float(b["stake"]), min_value=10.0, step=100.0, format="%.0f", key=f"{key_prefix}_stake")
            with ec3:
                if st.button("\u2713 Save", key=f"{key_prefix}_save", use_container_width=True):
                    if src == "pre":
                        st.session_state.pre_bets[idx]["odds"] = new_odds
                        st.session_state.pre_bets[idx]["stake"] = new_stake
                    elif src == "live":
                        st.session_state.bets[idx]["odds"] = new_odds
                        st.session_state.bets[idx]["stake"] = new_stake
                    elif src == "misc":
                        st.session_state.misc_bets[idx]["odds"] = new_odds
                        st.session_state.misc_bets[idx]["stake"] = new_stake
                    sync_match_to_history()
                    st.rerun()


def render_edit_fund_dialog():
    """Modify fund log entries."""
    h = st.session_state.history
    flog = h.get("fund_log", [])
    if not flog:
        return
    with st.expander(f"\u270f\ufe0f Edit Fund Entries ({len(flog)} total)", expanded=False):
        for i, fl in enumerate(flog):
            t = "Deposit" if fl.get("amount", 0) > 0 else "Withdrawal"
            fc1, fc2, fc3 = st.columns([2, 1, 1])
            with fc1:
                st.markdown(f'<div style="font-family:Rajdhani;font-size:13px;color:{TEXT};">{t}: {fmt_inr(abs(fl.get("amount", 0)))} ({str(fl.get("timestamp",""))[:16]})</div>', unsafe_allow_html=True)
            with fc2:
                new_amt = st.number_input("Amount", value=float(abs(fl.get("amount", 0))), min_value=0.0, step=100.0, format="%.0f", key=f"ef_{i}_amt")
            with fc3:
                if st.button("\u2713", key=f"ef_{i}_save", use_container_width=True):
                    sign = 1 if fl.get("amount", 0) > 0 else -1
                    h["fund_log"][i]["amount"] = new_amt * sign
                    save_history(h); update_user_index_stats(); st.rerun()



def render_conviction_panel(pnl, odds, remaining):
    """Conviction mode: pick your team, engine hedges to guarantee break-even on others."""
    t1n = st.session_state.t1_name or "Team 1"
    t2n = st.session_state.t2_name or "Team 2"
    name_map = {"t1": t1n, "t2": t2n, "tie": "Tie"}
    conviction = st.session_state.get("conviction")
    ponr = st.session_state.get("ponr_active", False)
    has_bets = len(get_all_bets()) > 0
    has_odds = all(odds[k] > 1 for k in odds)

    # ── PONR MODE ──
    if ponr and conviction:
        dead = conviction  # the team we believed in is now dead
        dead_name = name_map[dead]
        alive = [o for o in ["t1", "t2", "tie"] if o != dead]
        alive_names = [name_map[o] for o in alive]

        st.markdown(f"""
        <div class="dash-card" style="border-top:3px solid {RED};padding:18px;">
            <div style="font-family:Orbitron,monospace;font-size:14px;font-weight:700;color:{RED};letter-spacing:2px;">
                \U0001f6a8 POINT OF NO RETURN \u2014 {dead_name} can't win
            </div>
            <div style="font-family:Rajdhani;font-size:15px;color:{TEXT};margin-top:8px;">
                Strategy: minimize losses across {alive_names[0]} and {alive_names[1]}.
            </div>
        </div>
        """, unsafe_allow_html=True)

        if has_bets and has_odds:
            sol = solve_ponr_hedge(pnl, odds, remaining, dead)
            if sol and sol["success"]:
                min_pnl = sol["min_pnl"]
                pc = GREEN if min_pnl >= 0 else (AMBER if min_pnl > -100 else RED)
                stakes_html = ""
                for o, s in sol["stakes"].items():
                    if s >= 5:
                        stakes_html += f'<div style="color:{OUTCOME_COLORS.get(o, TEXT)};margin:4px 0;">\u279c <strong>{name_map[o]}</strong>: {fmt_inr(s)} @ {odds[o]:.2f}</div>'
                npnl = sol["new_pnl"]
                pnl_parts = " \u00b7 ".join(f'{name_map[o]}: <span style="color:{GREEN if npnl[o]>=0 else RED};">{fmt_inr(npnl[o])}</span>' for o in alive)
                st.markdown(f"""
                <div class="rec-card" style="border-color:{pc};">
                    <div class="rec-headline" style="color:{pc};">
                        {"Minimize Loss" if min_pnl < 0 else "Lock Profit"}: {fmt_inr(min_pnl)} worst-case
                    </div>
                    <div class="rec-detail">Place these hedges to equalize across the remaining outcomes:</div>
                    <div class="rec-stakes">{stakes_html}</div>
                    <div style="margin-top:10px;font-size:13px;color:{MUTED};">After hedge: {pnl_parts}</div>
                </div>
                """, unsafe_allow_html=True)

        # Undo PONR
        if st.button("\u21a9\ufe0f Undo Point of No Return", key="undo_ponr"):
            st.session_state.ponr_active = False
            sync_match_to_history(); st.rerun()
        return

    # ── CONVICTION SELECTION ──
    st.markdown(f'<div class="section-title">\U0001f3af  Conviction Mode</div>', unsafe_allow_html=True)

    if not conviction:
        st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};margin-bottom:8px;">Pick the team you think will win. The engine will hedge the other outcomes to break-even, so you profit only if your pick wins and lose nothing if it doesn\'t.</div>', unsafe_allow_html=True)
        cc1, cc2, cc3 = st.columns(3)
        if cc1.button(f"\U0001f3af {t1n}", key="conv_t1", use_container_width=True):
            st.session_state.conviction = "t1"; sync_match_to_history(); st.rerun()
        if cc2.button(f"\U0001f3af {t2n}", key="conv_t2", use_container_width=True):
            st.session_state.conviction = "t2"; sync_match_to_history(); st.rerun()
        if cc3.button("\u23e9 Skip (Standard Mode)", key="conv_skip", use_container_width=True, type="secondary"):
            pass  # just don't set conviction, engine uses standard mode
    else:
        conv_name = name_map[conviction]
        conv_color = OUTCOME_COLORS.get(conviction, GREEN)
        st.markdown(f"""
        <div class="dash-card" style="border-top:3px solid {conv_color};padding:16px 18px;">
            <div style="display:flex;justify-content:space-between;align-items:center;">
                <div>
                    <div style="font-family:Orbitron,monospace;font-size:11px;color:{MUTED};letter-spacing:2px;">YOUR CONVICTION</div>
                    <div style="font-family:Orbitron,monospace;font-size:22px;font-weight:700;color:{conv_color};">{conv_name} WINS</div>
                </div>
                <div style="text-align:right;">
                    <div style="font-family:Rajdhani;font-size:13px;color:{MUTED};">Win \u2192 profit \u00b7 Lose \u2192 break even</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Show conviction hedge calculation
        if has_bets and has_odds:
            sol = solve_conviction_hedge(pnl, odds, remaining, conviction)
            if sol and sol["success"]:
                if sol["feasible"]:
                    # Profitable scenario
                    cp = sol["conviction_profit"]
                    cp_color = GREEN if cp > 0 else (AMBER if cp >= 0 else RED)
                    stakes_html = ""
                    for o, s in sol["stakes"].items():
                        if s >= 5:
                            stakes_html += f'<div style="color:{OUTCOME_COLORS.get(o, TEXT)};margin:4px 0;">\u279c Hedge <strong>{name_map[o]}</strong>: {fmt_inr(s)} @ {odds[o]:.2f}</div>'
                    other_pnl_text = " \u00b7 ".join(f'{name_map[o]}: {fmt_inr(sol["other_pnls"][o])}' for o in sol["other_pnls"])
                    st.markdown(f"""
                    <div class="rec-card" style="border-color:{cp_color};">
                        <div style="font-size:28px;margin-bottom:4px;">\U0001f512</div>
                        <div class="rec-headline" style="color:{cp_color};">
                            If {conv_name} wins: {fmt_inr(cp)} profit
                        </div>
                        <div class="rec-detail">
                            If {conv_name} loses: break even ({other_pnl_text}). Hedge cost: {fmt_inr(sol["total_hedge"])}.
                        </div>
                        <div class="rec-stakes">{stakes_html}</div>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f'<div class="dash-card" style="border-left:3px solid {AMBER};"><div style="font-family:Rajdhani;font-size:15px;color:{AMBER};">\u26a0 Break-even hedge not feasible at current odds. The engine will minimize loss instead.</div></div>', unsafe_allow_html=True)

        # PONR + Change conviction buttons
        bc1, bc2, bc3 = st.columns(3)
        if bc1.button("\U0001f6a8 Point of No Return", key="ponr_btn", use_container_width=True, type="primary"):
            st.session_state["confirm_ponr"] = True
        if bc2.button("\U0001f504 Change Conviction", key="change_conv", use_container_width=True, type="secondary"):
            st.session_state.conviction = None
            st.session_state.ponr_active = False
            sync_match_to_history(); st.rerun()
        if bc3.button("\u23e9 Switch to Standard", key="to_standard", use_container_width=True, type="secondary"):
            st.session_state.conviction = None
            sync_match_to_history(); st.rerun()

        if st.session_state.get("confirm_ponr"):
            st.error(f"\U0001f6a8 Confirm: {conv_name} CANNOT win this match anymore. Strategy will switch to minimize losses.")
            pc1, pc2 = st.columns(2)
            if pc1.button(f"\u2713 Yes, {conv_name} is done", key="ponr_yes", type="primary", use_container_width=True):
                st.session_state.ponr_active = True
                st.session_state["confirm_ponr"] = False
                sync_match_to_history(); st.rerun()
            if pc2.button("\u2717 Cancel", key="ponr_no", use_container_width=True):
                st.session_state["confirm_ponr"] = False; st.rerun()


# ── Current Match Tab (preserves entire original UI + new features) ──

def render_current_match_tab():
    """Renders the complete single-match UI with all original + new features."""
    remaining = render_sidebar()

    # Risk guard banner (shown at top of main area when any guard is active)
    render_risk_banner()

    _, center, _ = st.columns([1, 2, 1])
    with center:
        render_header()

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # Portfolio Mode badge
    pmode, pmc, pmdesc = get_portfolio_mode()

    # Match label
    cm = get_current_match()
    if cm:
        st.markdown(f'<div style="text-align:center;font-family:Rajdhani;font-size:14px;color:{MUTED};margin-bottom:4px;">{cm.get("label", "")}</div>', unsafe_allow_html=True)

    # Mode badges (portfolio mode + match mode)
    mode = st.session_state.mode
    phase = st.session_state.match_phase
    mc = AMBER if mode == "Pre-Match" else GREEN
    st.markdown(f"""
    <div style="text-align:center;margin-bottom:16px;">
        <span class="health-badge" style="background:{pmc}22;color:{pmc};border:1px solid {pmc}44;font-size:11px;padding:4px 14px;">{pmode}</span>
        <span style="background:{mc}22;color:{mc};padding:4px 14px;border-radius:20px;font-family:Orbitron,monospace;font-size:12px;font-weight:600;letter-spacing:2px;border:1px solid {mc}44;">
            {"\U0001f4cb PRE-MATCH" if mode == "Pre-Match" else "\U0001f534 LIVE"} \u00b7 {phase.upper()}
        </span>
    </div>
    """, unsafe_allow_html=True)

    # Settlement dialog
    render_settle_match_dialog()

    # Check if match is already settled
    if cm and cm.get("status") == "settled":
        rpnl = get_match_total_pnl(cm)
        rc = GREEN if rpnl >= 0 else RED
        result_labels = {
            "t1_win": f"{st.session_state.t1_name} Won",
            "t2_win": f"{st.session_state.t2_name} Won",
            "tie": "Tie", "no_result": "No Result", "abandoned": "Abandoned",
        }
        rl = result_labels.get(cm.get("result", ""), "Unknown")
        st.markdown(f"""
        <div class="dash-card" style="border-top:3px solid {rc};text-align:center;">
            <div style="font-family:Orbitron,monospace;font-size:14px;color:{MUTED};letter-spacing:2px;">MATCH SETTLED</div>
            <div style="font-family:Orbitron,monospace;font-size:28px;font-weight:800;color:{rc};margin:8px 0;">{fmt_inr(rpnl)}</div>
            <div style="font-family:Rajdhani;font-size:16px;color:{TEXT};">{rl}</div>
        </div>
        """, unsafe_allow_html=True)
        db = cm.get("debrief", "")
        if db:
            with st.expander("\U0001f4cb Match Debrief by Gemini"):
                st.markdown(db)
        # Unsettle / reopen match (undo settlement)
        with st.expander("\u21a9\ufe0f Undo Settlement (re-open this match)", expanded=False):
            st.markdown(f'<div style="font-family:Rajdhani;font-size:14px;color:{MUTED};">Made an error settling? This will re-open the match and remove the P&L from your portfolio.</div>', unsafe_allow_html=True)
            if st.button("\u21a9\ufe0f Re-open Match", key="unsettle_btn", type="primary", use_container_width=True):
                for m_rec in st.session_state.history["matches"]:
                    if m_rec["id"] == cm["id"]:
                        m_rec["status"] = "in_progress"
                        m_rec.pop("result", None)
                        m_rec.pop("realized_pnl", None)
                        m_rec.pop("closing_capital", None)
                        m_rec.pop("settled_at", None)
                        m_rec.pop("debrief", None)
                        break
                st.session_state.history["learnings"] = extract_learnings()
                save_history(st.session_state.history); update_user_index_stats()
                st.success("Match re-opened!"); time.sleep(0.5); st.rerun()
        return

    # Pre-existing position import
    render_pre_existing_panel()

    # Live odds
    render_odds_panel()
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

    # Odds shift summary
    render_edge_shift_summary()

    # Compute unified P&L
    all_bets = get_all_bets()
    pnl = compute_pnl(all_bets)
    odds = get_current_odds()
    pre_bets = st.session_state.pre_bets
    live_bets = st.session_state.bets
    learnings = st.session_state.history.get("learnings", {})

    # P&L metrics
    render_pnl_metrics(pnl)

    # EV summary
    render_ev_summary(pnl, odds)

    # Conviction mode panel
    render_conviction_panel(pnl, odds, remaining)

    # Cash-out value + hedge cost sensitivity (skip if PONR active — it has its own recs)
    render_cashout_and_hedge_projection(pnl, odds, remaining)

    # Recommendation (with portfolio learnings + cross-session overlay)
    rec = generate_recommendation(all_bets, pnl, odds, remaining, phase, mode, pre_bets, learnings)
    render_recommendation_card(rec, learnings)

    # Gemini insight card
    render_gemini_insight_card()

    # Celebration
    if all_bets and all(pnl[sc] > 0 for sc in pnl):
        if not st.session_state.celebration_fired:
            st.balloons()
            st.session_state.celebration_fired = True
    else:
        st.session_state.celebration_fired = False

    # Add live bet
    render_add_bet_form(remaining)

    # Add misc bet
    render_misc_bet_form(remaining)

    # Bet ledger
    render_bet_ledger(pnl)

    # Edit / Modify bets
    render_edit_bet_dialog()

    # P&L chart
    render_pnl_chart(pnl)

    # Strategy explanation
    with st.expander("\U0001f4d6 Strategy Explanation", expanded=False):
        text = build_strategy_text(all_bets, pnl, odds, remaining, phase, pre_bets, live_bets)
        st.markdown(text)

    # Auto-sync to history
    sync_match_to_history()

    # Footer
    st.markdown(
        f'<div style="text-align:center;padding:30px 0 10px 0;font-family:Rajdhani;'
        f'font-size:12px;color:{MUTED};letter-spacing:1px;">'
        f'IPL Hedge Engine v4 \u00b7 Multi-User \u00b7 Portfolio AI \u00b7 Not financial advice</div>',
        unsafe_allow_html=True,
    )


# ── Main ─────────────────────────────────────────────────────

def main():
    st.set_page_config(
        page_title="IPL Hedge Engine",
        page_icon="\U0001f3cf",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    inject_styles()
    init_state()

    # Gate 1: User selection (first thing shown)
    if not st.session_state.active_user:
        render_user_selection()
        return

    # Gate 2: Match active or lobby
    if st.session_state.current_match_id:
        render_current_match_tab()
    else:
        render_match_lobby()


if __name__ == "__main__":
    main()